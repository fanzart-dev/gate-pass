#!/usr/bin/env python3
"""Plain-script test suite for the gate pass app. Not pytest, run directly.

Builds its own temp SQLite database and temp storage directory per test
group, so it never touches storage/ (the real book).
"""

import functools
import io
import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import db  # noqa: E402
import app as appmod  # noqa: E402
import app as app_module  # noqa: E402
import invoice_parser  # noqa: E402
from app import create_app  # noqa: E402

sys.path.insert(0, str(Path(__file__).parent))
import make_invoice_pdf  # noqa: E402
import stock_transfer_parser  # noqa: E402

from werkzeug.datastructures import MultiDict  # noqa: E402

SAMPLE_INVOICE = Path(__file__).parent / "sample_invoices" / "golden_touch_sample.pdf"
SCANNED_INVOICE = Path(__file__).parent / "sample_invoices" / "mangaldeep_scan.pdf"
TWO_PAGE_INVOICE = Path(__file__).parent / "sample_invoices" / "golden_touch_two_page.pdf"
SERVICE_LINE_INVOICE = Path(__file__).parent / "sample_invoices" / "golden_touch_service_line.pdf"

# The fixtures are real supplier invoices and are deliberately NOT in git — they
# carry a supplier's GSTIN and named customers. Tests that need them are skipped
# rather than failed on a fresh clone, and everything else still runs. Put the
# real PDFs in tests/sample_invoices/ to run the parser tests in full.
SAMPLE_DIR = Path(__file__).parent / "sample_invoices"
BI_INVOICE = SAMPLE_DIR / "golden_touch_bi_series.pdf"
TRANSFER_MEMOS = tuple(SAMPLE_DIR / f"transfer_memo_doc{n}.pdf" for n in (63, 66, 69)) \
                 + tuple(SAMPLE_DIR / f"transfer_memo_{n}.pdf"
                         for n in ("wrapped_name", "short_names", "narrow_page"))
FIXTURES = (SAMPLE_INVOICE, SCANNED_INVOICE, TWO_PAGE_INVOICE, SERVICE_LINE_INVOICE,
            BI_INVOICE) + TRANSFER_MEMOS
HAVE_FIXTURES = all(p.exists() for p in FIXTURES)

RESULTS = []
SKIPPED = []


def needs_fixtures(fn):
    """Skip a test that reads one of the sample invoice PDFs, if they are absent."""
    @functools.wraps(fn)
    def wrapped(*args, **kwargs):
        if not HAVE_FIXTURES:
            SKIPPED.append(fn.__name__)
            return None
        return fn(*args, **kwargs)
    return wrapped


def check(name, condition):
    RESULTS.append((name, bool(condition)))
    if not condition:
        print(f"FAIL: {name}")


def raises(exc_type, fn, *args, **kwargs):
    try:
        fn(*args, **kwargs)
    except exc_type:
        return True
    except Exception as exc:  # wrong exception type
        print(f"  (raised {type(exc).__name__} instead of {exc_type.__name__}: {exc})")
        return False
    return False


TEST_USER = {"username": "ops", "display_name": "Ravi Kumar", "password": "gatepass-test-pw"}


def make_user(conn, username="ops", display_name="Ravi Kumar", password="gatepass-test-pw",
               status=None, is_admin=False):
    """Test accounts are approved by default — approval itself is tested directly."""
    return db.create_user(conn, username, display_name, password,
                           status=status or db.APPROVED, is_admin=is_admin)


def sign_in(client, username="ops", password="gatepass-test-pw"):
    return client.post("/login", data={"username": username, "password": password})


def logged_in_app(tmpdir, name, users=(("ops", "Ravi Kumar"),)):
    """A Flask test client already signed in as the first listed user."""
    storage = Path(tmpdir) / name
    flask_app = create_app(db_path=storage / "gate_pass.db", storage_dir=storage)
    flask_app.testing = True
    conn = db.connect(flask_app.config["DB_PATH"])
    for i, (username, display_name) in enumerate(users):
        db.create_user(conn, username, display_name, "gatepass-test-pw",
                        status=db.APPROVED, is_admin=(i == 0))
    conn.close()
    client = flask_app.test_client()
    sign_in(client, users[0][0])
    return flask_app, client


def sample_items(**overrides):
    item = {"item_name": "MICRON MODREN OAK", "quantity": "1", "cartons": "1"}
    item.update(overrides)
    return [item]


# ---------------------------------------------------------------------------

def test_serial_format():
    check("a serial reads like FZ-00001", db.serial_for("FZ", 1) == "FZ-00001")
    check("the sequence is padded to five digits", db.serial_for("FZ", 42) == "FZ-00042")
    check("a five digit sequence is not truncated", db.serial_for("FZ", 99999) == "FZ-99999")
    check("the sequence grows past five digits rather than wrapping",
          db.serial_for("FZ", 100000) == "FZ-100000")

    check("a prefix is upper-cased", db.validate_prefix(" fz ") == "FZ")
    check("digits are allowed in a prefix", db.validate_prefix("FZ27") == "FZ27")
    check("an empty prefix is refused", raises(ValueError, db.validate_prefix, ""))
    check("a prefix with a space is refused", raises(ValueError, db.validate_prefix, "FZ 27"))
    check("a prefix with a dash is refused", raises(ValueError, db.validate_prefix, "FZ-27"))
    check("an over-long prefix is refused",
          raises(ValueError, db.validate_prefix, "ABCDEFGHIJK"))


def test_triggers(tmpdir):
    conn = db.connect(Path(tmpdir) / "triggers.db")
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                              "FR 262702171", "03-08-2026", "KA 01 AB 1234", sample_items())

    check("serial_no cannot be updated directly", raises(
        sqlite3.IntegrityError, conn.execute,
        "UPDATE gate_passes SET serial_no = 'HACKED' WHERE id = ?", (gp["id"],)))
    check("serial_seq cannot be updated directly", raises(
        sqlite3.IntegrityError, conn.execute,
        "UPDATE gate_passes SET serial_seq = 999 WHERE id = ?", (gp["id"],)))
    check("serial_scope cannot be updated directly", raises(
        sqlite3.IntegrityError, conn.execute,
        "UPDATE gate_passes SET serial_scope = '99-00' WHERE id = ?", (gp["id"],)))
    check("gate_passes rows cannot be deleted", raises(
        sqlite3.IntegrityError, conn.execute,
        "DELETE FROM gate_passes WHERE id = ?", (gp["id"],)))

    log_row = conn.execute(
        "SELECT id FROM audit_log WHERE gate_pass_id = ?", (gp["id"],)
    ).fetchone()
    check("issuing wrote an audit log entry", log_row is not None)
    check("audit_log rows cannot be updated", raises(
        sqlite3.IntegrityError, conn.execute,
        "UPDATE audit_log SET action = 'tampered' WHERE id = ?", (log_row["id"],)))
    check("audit_log rows cannot be deleted", raises(
        sqlite3.IntegrityError, conn.execute,
        "DELETE FROM audit_log WHERE id = ?", (log_row["id"],)))

    # Sanity: the trigger doesn't over-block ordinary columns.
    conn.execute("UPDATE gate_passes SET vehicle_no = 'KA 05 ZZ 9999' WHERE id = ?", (gp["id"],))
    conn.commit()
    row = conn.execute("SELECT vehicle_no FROM gate_passes WHERE id = ?", (gp["id"],)).fetchone()
    check("non-serial columns remain editable", row["vehicle_no"] == "KA 05 ZZ 9999")

    conn.close()


def test_serial_allocation(tmpdir):
    conn = db.connect(Path(tmpdir) / "serials.db")

    gp1 = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                               "INV-1", "03-08-2026", "", sample_items())
    gp2 = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                               "INV-2", "03-08-2026", "", sample_items())

    check("the first serial of a run is 00001", gp1["serial_no"] == "FZ-00001")
    check("the second serial follows on", gp2["serial_no"] == "FZ-00002")
    check("the serial carries the current prefix", gp1["serial_no"].startswith("FZ-"))
    check("serials in the same run share a prefix",
          gp1["serial_scope"] == gp2["serial_scope"] == "FZ")

    check("issuing with no items raises ValueError", raises(
        ValueError, db.create_gate_pass, conn, None, "S", "C", "I", "01-01-2026", "", []))
    check("issuing with a blank quantity raises ValueError", raises(
        ValueError, db.create_gate_pass, conn, None, "S", "C", "I", "01-01-2026", "",
        [{"item_name": "Fan", "quantity": "", "cartons": "1"}]))
    blank_cartons = db.create_gate_pass(
        conn, None, "S", "C", "I", "01-01-2026", "",
        [{"item_name": "Fan", "quantity": "1", "cartons": ""}])
    check("a gate pass issues with the carton count left blank",
          blank_cartons["serial_no"] == "FZ-00003")
    check("the blank carton count is stored as empty",
          blank_cartons["items"][0]["cartons"] == "")
    check("an item with no cartons key at all still issues",
          db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                               [{"item_name": "Fan", "quantity": "1"}]) is not None)
    check("issuing with an unnamed item raises ValueError", raises(
        ValueError, db.create_gate_pass, conn, None, "S", "C", "I", "01-01-2026", "",
        [{"item_name": "  ", "quantity": "1", "cartons": "1"}]))

    # A long invoice is NOT refused and is NOT given a second number: it stays
    # one gate pass, printed over more pages.
    long_invoice = [{"item_name": f"Item {i}", "quantity": "1", "cartons": "1"}
                    for i in range(db.ITEMS_PER_PAGE + 6)]
    gp_long = db.create_gate_pass(conn, None, "S", "C", "LONG", "01-01-2026", "",
                                   long_invoice)
    check("an invoice longer than one page still issues",
          len(gp_long["items"]) == db.ITEMS_PER_PAGE + 6)
    check("it takes exactly one serial number, with no suffix",
          gp_long["serial_no"].count("-") == 1)

    exactly_full = long_invoice[:db.ITEMS_PER_PAGE]
    gp_full = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "", exactly_full)
    check("a full sheet of items issues fine",
          len(gp_full["items"]) == db.ITEMS_PER_PAGE)

    # Six passes issued so far (2 + blank cartons + no cartons key + the long
    # invoice + a full sheet); every rejected attempt in between must have left
    # the numbering alone.
    gp3 = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                               sample_items(item_name="Second Model"))
    check("failed issue attempts did not burn a number", gp3["serial_no"] == "FZ-00007")

    conn.close()


def test_draft_never_burns_a_number(tmpdir):
    conn = db.connect(Path(tmpdir) / "drafts.db")

    for _ in range(5):
        draft_id = db.create_draft(conn, supplier_name="S", customer_name="C",
                                    invoice_no="I", invoice_date="01-01-2026",
                                    items=[{"item_name": "Fan", "quantity": "1", "cartons": ""}])
        db.delete_draft(conn, draft_id)

    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "", sample_items())
    check("five abandoned drafts did not advance the counter", gp["serial_no"] == "FZ-00001")

    check("the number came from the book itself, not a separate counter",
          conn.execute(
              "SELECT COUNT(*) AS n FROM sqlite_master WHERE name = 'serial_counters'"
          ).fetchone()["n"] == 0)

    conn.close()


def test_draft_lifecycle(tmpdir):
    conn = db.connect(Path(tmpdir) / "draft_lifecycle.db")

    draft_id = db.create_draft(
        conn, supplier_name="Golden Touch Exports", customer_name="FANZART LLP",
        invoice_no="FR 262702171", invoice_date="03-08-2026",
        items=[{"sl_no": 1, "item_name": "MICRON MODREN OAK", "quantity": "1", "cartons": ""}],
    )
    draft = db.get_draft(conn, draft_id)
    check("draft round-trips supplier name", draft["supplier_name"] == "Golden Touch Exports")
    check("draft round-trips one item", len(draft["items"]) == 1)

    db.update_draft(conn, draft_id, "Golden Touch Exports", "FANZART LLP", "FR 262702171",
                     "03-08-2026", "KA 01 AB 1234",
                     [{"item_name": "MICRON MODREN OAK", "quantity": "1", "cartons": "2"},
                      {"item_name": "Second Item", "quantity": "3", "cartons": "3"}])
    draft = db.get_draft(conn, draft_id)
    check("update_draft replaces items rather than appending", len(draft["items"]) == 2)
    check("update_draft stores vehicle number", draft["vehicle_no"] == "KA 01 AB 1234")

    check("list_drafts finds the draft", any(d["id"] == draft_id for d in db.list_drafts(conn)))

    db.delete_draft(conn, draft_id)
    check("draft is gone after delete", db.get_draft(conn, draft_id) is None)
    check("draft items are gone after delete",
          conn.execute("SELECT COUNT(*) AS n FROM draft_items WHERE draft_id = ?",
                       (draft_id,)).fetchone()["n"] == 0)

    conn.close()


def test_cancel(tmpdir):
    conn = db.connect(Path(tmpdir) / "cancel.db")
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "", sample_items())

    db.cancel_gate_pass(conn, gp["id"], "damaged goods")
    cancelled = db.get_gate_pass(conn, gp["id"])
    check("cancelled gate pass keeps its serial number", cancelled["serial_no"] == gp["serial_no"])
    check("cancelled gate pass has status cancelled", cancelled["status"] == "cancelled")
    check("cancel reason is recorded", cancelled["cancel_reason"] == "damaged goods")

    only_cancelled = db.list_gate_passes(conn, status="cancelled")
    only_issued = db.list_gate_passes(conn, status="issued")
    check("cancelled filter finds it", any(g["id"] == gp["id"] for g in only_cancelled))
    check("issued filter excludes it", all(g["id"] != gp["id"] for g in only_issued))

    conn.close()


def test_settings(tmpdir):
    conn = db.connect(Path(tmpdir) / "settings.db")

    defaults = db.get_settings(conn)
    check("default paper mode is a4x2", defaults["paper_mode"] == "a4x2")
    check("both totals are off by default, and are separate settings",
          defaults["show_total_qty"] == "0" and defaults["show_total_cartons"] == "0")
    check("the old combined setting is gone", "show_totals" not in defaults)
    # No show_vehicle setting: the vehicle prints when one was recorded.
    check("there is no vehicle tick box to get wrong",
          "show_vehicle" not in defaults)

    db.update_settings(conn, paper_mode="a5", show_total_qty="1")
    updated = db.get_settings(conn)
    check("paper_mode persists", updated["paper_mode"] == "a5")
    check("show_total_qty persists", updated["show_total_qty"] == "1")
    check("and the other one is untouched by it", updated["show_total_cartons"] == "0")
    check("unrelated setting keeps its default", updated["company_name"] == "fanzart")

    conn.close()


@needs_fixtures
def test_invoice_parser():
    result = invoice_parser.parse_invoice(SAMPLE_INVOICE)
    check("parser reads supplier name", result["supplier_name"] == "Golden Touch Exports")
    check("parser reads customer name without M/S prefix", result["customer_name"] == "FANZART LLP")
    check("parser reads invoice number", result["invoice_no"] == "FR 262702171")
    check("parser reads invoice date", result["invoice_date"] == "03-08-2026")
    check("parser reads exactly one item", len(result["items"]) == 1)
    check("parser reads item name from the Model column",
          result["items"] and result["items"][0]["item_name"] == "MICRON MODREN OAK")
    check("parser reads quantity", result["items"] and result["items"][0]["quantity"] == "1")
    check("no parse notes on a clean read", result["notes"] == [])

    garbage = invoice_parser.parse_invoice(io.BytesIO(b"not a pdf at all"))
    check("garbage input degrades instead of raising", garbage["supplier_name"] == "")
    check("garbage input reports a note", len(garbage["notes"]) > 0)
    check("garbage input yields no items", garbage["items"] == [])


def test_multi_item_and_wrapped_names(tmpdir):
    """The MANGALDEEP invoice: 3 items, every name wrapping onto a second line.

    The item box has no rules between rows, so extract_tables() collapses it
    into one row and loses the name-to-quantity mapping. The parser reads word
    coordinates instead, and must pair each quantity with its full name.

    This builds a real PDF and runs the whole parser over it, so pdfplumber's
    own word and edge extraction is exercised rather than a hand-made stand-in.
    """
    pdf_path = Path(tmpdir) / "mangaldeep.pdf"
    make_invoice_pdf.build_invoice(
        pdf_path, supplier="Golden Touch Exports", customer="MANGALDEEP",
        invoice_no="FR 262702169", invoice_date="03-08-2026",
        items=make_invoice_pdf.MANGALDEEP_ITEMS,
    )
    parsed = invoice_parser.parse_invoice(pdf_path)
    items = parsed["items"]

    check("multi-item invoice parses without notes", parsed["notes"] == [])
    check("reads the supplier as From", parsed["supplier_name"] == "Golden Touch Exports")
    check("reads the Bill To party as Customer", parsed["customer_name"] == "MANGALDEEP")
    check("reads the invoice number", parsed["invoice_no"] == "FR 262702169")
    check("reads the invoice date", parsed["invoice_date"] == "03-08-2026")
    check("reads all three items", len(items) == 3)
    if len(items) == 3:
        check("first wrapped name is rejoined",
              items[0]["item_name"] == "VIENNA (52) (DC) (ABS) MATTE WHITE")
        check("second wrapped name is rejoined",
              items[1]["item_name"] == "VIENNA (52) (DC) (ABS) BRUSH NICKEL")
        check("third wrapped name is rejoined",
              items[2]["item_name"] == "VIENNA (52) (DC) (ABS) ANTIQUE BRASS")
        check("quantity stays with its own item",
              [i["quantity"] for i in items] == ["2", "1", "1"])
        check("items are numbered in order",
              [i["sl_no"] for i in items] == [1, 2, 3])
        check("total quantity matches the transporter's 4 fans",
              sum(int(i["quantity"]) for i in items) == 4)
        check("the HSN code is never mistaken for a quantity",
              all("84145120" not in i["quantity"] for i in items))
        check("the unit price is never mistaken for a quantity",
              all("." not in i["quantity"] for i in items))
        check("the item name excludes the description column",
              all("Ceiling" not in i["item_name"] for i in items))
    check("the footer line is not read as an item",
          all("Forty" not in i["item_name"] for i in items))


@needs_fixtures
def test_two_page_invoice():
    """A long invoice continues its item table onto a second page, under a
    repeated letterhead. All 13 items must come through, numbered continuously,
    and the terms and conditions that follow the table must not be read as an
    item — this invoice prints no "Total Amount in words" line on page one, so
    the end of the item box has to come from the rule drawn across it."""
    result = invoice_parser.parse_invoice(TWO_PAGE_INVOICE)
    items = result["items"]

    check("two-page invoice parses without notes", result["notes"] == [])
    check("reads the invoice number once", result["invoice_no"] == "FR 262702140")
    check("reads items from both pages", len(items) == 13)
    check("items are numbered continuously across pages",
          [i["sl_no"] for i in items] == list(range(1, 14)))
    check("the first item is from page one",
          items and items[0]["item_name"] == "PHOENIX 52 WALNUT")
    check("the last item is from page two",
          items and items[-1]["item_name"] == "REMOTE METALLICA")
    check("quantities survive the page break",
          [i["quantity"] for i in items] ==
          ["4", "8", "1", "1", "1", "1", "1", "2", "1", "1", "2", "1", "1"])
    check("total quantity is 25",
          sum(int(i["quantity"]) for i in items) == 25)
    check("terms and conditions are not swallowed into an item name",
          all("Goods once sold" not in i["item_name"] for i in items))
    check("bank details are not swallowed into an item name",
          all("Account Name" not in i["item_name"] for i in items))
    check("no item name runs longer than a real model name",
          all(len(i["item_name"]) < 60 for i in items))
    check("the wrapped Description text never lands in the item name",
          all("Spares" not in i["item_name"] for i in items))
    # Rule 5: cartons are typed by hand. The parser must not offer a value for
    # them at all, not even one copied from the quantity.
    check("the parser never returns a carton count",
          all("cartons" not in i for i in items))


def test_name_column_is_found_by_its_heading():
    """The item name is whichever column is *headed* as the name.

    No PDF needed: this is the column-picking logic on its own, so it runs on a
    clone that has no sample invoices. It is the part that actually broke.
    """
    def header(labels, width=100):
        """Fake a heading row: one label per column, evenly spaced."""
        words, bounds = [], [i * width for i in range(len(labels) + 1)]
        for i, label in enumerate(labels):
            x = i * width + 5
            for part in label.split():
                words.append({"text": part, "x0": x, "x1": x + 20, "top": 10, "bottom": 18})
                x += 25
        return words, bounds

    # The layout that was being misread: Model No sits where Model used to.
    words, bounds = header(["S.no", "Model No", "Model Name", "HSN", "Description", "Qty"])
    check("'Model Name' is picked over 'Model No'",
          invoice_parser._name_column_index(words, bounds, words[0]) == 2)

    # The original layout must still resolve to the same column as before.
    words, bounds = header(["S.no", "Model", "HSN", "Description", "Qty"])
    check("a plain 'Model' column is still the name",
          invoice_parser._name_column_index(words, bounds, words[0]) == 1)

    words, bounds = header(["Sl", "Item Name", "Qty"])
    check("'Item Name' is recognised too",
          invoice_parser._name_column_index(words, bounds, words[0]) == 1)

    words, bounds = header(["S.no", "Item Code", "Item", "Qty"])
    check("a code column is never mistaken for the name",
          invoice_parser._name_column_index(words, bounds, words[0]) == 2)

    # A row with a quantity but no name would reach the operator blank, and a
    # pass cannot be issued without an item name. The code is a poor label but
    # one the warehouse can match against the invoice.
    bounds = [0, 100, 200, 300, 400, 500]
    row = [
        {"text": "1", "x0": 5, "x1": 20, "top": 10, "bottom": 18},          # Sl No.
        {"text": "0007", "x0": 105, "x1": 140, "top": 10, "bottom": 18},    # Model No
        # the Model Name cell is deliberately empty
        {"text": "2", "x0": 405, "x1": 420, "top": 10, "bottom": 18},       # Qty
    ]
    fell_back = invoice_parser._assemble_items(row, bounds, name_idx=2, code_idx=1)
    check("an empty name falls back to the model code",
          fell_back and fell_back[0]["item_name"] == "0007")
    check("and the quantity still comes through",
          fell_back and fell_back[0]["quantity"] == "2")
    # The fallback must not fire when there is a name; that would undo the fix.
    row.append({"text": "BUDDY", "x0": 205, "x1": 250, "top": 10, "bottom": 18})
    named = invoice_parser._assemble_items(row, bounds, name_idx=2, code_idx=1)
    check("the code is never used when a name is present",
          named and named[0]["item_name"] == "BUDDY")


def test_stock_transfer_memo(tmpdir):
    """A STOCK TRANSFER MEMO is a different document, read by its own module.

    Goods moving between Golden Touch branches still leave the gate, so they
    still need a pass — but the memo has a TO number rather than an invoice
    number, and its destination is a branch rather than a customer. Recording
    just "Golden Touch Exports" would make every transfer look identical in the
    register.

    The memo is generated to the real WEBPOS layout (see build_stock_transfer),
    so this runs without any sample document. What makes that layout awkward:
    there are NO drawn rules — the separators are runs of hyphens — so columns
    are held by alignment alone, and the two addresses are printed side by side.
    """
    path = Path(tmpdir) / "transfer.pdf"
    make_invoice_pdf.build_stock_transfer(path, items=make_invoice_pdf.TRANSFER_ITEMS)
    result = invoice_parser.parse_invoice(path)

    check("a transfer memo parses without notes", result["notes"] == [])
    check("the TO number replaces the invoice number, label and all",
          result["invoice_no"] == "TO NO: FR 262700512")
    check("the destination branch is recorded, not just the company",
          result["customer_name"] == "Golden Touch Exports, Indospace")
    check("the transfer date is read", result["invoice_date"] == "01-08-2026")
    check("the supplier is the company", result["supplier_name"] == "Golden Touch Exports")
    check("all three lines come through", len(result["items"]) == 3)

    # The heading "Model Name" is 54pt wide over names up to 140pt. Splitting
    # the gap between headings put the last word of the longest name into the
    # Description column, so it came through as "GRANDMASTER MATTE BLACK 70" —
    # a product name quietly missing its last word.
    check("a name wider than its heading keeps every word",
          result["items"] and result["items"][0]["item_name"]
          == "GRANDMASTER MATTE BLACK 70 INCHES")
    check("model names, not HSN codes",
          all(not i["item_name"].strip().isdigit() for i in result["items"]))
    check("quantities come through",
          [i["quantity"] for i in result["items"]] == ["1", "2", "5"])
    # The TOTAL row is not an item, and neither are the hyphen separators.
    check("the TOTAL row is not read as an item",
          all("total" not in i["item_name"].lower() for i in result["items"]))
    check("no separator rule is read as an item",
          all("---" not in i["item_name"] for i in result["items"]))
    # Cartons are never parsed from any document — rule 5, no exceptions.
    check("no line carries a cartons value",
          all("cartons" not in i for i in result["items"]))

    # Both addresses are printed side by side and BOTH are Golden Touch. The
    # sender is in Bommasandra, so a keyword search over the flattened text is
    # reading the right answer out of the wrong column.
    for keyword, address in (
            ("Indiranagar", ("61, Defence Colony , 100 ft Road", "Indiranagar.",
                              "Bengaluru 560038")),
            ("Sadashivanagar", ("No 5, 1st Floor, opp. White Petals,",
                                 "Bellary Road, Sadashivanagar,", "Bengaluru 560080"))):
        p = Path(tmpdir) / f"transfer_{keyword}.pdf"
        make_invoice_pdf.build_stock_transfer(
            p, to_address=address,
            from_address=("15 Krishnanagar Ind Area,", "Indospace Bommasandra,",
                           "Bangalore 560029"),
            items=make_invoice_pdf.TRANSFER_ITEMS)
        r = invoice_parser.parse_invoice(p)
        check(f"{keyword} maps to its branch",
              r["customer_name"] == f"Golden Touch Exports, {keyword}")
        check(f"the sender's branch does not win over {keyword}",
              "Indospace" not in r["customer_name"])

    # An unrecognised branch is not a failure — a new one may have opened.
    p = Path(tmpdir) / "transfer_unknown.pdf"
    make_invoice_pdf.build_stock_transfer(
        p, to_address=("New Estate, Whitefield", "Bengaluru 560066"),
        from_address=("15 Krishnanagar Ind Area,", "Bangalore 560029"),
        items=make_invoice_pdf.TRANSFER_ITEMS)
    r = invoice_parser.parse_invoice(p)
    check("an unknown branch still yields the company name",
          r["customer_name"] == "Golden Touch Exports")
    check("and says the branch needs adding by hand",
          any("which branch" in n for n in r["notes"]))
    check("an unknown branch does not block the items", len(r["items"]) == 3)

    # Routing: neither document type may be read by the other's rules.
    check("a memo is detected by either marker",
          stock_transfer_parser.looks_like_stock_transfer("STOCK TRANSFER MEMO")
          and stock_transfer_parser.looks_like_stock_transfer("Transfer Out"))
    check("a tax invoice is never routed to the transfer parser",
          not stock_transfer_parser.looks_like_stock_transfer(
              "Tax Invoice cum Delivery Note ... Bill To ... Name : M/S FANZART LLP"))


@needs_fixtures
def test_transfer_rows_are_read_by_shape_not_by_column_edges():
    """The row parser, without a PDF, so a fresh clone still guards it.

    The fixture PDFs are real documents carrying a GSTIN and are not in git.
    The logic they broke is small enough to exercise directly, with word boxes
    laid out at the coordinates the real documents actually use.
    """
    def w(text, x0, x1):
        return {"text": text, "x0": x0, "x1": x1, "top": 0, "bottom": 10}

    # The layout that broke it: description at x=232 and quantity at x=477,
    # both on the wrong side of boundaries derived from the heading text.
    row = [w("1", 21.3, 25.5), w("84145120", 58.5, 92.1),
           w("MUSTANG", 113.8, 143.2), w("WHITE", 147.4, 168.4),
           w("Ceiling", 230.9, 260.3), w("Fan", 264.5, 277.1),
           w("1", 477.6, 481.8), w("22872.87", 524.7, 558.3)]
    got = stock_transfer_parser._read_item_row(row)
    check("a row is read whatever the columns line up with", got is not None)
    check("the name stops before the description",
          got and got["item_name"] == "MUSTANG WHITE")
    check("the quantity is the integer before the money, not the money",
          got and got["quantity"] == "1")

    # The long name that crowds its description: only 19.2pt between them,
    # against 4.2pt between words inside the name.
    long_row = [w("1", 21, 25), w("84145930", 58, 92),
                w("GRANDMASTER", 113, 175), w("MATTE", 179.2, 208),
                w("WHITE", 212.2, 240), w("60", 244.2, 254),
                w("INCHES", 258.2, 290), w("Industrial", 309.2, 350),
                w("Fan", 354.2, 370), w("1", 480, 484), w("50838.98", 520, 560)]
    long_got = stock_transfer_parser._read_item_row(long_row)
    check("a long name keeps every one of its words",
          long_got and long_got["item_name"] == "GRANDMASTER MATTE WHITE 60 INCHES")
    check("and still drops the description beside it",
          long_got and "Industrial" not in long_got["item_name"])

    # A name with no description at all must survive untouched.
    bare = [w("2", 21, 25), w("84145120", 58, 92),
            w("DRIFT", 113, 140), w("DARK", 144.2, 172), w("COFFEE", 176.2, 215),
            w("5", 480, 484), w("1234.50", 520, 560)]
    check("a row with no description keeps the whole name",
          stock_transfer_parser._read_item_row(bare)["item_name"] == "DRIFT DARK COFFEE")

    # Things that are not item rows.
    check("a total row is not an item",
          stock_transfer_parser._read_item_row(
              [w("TOTAL", 113, 150), w("5", 480, 484), w("999.00", 520, 560)]) is None)
    check("a line with no money value is not an item",
          stock_transfer_parser._read_item_row(
              [w("1", 21, 25), w("84145120", 58, 92), w("NAME", 113, 150),
               w("ONLY", 154, 180), w("HERE", 184, 210)]) is None)

    # A name containing digits must not be mistaken for the money column, which
    # is why the value is found from the right.
    digits = [w("1", 21, 25), w("84145120", 58, 92),
              w("HUGGER", 113, 150), w("52", 154.2, 166),
              w("1", 480, 484), w("2500.00", 520, 560)]
    check("digits inside a name do not end the row early",
          stock_transfer_parser._read_item_row(digits)["item_name"] == "HUGGER 52")


@needs_fixtures
def test_real_transfer_memos():
    """The three real WEBPOS memos, one per branch.

    Each document states its own TOTAL quantity, which is checked against the
    sum of the lines read — an independent confirmation that nothing was
    dropped or duplicated, from the document itself rather than from this test.
    """
    expected = {
        "transfer_memo_doc69.pdf": ("TO NO: FR 262700512", "Indospace",
                                     "01-08-2026", 2, 10),
        "transfer_memo_doc66.pdf": ("TO NO: FR 262700531", "Indiranagar",
                                     "04-08-2026", 4, 9),
        "transfer_memo_doc63.pdf": ("TO NO: FR 262700536", "Sadashivanagar",
                                     "06-08-2026", 1, 1),
        # The three that arrived later and broke the column-geometry parse.
        # Their description column starts at x=231 where the earlier documents
        # put it at 272, and one of them is on a 607pt page rather than 612.
        "transfer_memo_wrapped_name.pdf": ("TO NO: FR 262700652", "Indiranagar",
                                            "03-09-2026", 5, 5),
        "transfer_memo_short_names.pdf": ("TO NO: FR 262700644", "Indiranagar",
                                           "01-09-2026", 3, 3),
        "transfer_memo_narrow_page.pdf": ("TO NO: FR 262700648", "Sadashivanagar",
                                           "02-09-2026", 1, 1),
    }
    for name, (to_no, branch, date, count, total_qty) in expected.items():
        result = invoice_parser.parse_invoice(SAMPLE_DIR / name)
        check(f"{name} parses without notes", result["notes"] == [])
        check(f"{name} reads its TO number", result["invoice_no"] == to_no)
        check(f"{name} goes to {branch}",
              result["customer_name"] == f"Golden Touch Exports, {branch}")
        check(f"{name} reads its date", result["invoice_date"] == date)
        check(f"{name} reads all {count} line(s)", len(result["items"]) == count)
        check(f"{name} totals match the document's own TOTAL row",
              sum(int(i["quantity"]) for i in result["items"]) == total_qty)
        check(f"{name} has no empty item name",
              all(i["item_name"].strip() for i in result["items"]))

    # The description column must never end up inside the name. It did: the
    # boundary derived from the `Description` HEADING sat at x=258 while the
    # word `Ceiling` underneath it began at x=232, so every wrapped name came
    # through as "AVALON - MATTE Ceiling SILVER".
    wrapped = invoice_parser.parse_invoice(SAMPLE_DIR / "transfer_memo_wrapped_name.pdf")
    names = [i["item_name"] for i in wrapped["items"]]
    check("a name that wraps is joined with nothing in between",
          names[0] == "AVALON - MATTE SILVER")
    check("and no name carries the description column",
          not any("Ceiling" in n or "Fan" in n for n in names))
    check("every line is read, including the ones that wrap", len(names) == 5)

    short = invoice_parser.parse_invoice(SAMPLE_DIR / "transfer_memo_short_names.pdf")
    check("a hyphenated name keeps both halves",
          [i["item_name"] for i in short["items"]][0] == "HUGGER 52 - MUD BROWN")

    # This one returned nothing at all. Its quantity sits at x=477 where the
    # others put it at x=445 — across a boundary at 469 — so the quantity was
    # read as part of the value, the row failed its own check, and the table
    # came back empty with "could not find an item table".
    narrow = invoice_parser.parse_invoice(SAMPLE_DIR / "transfer_memo_narrow_page.pdf")
    check("a narrower page still finds its item table", len(narrow["items"]) == 1)
    check("and reads the quantity out of it rather than the price",
          narrow["items"][0]["quantity"] == "1")
    check("and the name is only the model name",
          narrow["items"][0]["item_name"] == "MUSTANG WHITE")

    # The longest real name, which is what the column boundaries have to clear.
    doc66 = invoice_parser.parse_invoice(SAMPLE_DIR / "transfer_memo_doc66.pdf")
    check("the longest real name keeps its last word",
          doc66["items"][0]["item_name"] == "GRANDMASTER MATTE BLACK 70 INCHES")


def test_a_gate_pass_uploaded_as_an_invoice(tmpdir):
    """Uploading one of our own printed passes must say so.

    It happens: the printed pass and the supplier invoice sit in the same
    folder. Left alone the parser produces confident nonsense rather than
    failing — the two A5 halves of an A4 sheet read as one line, so the invoice
    number comes back doubled ("FR 262702140 Date : 31-07-2026 Invoice No : FR
    262702140 ..."), the supplier reads as "Sl. No. : FZ-00057 Sl. No. :
    FZ-00057" and the item count doubles. The operator was then told only
    "no customer — could not be read from the PDF", which points at the wrong
    problem entirely.
    """
    flask_app, client = logged_in_app(tmpdir, "gatepassupload")
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                              "FR 262702140", "31-07-2026", "", sample_items(),
                              prepared_by="Ravi Kumar")
    printed = client.get(f"/print/{gp['id']}").get_data(as_text=True)
    conn.close()

    # The real thing is a PDF, but detection reads the page text, and that is
    # what the printed HTML carries.
    text = re.sub(r"<[^>]+>", " ", printed)
    check("a printed pass is recognised as one",
          invoice_parser._looks_like_a_gate_pass(text))

    # Both markers are required, so an invoice mentioning a gate pass in its
    # terms and conditions is not turned away.
    check("'gate pass' alone is not enough",
          not invoice_parser._looks_like_a_gate_pass(
              "Tax Invoice ... a gate pass must accompany the goods"))
    check("a signature block alone is not enough",
          not invoice_parser._looks_like_a_gate_pass("Delivery note ... Authorised by"))

    # The draft must report the real reason, not the first failed field check.
    draft = {"supplier_name": "", "customer_name": "", "invoice_no": "",
             "invoice_date": "", "items": [],
             "parse_notes": invoice_parser.GATE_PASS_UPLOADED_NOTE}
    check("the draft says a gate pass was uploaded, not 'no supplier'",
          db.draft_problem(draft) == invoice_parser.GATE_PASS_UPLOADED_NOTE)

    # And a genuine invoice with a missing field still reports that field.
    ordinary = {"supplier_name": "", "customer_name": "", "invoice_no": "",
                "invoice_date": "", "items": [], "parse_notes": ""}
    check("an ordinary unreadable invoice still names the missing field",
          db.draft_problem(ordinary) == "no supplier — could not be read from the PDF")


@needs_fixtures
def test_bi_series_invoice():
    """A BI-series invoice has two extra columns before the name.

    Its layout is  S.no | Model No | Model Name | HSN | Description | Batch_no |
    Qty | ...  where the older FR invoices are  S.no | Model | HSN | ... .
    Assuming the name was always the column after S.no put the model *number*
    on the gate pass: items came through as '0003 B' and '0006' instead of
    'HAWK BLACK' and 'BUDDY'.
    """
    result = invoice_parser.parse_invoice(BI_INVOICE)
    items = result["items"]

    check("BI invoice parses without notes", result["notes"] == [])
    check("reads the BI invoice number", result["invoice_no"].startswith("BI "))
    check("reads both items", len(items) == 2)
    check("reads the model NAME, not the model number",
          items and items[0]["item_name"] == "HAWK BLACK")
    check("second item name is right too",
          len(items) > 1 and items[1]["item_name"] == "BUDDY")
    check("no item name is a bare model number",
          all(not re.fullmatch(r"[\d ]+[A-Z]?", i["item_name"]) for i in items))
    check("quantities still read correctly",
          [i["quantity"] for i in items] == ["1", "1"])


@needs_fixtures
def test_service_line_invoice():
    """An invoice mixing fans, rods and a service line. Everything is offered to
    the operator; deciding what physically leaves the gate is their call."""
    result = invoice_parser.parse_invoice(SERVICE_LINE_INVOICE)
    items = result["items"]

    check("service-line invoice parses without notes", result["notes"] == [])
    # Asserted by shape, not by name: the point is that a customer who is a
    # person rather than Fanzart itself is read correctly, and the fixture is a
    # real invoice, so spelling the buyer's name out here would publish it.
    check("reads a non-Fanzart customer",
          result["customer_name"].startswith("MR ") and len(result["customer_name"]) > 5)
    check("reads the invoice number", result["invoice_no"] == "FR 262702176")
    check("reads all three lines", len(items) == 3)
    check("reads a two-digit quantity",
          items and items[1]["quantity"] == "18")
    check("keeps an item name containing brackets and underscores",
          items and items[1]["item_name"] == "FAN ROD19MM_(IN INCHES)")
    check("keeps the service line for the operator to remove",
          items and "ERECTION" in items[2]["item_name"])
    check("no item is left without a quantity",
          all(i["quantity"] for i in items))


@needs_fixtures
def test_cartons_come_from_the_master_list(tmpdir):
    """Carton counts are looked up, not counted by hand — and never guessed.

    This replaces the old rule that nothing may ever pre-fill the carton box.
    That rule existed because the code had no way to know how a fan is packed;
    the Box Qty master list is exactly that knowledge, so the rule now is
    narrower: fill it in when the master list actually says, and leave it
    visibly empty when it does not. An empty box asks a question. A wrong
    number is an assertion, on a document somebody signs at the gate.
    """
    flask_app, client = logged_in_app(tmpdir, "cartons_storage")
    conn = db.connect(flask_app.config["DB_PATH"])
    db.upsert_carton_mappings(conn, [
        ("MICRON MODREN OAK", 1), ("PHOENIX 52 WALNUT", 1),
        ("NIZAM ANTIQUE BRASS", 1), ("CHAKRA 28", 1),
        ("WINDFLOWER - FANDELIER", 1), ("AVALON - MATTE SILVER", 1),
        ("RACE (AC) MATTE WHITE", 1), ("CRYSTAL - FANDELIER", 2),
    ])

    # No parser invents a carton count; every value comes from the master list.
    for fixture in (SAMPLE_INVOICE, TWO_PAGE_INVOICE, SERVICE_LINE_INVOICE, SCANNED_INVOICE):
        parsed = invoice_parser.parse_invoice(fixture)
        check(f"{fixture.name}: the parser itself returns no carton value",
              all(not item.get("cartons") for item in parsed["items"]))

    with open(TWO_PAGE_INVOICE, "rb") as f:
        resp = client.post("/upload", data={"invoice": (f, TWO_PAGE_INVOICE.name)},
                            content_type="multipart/form-data")
    draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    review = client.get(f"/review/{draft_id}").data
    draft = db.get_draft(conn, draft_id)
    by_name = {i["item_name"]: i for i in draft["items"]}

    # The master sheet counts cartons per UNIT, so eight fans is eight boxes.
    # Taking the sheet value literally would print 1 against 8 fans, and
    # understating boxes is how a box goes missing without anyone noticing.
    check("a line for 8 fans gets 8 cartons, not the per-unit 1",
          by_name["MICRON MODREN OAK"]["cartons"] == "8")
    check("a line for 4 fans gets 4", by_name["PHOENIX 52 WALNUT"]["cartons"] == "4")
    check("a line for 1 fan gets 1", by_name["CHAKRA 28"]["cartons"] == "1")

    # Remotes are not in the master list, so nobody has said how they pack.
    check("an item missing from the master list is left blank",
          by_name["REMOTE VENETIAN"]["cartons"] == "")
    check("even when several of it were bought",
          by_name["REMOTE AVALON"]["cartons"] == "")

    check("the review page shows the looked-up values",
          b'name="cartons" value="8"' in review)
    check("and shows empty boxes for the unmatched ones",
          b'name="cartons" value=""' in review)
    check("every carton box stays editable",
          b"readonly" not in review.split(b'name="cartons"')[1][:80]
          and b"disabled" not in review.split(b'name="cartons"')[1][:80])
    check("no control offers to derive cartons from quantity",
          b"Cartons = Quantity" not in review and b"cartons-equal-qty" not in review)
    check("the cartons column is still there to type into",
          b"No. of Cartons" in review)

    # Spares and service lines: blank whatever the master list says, because
    # they travel inside somebody else's box or are not goods at all.
    with open(SERVICE_LINE_INVOICE, "rb") as f:
        resp = client.post("/upload", data={"invoice": (f, SERVICE_LINE_INVOICE.name)},
                            content_type="multipart/form-data")
    service_draft = db.get_draft(conn, int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1]))
    service = {i["item_name"]: i["cartons"] for i in service_draft["items"]}
    check("a real fan on the same invoice is still filled in",
          service["RACE (AC) MATTE WHITE"] == "3")
    check("the installation services line is left blank",
          all(v == "" for k, v in service.items() if "ERECTION" in k))
    conn.close()


@needs_fixtures
def test_parsing_a_batch(tmpdir):
    """A batch is parsed several files at a time, in worker PROCESSES.

    pdfplumber is Python nearly all the way down, so a thread pool is measurably
    SLOWER than doing nothing (2.21s sequential against 3.16s on four threads,
    for eight copies of a two-page invoice). Processes give about 3.7x on fifty.
    """
    fixtures = [SAMPLE_INVOICE, TWO_PAGE_INVOICE, SERVICE_LINE_INVOICE,
                SAMPLE_INVOICE, TWO_PAGE_INVOICE, SERVICE_LINE_INVOICE]
    one_by_one = [invoice_parser.parse_invoice(f) for f in fixtures]
    together = invoice_parser.parse_many(fixtures)

    check("a batch returns one result per file", len(together) == len(fixtures))
    check("in the order it was given, not the order they finished",
          [r["invoice_no"] for r in together] == [r["invoice_no"] for r in one_by_one])
    check("with the same items as parsing them one at a time",
          [[i["item_name"] for i in r["items"]] for r in together]
          == [[i["item_name"] for i in r["items"]] for r in one_by_one])

    # Small batches skip the pool entirely: starting worker processes costs
    # about 100ms each, which is most of the work for two files.
    check("a small batch still parses correctly",
          len(invoice_parser.parse_many(fixtures[:2])) == 2)
    check("an empty batch is not an error", invoice_parser.parse_many([]) == [])

    # One unreadable file must cost only itself.
    bad = Path(tmpdir) / "not-a-pdf.pdf"
    bad.write_bytes(b"this is not a PDF at all")
    mixed = invoice_parser.parse_many(list(fixtures) + [bad])
    check("a broken file in the batch still yields a result", len(mixed) == len(fixtures) + 1)
    check("and it is marked unreadable rather than silently empty",
          bool(mixed[-1]["notes"]) and not mixed[-1]["items"])
    check("while the good ones parsed fine", all(r["items"] for r in mixed[:len(fixtures)]))

    # If the pool cannot start, EVERY file fails and the failures look exactly
    # like unreadable PDFs — which would turn a batch of fifty good invoices
    # into fifty drafts marked "could not be read". It must fall back instead.
    import concurrent.futures
    broken = concurrent.futures.ProcessPoolExecutor
    class Unusable:
        def __init__(self, *a, **k): raise OSError("no processes available")
    concurrent.futures.ProcessPoolExecutor = Unusable
    try:
        fallback = invoice_parser.parse_many(fixtures)
    finally:
        concurrent.futures.ProcessPoolExecutor = broken
    check("a pool that will not start falls back to parsing in order",
          [[i["item_name"] for i in r["items"]] for r in fallback]
          == [[i["item_name"] for i in r["items"]] for r in one_by_one])


def test_vehicle_prints_only_when_there_is_one(tmpdir):
    """No empty "Vehicle :" on a pass that did not travel by road.

    An empty label reads as a vehicle nobody bothered to record, which is a
    different claim from a consignment that had no vehicle. There is no longer
    a setting for it either: a tick box that only ever HIDES something already
    entered is a way to print a document missing information somebody typed.
    """
    flask_app, client = logged_in_app(tmpdir, "vehicle")
    conn = db.connect(flask_app.config["DB_PATH"])

    without = db.create_gate_pass(conn, None, "S", "C", "I1", "01-01-2026", "",
                                   sample_items(), prepared_by="Ravi Kumar")
    page = client.get(f"/print/{without['id']}").get_data(as_text=True)
    check("no vehicle recorded, no vehicle label", "Vehicle" not in page)
    check("and nothing is left dangling", 'class="totals-vehicle"' not in page)

    with_one = db.create_gate_pass(conn, None, "S", "C", "I2", "01-01-2026",
                                    "KA 01 AB 1234", sample_items(),
                                    prepared_by="Ravi Kumar")
    page = client.get(f"/print/{with_one['id']}").get_data(as_text=True)
    check("a recorded vehicle is printed", "KA 01 AB 1234" in page)
    check("with a short label so a bare registration is not left unexplained",
          'class="totals-vehicle"' in page)
    check("and it does not depend on a setting being ticked",
          "show_vehicle" not in (ROOT / "templates" / "settings.html").read_text())
    conn.close()


def test_totals_sit_between_the_items_and_the_signatures(tmpdir):
    """Totals below the table, signatures below the totals, both at the end.

    They used to be a third column inside the footer, wedged between "Prepared
    by" and "Authorised by".
    """
    flask_app, client = logged_in_app(tmpdir, "totals_pos")
    conn = db.connect(flask_app.config["DB_PATH"])
    db.update_settings(conn, show_total_qty="1", show_total_cartons="1")
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "KA 01 AB 1234",
                              [{"item_name": "MICRON", "quantity": "3", "cartons": "2"},
                               {"item_name": "PHOENIX", "quantity": "4", "cartons": "1"}],
                              prepared_by="Ravi Kumar")
    page = client.get(f"/print/{gp['id']}").get_data(as_text=True)

    # The totals belong to the table, as its last row. As a strip underneath
    # they were a floating line whose numbers did not sit under the columns
    # they totalled.
    print_css_now = (ROOT / "static" / "css" / "print.css").read_text()
    check("there is no standalone totals strip any more",
          ".pass-totals" not in print_css_now and "pass-totals" not in page)
    check("the totals are the table's own last row", "<tfoot>" in page)
    check("and they come after every item row",
          page.index("</tbody>") < page.index("<tfoot>"))
    # Compared against the ITEM table's closing tag: the meta box above is a
    # table too, and its </table> comes first.
    items_table = re.search(r'<table class="items-table".*?</table>', page, re.S).group(0)
    check("inside the table, not after it", "<tfoot>" in items_table)

    # The numbers land in the columns they total, which is the whole point.
    tfoot = re.search(r"<tfoot>.*?</tfoot>", page, re.S).group(0)
    check("the quantity total is in the Quantity column",
          re.search(r'<td class="col-qty"><strong>7</strong>', tfoot))
    check("the carton total is in the No. of Cartons column",
          re.search(r'<td class="col-ctn"><strong>3</strong>', tfoot))
    check("and the row is labelled", "totals-label" in tfoot)

    # A td with display:flex stops being a table-cell: it leaves the column
    # grid and takes its share of the ruled border with it, so the rule above
    # the totals ended halfway across. Already documented on .actions-cell in
    # register.html, and hit again here — so it is a check now, not a comment.
    for rule_text in re.findall(r"table\.items-table[^{]*\{[^}]*\}", print_css_now):
        selector = rule_text.split("{")[0]
        if "display: flex" in rule_text:
            check(f"no table cell is turned into a flex container ({selector.strip()})",
                  " td" not in selector and not selector.rstrip().endswith(("col-item",
                      "col-qty", "col-ctn", "col-sl")))

    # Its height comes out of the table's budget, not out of the page, so the
    # ruled box still ends where it always did and no item is pushed off a
    # full page.
    check("the row height allows for the totals row",
          db.print_row_height_mm(26, True) < db.print_row_height_mm(26))
    check("and the table still fills its budget either way",
          abs(db.print_row_height_mm(26, True) * 26 + db.TOTALS_ROW_MM
              - db.ITEM_BODY_MM) < 0.2)

    # A blank carton column must not be totalled as a confident 0.
    blank = db.create_gate_pass(conn, None, "S", "C", "I2", "01-01-2026", "",
                                 sample_items(cartons=""), prepared_by="Ravi Kumar")
    page = client.get(f"/print/{blank['id']}").get_data(as_text=True)
    blank_foot = re.search(r"<tfoot>.*?</tfoot>", page, re.S).group(0)
    check("cartons nobody counted are not totalled as 0",
          not re.search(r'col-ctn"><strong>', blank_foot))
    check("but the quantity still is", re.search(r'col-qty"><strong>', blank_foot))
    conn.close()


def test_the_two_totals_are_independent(tmpdir):
    """Quantity and cartons are separate ticks.

    An office that counts cartons by hand at the gate wants the quantity
    totalled and the carton column left alone. One combined setting could not
    say that.
    """
    flask_app, client = logged_in_app(tmpdir, "two_totals")
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                              [{"item_name": "MICRON", "quantity": "3", "cartons": "2"},
                               {"item_name": "PHOENIX", "quantity": "4", "cartons": "1"}],
                              prepared_by="Ravi Kumar")

    def printed():
        return client.get(f"/print/{gp['id']}").get_data(as_text=True)

    def totals_row():
        found = re.search(r"<tfoot>.*?</tfoot>", printed(), re.S)
        return found.group(0) if found else ""

    def figure(column):
        found = re.search(rf'<td class="col-{column}"><strong>(\d+)</strong>', totals_row())
        return found.group(1) if found else None

    check("both off prints no totals row at all", totals_row() == "")

    db.update_settings(conn, show_total_qty="1", show_total_cartons="0")
    check("quantity alone prints the quantity", figure("qty") == "7")
    check("and not the cartons", figure("ctn") is None)

    db.update_settings(conn, show_total_qty="0", show_total_cartons="1")
    check("cartons alone prints the cartons", figure("ctn") == "3")
    check("and not the quantity", figure("qty") is None)

    db.update_settings(conn, show_total_qty="1", show_total_cartons="1")
    check("both on prints both", figure("qty") == "7" and figure("ctn") == "3")
    check("each under the column it totals, not run together",
          'class="col-qty"><strong>7' in totals_row()
          and 'class="col-ctn"><strong>3' in totals_row())

    # Cartons nobody typed are still not totalled as a confident 0, whatever
    # the tick box says.
    blank = db.create_gate_pass(conn, None, "S", "C", "I2", "01-01-2026", "",
                                 sample_items(cartons=""), prepared_by="Ravi Kumar")
    page = client.get(f"/print/{blank['id']}").get_data(as_text=True)
    check("an untouched carton column is not totalled as 0",
          "Total Cartons" not in page)

    # Both tick boxes exist on the settings page and post independently.
    settings_page = client.get("/settings").get_data(as_text=True)
    check("the settings page offers two separate ticks",
          'name="show_total_qty"' in settings_page
          and 'name="show_total_cartons"' in settings_page)
    check("and no longer offers the combined one",
          'name="show_totals"' not in settings_page)
    client.post("/settings", headers={"Origin": "http://localhost"},
                data={"company_name": "fanzart", "paper_mode": "a4x2",
                      "show_total_qty": "on"})
    saved = db.get_settings(db.connect(flask_app.config["DB_PATH"]))
    check("posting one tick turns only that one on",
          saved["show_total_qty"] == "1" and saved["show_total_cartons"] == "0")
    conn.close()


def test_upgrading_keeps_a_totals_preference(tmpdir):
    """Splitting the setting must not silently switch it off.

    Someone who had totals printing keeps both halves; the pass they print the
    day after an update looks like the one they printed the day before.
    """
    path = Path(tmpdir) / "old_totals.db"
    conn = db.connect(path)
    db.update_settings(conn, show_total_qty="1", show_total_cartons="1")
    # Put the database back the way version 13 left it.
    with db.writing(conn):
        conn.execute("DELETE FROM settings WHERE key LIKE 'show_total_%'")
        conn.execute("INSERT INTO settings (key, value) VALUES ('show_totals', '1')")
    conn.execute("PRAGMA user_version = 13")
    conn.close()

    upgraded = db.get_settings(db.connect(path))
    check("an upgraded database keeps the quantity total on",
          upgraded["show_total_qty"] == "1")
    check("and the carton total on", upgraded["show_total_cartons"] == "1")
    check("and the old key is cleared away", "show_totals" not in upgraded)


def test_a_two_page_pass_signs_off_only_at_the_end(tmpdir):
    """Totals and signatures belong at the end of the document, once.

    A signature on page 1 of 2 says that half of a gate pass was authorised on
    its own.
    """
    flask_app, client = logged_in_app(tmpdir, "twopage")
    conn = db.connect(flask_app.config["DB_PATH"])
    db.update_settings(conn, show_total_qty="1", show_total_cartons="1")
    many = [{"item_name": f"ITEM {i}", "quantity": "1", "cartons": "1"}
            for i in range(db.ITEMS_PER_PAGE + 4)]
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "KA 01 AB 1234",
                              many, prepared_by="Ravi Kumar")
    page = client.get(f"/print/{gp['id']}").get_data(as_text=True)

    sheets = page.count('class="sheet')
    check("a long pass prints on more than one sheet", sheets >= 2)
    check("the pages are numbered so a loose sheet is identifiable",
          "Page 1 of 2" in page and "Page 2 of 2" in page)

    # Two copies per A4 sheet, so the last page contributes two of each.
    check("the signatures appear only on the last page",
          page.count('class="pass-foot"') == 2)
    check("and so do the totals", page.count("<tfoot>") == 2)
    check("every page still has the end block, so the geometry does not shift",
          page.count('class="pass-end"') == len(re.findall(r'class="pass[ "]', page)))
    conn.close()


def test_duplicate_document_numbers_are_flagged_not_blocked(tmpdir):
    """The same document number arriving twice is a warning, never a block.

    A file uploaded twice in one batch, or a document already turned into a
    gate pass last week, would otherwise spend a second number on a
    consignment the register already accounts for. But the same number
    legitimately recurs — a split delivery, a corrected reissue — and the
    person at the desk knows which it is. What they cannot do is spot it
    unaided across fifty uploaded files.
    """
    flask_app, client = logged_in_app(tmpdir, "dupes")
    conn = db.connect(flask_app.config["DB_PATH"])

    # The comparison has to survive the ways the same number is written.
    for written in ("FR 262700644", "TO NO: FR 262700644", "fr262700644",
                    "FR-262700644", "Document No: FR 262700644"):
        check(f"{written!r} normalises to the same key",
              db.normalize_document_no(written) == "FR262700644")
    check("a different number does not collide",
          db.normalize_document_no("FR 262700645") != "FR262700644")
    check("an empty number has no key", db.normalize_document_no("") == "")

    issued = db.create_gate_pass(conn, None, "S", "C", "FR 262700644",
                                  "01-01-2026", "", sample_items(),
                                  prepared_by="Ravi Kumar")
    # Written differently from the issued one on purpose.
    already = db.create_draft(conn, supplier_name="S", customer_name="C",
                               invoice_no="TO NO: FR 262700644",
                               invoice_date="01-01-2026",
                               invoice_pdf_path="invoices/20260101_TO 1.pdf",
                               items=sample_items())
    twin_a = db.create_draft(conn, supplier_name="S", customer_name="C",
                              invoice_no="FR 262700652", invoice_date="01-01-2026",
                              invoice_pdf_path="invoices/20260101_TO 2.pdf",
                              items=sample_items())
    twin_b = db.create_draft(conn, supplier_name="S", customer_name="C",
                              invoice_no="FR-262700652", invoice_date="01-01-2026",
                              invoice_pdf_path="invoices/20260101_TO 3.pdf",
                              items=sample_items())
    unique = db.create_draft(conn, supplier_name="S", customer_name="C",
                              invoice_no="FR 262700999", invoice_date="01-01-2026",
                              items=sample_items())

    report = db.duplicate_document_report(conn, db.list_drafts(conn))
    check("a number already issued is caught", already in report)
    check("and the message names the gate pass",
          issued["serial_no"] in report[already]["message"])
    check("but it does not stop the draft being issued",
          report[already]["blocking"] is False)
    check("two files in the same batch are caught",
          twin_a in report and twin_b in report)
    check("and the message names both files",
          "TO 2.pdf" in report[twin_a]["message"]
          and "TO 3.pdf" in report[twin_a]["message"])
    # Two drafts in hand with one number cannot both go: the register would
    # show a consignment leaving twice, and nothing could tell the passes apart.
    check("and that one DOES stop them being issued",
          report[twin_a]["blocking"] and report[twin_b]["blocking"])
    check("a number nobody has seen is not flagged", unique not in report)

    page = client.get("/drafts").get_data(as_text=True)
    check("duplicate rows are highlighted", "is-duplicate" in page)
    check("with a badge", "DUPLICATE" in page.upper())
    check("and the reason under the number", "duplicate-note" in page)
    check("the reason names the issued pass", issued["serial_no"] in page)

    # The button must stay usable: the operator decides, not the app.
    button = re.search(r'<button[^>]*id="issue-selected"[^>]*>(.*?)</button>', page, re.S)
    check("the issue button is not disabled by a duplicate",
          button and "disabled" not in button.group(0))
    check("and it asks to be confirmed", "Proceed" in button.group(1))
    check("a confirmation is wired up", "confirm(" in page)
    # Only for duplicates that are actually still ticked — warning about one
    # the operator already unticked is how a prompt becomes noise.
    check("and only for duplicates still selected",
          ".draft-tick:checked" in page)
    conn.close()


def test_two_drafts_with_one_number_cannot_both_be_issued(tmpdir):
    """A duplicate inside the batch is refused; one already issued is only a warning.

    They are not the same problem. Two drafts in hand carrying one number would
    spend two gate pass numbers on one consignment, and afterwards nothing
    could tell the two passes apart — there is no reading of that which is
    correct. A number that is on an already-issued pass is different: a split
    delivery against one invoice, or a reissue after a correction, are both
    ordinary, and the person at the desk knows which it is.
    """
    flask_app, client = logged_in_app(tmpdir, "blockdupes")
    conn = db.connect(flask_app.config["DB_PATH"])

    def draft(number, source):
        return db.create_draft(conn, supplier_name="S", customer_name="C",
                                invoice_no=number, invoice_date="01-01-2026",
                                invoice_pdf_path=f"invoices/20260101120000_{source}",
                                items=sample_items())

    twin_a, twin_b = draft("FR 262700644", "TO 1_2.pdf"), draft("FR-262700644", "TO 2_2.pdf")
    clean = draft("FR 262700999", "TO 3.pdf")

    blocked = db.blocking_duplicates(conn, [twin_a, twin_b, clean])
    check("both halves of the pair are blocked", set(blocked) == {twin_a, twin_b})
    check("the message names both files",
          "TO 1_2.pdf" in blocked[twin_a] and "TO 2_2.pdf" in blocked[twin_a])
    check("and says what to do", "remove one" in blocked[twin_a].lower())
    check("the clean draft is not blocked", clean not in blocked)

    page = client.get("/drafts").get_data(as_text=True)
    # The badge reads "Duplicate" and the line under it carries the rest; the
    # longer wording did not fit the column.
    check("a blocked draft is marked as such", "&#9888; Duplicate<" in page)
    check("and says how to clear it", "Remove one with" in page)
    # No tick for a draft that cannot go — but that is presentation, not a rule.
    check("only the clean draft can be selected",
          page.count('class="draft-tick"') == 1)

    # The rule is on the server. A checkbox that was never rendered does not
    # stop the form being posted with the id in it.
    before = len(db.list_gate_passes(conn))
    resp = client.post("/drafts/issue", headers={"Origin": "http://localhost"},
                       data={"draft_id": [str(twin_a), str(twin_b), str(clean)]})
    check("posting them anyway is refused", resp.status_code == 302)
    check("and nothing at all was numbered", len(db.list_gate_passes(conn)) == before)
    check("not even the clean one in the batch",
          db.get_draft(conn, clean) is not None)
    check("and the reason is shown", b"Not issued" in client.get("/drafts").data)

    # Remove one and the rest go through.
    client.post(f"/drafts/{twin_b}/delete", headers={"Origin": "http://localhost"})
    resp = client.post("/drafts/issue", headers={"Origin": "http://localhost"},
                       data={"draft_id": [str(twin_a), str(clean)]})
    check("with one of them gone the batch issues",
          len(db.list_gate_passes(conn)) == before + 2)

    # An already-issued number is a warning, and warnings do not block.
    again = draft("FR 262700644", "TO 4.pdf")
    report = db.duplicate_document_report(conn, db.list_drafts(conn))
    check("a number already issued is flagged", again in report)
    check("but not blocked", report[again]["blocking"] is False)
    check("and it can still be issued",
          client.post("/drafts/issue", headers={"Origin": "http://localhost"},
                      data={"draft_id": str(again)}).status_code == 302)
    check("and it really was", len(db.list_gate_passes(conn)) == before + 3)

    # A CANCELLED pass must not raise the warning at all. Its number is spent
    # but its document is not, and reissuing against it is exactly how a pass
    # cancelled for being wrong gets replaced.
    # A number of its own: FR 262700644 is on two issued passes by now, so
    # cancelling one of them would leave the other still matching and prove
    # nothing about cancellation.
    doomed = db.create_gate_pass(conn, None, "S", "C", "FR 262700777",
                                  "01-01-2026", "", sample_items(),
                                  prepared_by="Ravi Kumar")
    fresh = draft("FR 262700777", "TO 5.pdf")
    check("while that pass is issued, the number is flagged",
          fresh in db.duplicate_document_report(conn, [db.get_draft(conn, fresh)]))
    db.cancel_gate_pass(conn, doomed["id"], "printed in error", cancelled_by="Ravi Kumar")
    check("once it is cancelled, the number is free again",
          fresh not in db.duplicate_document_report(conn, [db.get_draft(conn, fresh)]))
    conn.close()


def test_totals_are_shown_live_and_derived_on_save(tmpdir):
    """Running totals while typing, and totals on the record that cannot drift.

    The screen figure is a convenience. The one that matters is stored — and it
    is not stored at all: total_qty and total_cartons are computed from the item
    rows every time a pass or draft is read, so there is no saved total that can
    disagree with the items under it.
    """
    flask_app, client = logged_in_app(tmpdir, "livetotals")
    conn = db.connect(flask_app.config["DB_PATH"])
    draft_id = db.create_draft(conn, supplier_name="S", customer_name="C",
                                invoice_no="I", invoice_date="01-01-2026",
                                items=[{"sl_no": 1, "item_name": "MICRON",
                                        "quantity": "3", "cartons": "2"},
                                       {"sl_no": 2, "item_name": "PHOENIX",
                                        "quantity": "4", "cartons": "1"}])

    for screen, url in (("review", f"/review/{draft_id}"),):
        page = client.get(url).get_data(as_text=True)
        check(f"{screen} shows a running quantity total", "total-qty-display" in page)
        check(f"{screen} shows a running carton total", "total-cartons-display" in page)
        # Delegated from the tbody, so a row added later is covered without
        # anything being bound to it — a per-row listener is one that gets
        # forgotten on the next "+ Add item".
        check(f"{screen} listens on the table, not on each row",
              'tbody.addEventListener("input"' in page)
        check(f"{screen} ignores blanks rather than summing NaN",
              "Number.isNaN" in page)

    check("the edit screen has them too",
          "total-qty-display" in (ROOT / "templates" / "edit_pass.html").read_text())

    # A DRAFT carries no totals at all — it is not a record of anything yet,
    # and the running figure on the review screen is what the operator checks
    # against the invoice in their hand.
    check("a draft has no stored totals to go stale",
          "total_qty" not in db.get_draft(conn, draft_id).keys())

    # An issued pass does have them, and they are derived on every read rather
    # than written down, so there is nothing that can disagree with the items.
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                              [{"item_name": "MICRON", "quantity": "10", "cartons": "4"}],
                              prepared_by="Ravi Kumar")
    check("an issued pass totals its own items",
          gp["total_qty"] == 10 and gp["total_cartons"] == 4)
    db.update_gate_pass(conn, gp["id"], edited_by="Ravi Kumar",
                        items=[{"item_name": "MICRON", "quantity": "2", "cartons": "1"},
                               {"item_name": "PHOENIX", "quantity": "5", "cartons": "2"}])
    after = db.get_gate_pass(conn, gp["id"])
    check("and follows an edit without anything recalculating a stored figure",
          after["total_qty"] == 7 and after["total_cartons"] == 3)
    check("there is no stored total column to drift",
          "total_qty" not in [c[1] for c in conn.execute("PRAGMA table_info(gate_passes)")])
    conn.close()


def test_the_register_shows_what_has_reached_paper(tmpdir):
    """Which passes are printed, and how many times.

    A pass can be issued, hold its number and sit in the register for an hour
    before anybody prints it. There was no way to see which ones those were.
    """
    flask_app, client = logged_in_app(tmpdir, "printed")
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")

    check("a new pass has not been printed", gp["is_printed"] == 0)
    check("and has no count", gp["print_count"] == 0)
    check("and no time", gp["printed_at"] is None)

    page = client.get("/register").get_data(as_text=True)
    # The state is on the Print button, not on a badge and not on the row. The
    # register is read by scanning down it, and a mark on every pass that has
    # not been printed yet — which is every new pass — was a column of warnings
    # about nothing.
    check("the register says so", "print-action not-printed" in page)
    check("and does not tint the row", "awaiting-print" not in page)
    check("nor add a second badge", "Unprinted" not in page)

    # OPENING the print page is a preview — from the register, from a link, by
    # somebody checking a number. Marking it there would fill the register with
    # passes marked printed that nobody put on paper, which is worse than no
    # marking at all because it looks like information.
    client.get(f"/print/{gp['id']}")
    check("opening the print page does NOT mark it printed",
          db.get_gate_pass(conn, gp["id"])["print_count"] == 0)
    check("the print page posts only once the dialog finishes",
          "afterprint" in (ROOT / "templates" / "print.html").read_text())

    resp = client.post(f"/gate-passes/{gp['id']}/printed",
                       headers={"Origin": "http://localhost"})
    check("printing it is recorded", resp.get_json()["print_count"] == 1)
    after = db.get_gate_pass(conn, gp["id"])
    check("the flag is set", after["is_printed"] == 1)
    check("and the time", after["printed_at"] is not None)
    check("the serial is untouched by any of it",
          after["serial_no"] == gp["serial_no"])

    page = client.get("/register").get_data(as_text=True)
    check("the register shows it printed", "print-action is-printed" in page)
    check("and stops asking for paper", "not-printed" not in page)

    # Reprints are worth seeing: a pass run off four times is worth a look, and
    # the count is the only place that shows.
    client.post(f"/gate-passes/{gp['id']}/printed", headers={"Origin": "http://localhost"})
    check("a reprint increments rather than re-flagging",
          db.get_gate_pass(conn, gp["id"])["print_count"] == 2)
    check("and the register shows how many",
          "2 times" in client.get("/register").get_data(as_text=True))

    # A cancelled pass is not waiting for paper, so it is not chased for it.
    doomed = db.create_gate_pass(conn, None, "S", "C", "I2", "01-01-2026", "",
                                  sample_items(), prepared_by="Ravi Kumar")
    db.cancel_gate_pass(conn, doomed["id"], "printed in error", cancelled_by="Ravi Kumar")
    page = client.get("/register").get_data(as_text=True)
    check("a cancelled pass is not marked as awaiting print",
          "awaiting-print" not in page)
    check("nor labelled unprinted", "Unprinted" not in page)
    # And its Print button gets neither state: solid would be a standing
    # instruction to print a pass that should not go out.
    check("nor is its Print button styled as owed",
          page.count("print-action not-printed") == 0)

    check("printing an unknown pass is a 404",
          client.post("/gate-passes/999999/printed",
                      headers={"Origin": "http://localhost"}).status_code == 404)

    # A batch print marks every pass in it, not only the first.
    batch = (ROOT / "templates" / "print_batch.html").read_text()
    check("the batch page marks each pass it printed",
          "{% for gate_pass in passes %}" in batch
          and batch.count("mark_gate_pass_printed") == 1)
    conn.close()


def test_the_printed_totals_are_bolder_than_the_rows(tmpdir):
    """The summary row reads as a summary, and still fits its budgeted height."""
    css = (ROOT / "static" / "css" / "print.css").read_text()
    foot = re.search(r"table\.items-table tfoot td \{(.*?)\}", css, re.S).group(1)
    size = float(re.search(r"font-size: ([\d.]+)pt", foot).group(1))
    weight = int(re.search(r"font-weight: (\d+)", foot).group(1))
    rows = re.search(r"table\.items-table \{(.*?)\}", css, re.S).group(1)
    row_size = float(re.search(r"font-size: ([\d.]+)pt", rows).group(1))

    check("the totals are heavier than normal text", weight >= 700)
    check("and larger than the item rows", size > row_size)
    check("the figures inside are heavier still",
          re.search(r"tfoot strong \{[^}]*font-weight: 800", css))
    # Weight is dropped along with colour on some print paths unless the page
    # is rendered as authored.
    check("and it survives printing", "print-color-adjust: exact" in foot)

    # The row height is fixed and comes out of the table's budget, so bigger
    # text must still fit inside it or the pass overflows the sheet.
    height = float(re.search(r"height: ([\d.]+)mm", foot).group(1))
    line_mm = size * 25.4 / 72 * 1.2
    check(f"{size}pt still fits the {height}mm row", line_mm < height)
    check("and the row is still the height the budget allows for",
          abs(height - db.TOTALS_ROW_MM) < 0.01)


def test_correcting_an_issued_pass(tmpdir):
    """An issued pass can be corrected, and the correction is on the record.

    This is the heaviest permission in the app: everything else changes what
    happens next, this changes what the register says already happened. What
    makes it defensible is that the pass stays the same pass — same number,
    same issue date, same preparer — and the old values survive in the audit
    log even though they no longer survive in the row.
    """
    flask_app, admin = logged_in_app(tmpdir, "editpass",
                                      users=(("dinesh", "Dinesh D"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "OLD CUSTOMER",
                              "FR 1", "01-01-2026", "",
                              [{"item_name": "MICRON MODREN OAK", "quantity": "3", "cartons": ""}],
                              prepared_by="Rakesh")

    check("the permission exists and is a real one",
          "can_edit_issued_pass" in db.PERMISSIONS)
    check("and it is explained on the people screen",
          "can_edit_issued_pass" in db.PERMISSION_HINTS)
    check("a fresh account does not have it",
          not db.user_can(db.get_user_by_username(conn, "asha"), "can_edit_issued_pass"))

    # Someone without it cannot reach the screen, and is not shown the way in.
    staff = flask_app.test_client()
    sign_in(staff, "asha")
    check("without the permission the edit screen is refused",
          staff.get(f"/gate-passes/{gp['id']}/edit").status_code == 302)
    check("and no Edit button is offered", b"/edit" not in staff.get("/register").data)
    check("with it, the button is there", b"/edit" in admin.get("/register").data)

    form = MultiDict([
        ("supplier_name", "Golden Touch Exports"), ("customer_name", "NEW CUSTOMER"),
        ("invoice_no", "FR 1"), ("invoice_date", "01-01-2026"),
        ("vehicle_no", "KA 01 AB 1234"), ("remarks", "corrected at the gate"),
        ("item_name", "MICRON MODREN OAK"), ("quantity", "5"), ("cartons", "5")])
    resp = admin.post(f"/gate-passes/{gp['id']}/edit", data=form,
                      headers={"Origin": "http://localhost"})
    check("saving returns to the register", resp.headers.get("Location", "").endswith("/register"))
    check("and says so", b"updated successfully" in admin.get("/register").data)

    after = db.get_gate_pass(conn, gp["id"])
    # The three things that must not move, whatever else does.
    check("the serial number is unchanged", after["serial_no"] == gp["serial_no"])
    check("the issue date is unchanged", after["issued_at"] == gp["issued_at"])
    check("the preparer is still whoever issued it", after["prepared_by"] == "Rakesh")
    # And the things that were meant to change.
    check("the customer is corrected", after["customer_name"] == "NEW CUSTOMER")
    check("the vehicle is recorded", after["vehicle_no"] == "KA 01 AB 1234")
    check("the remark is recorded", after["remarks"] == "corrected at the gate")
    check("the quantity is corrected", after["items"][0]["quantity"] == "5")
    check("the cartons are corrected", after["items"][0]["cartons"] == "5")

    # The whole justification for allowing this at all.
    logged = conn.execute(
        "SELECT details FROM audit_log WHERE gate_pass_id = ? AND action = 'edit'",
        (gp["id"],)).fetchall()
    check("the edit is written to the audit log", len(logged) == 1)
    details = logged[0]["details"]
    check("naming who made it", "Dinesh D" in details)
    check("and keeping the old value beside the new",
          "OLD CUSTOMER -> NEW CUSTOMER" in details)
    check("including the quantity that changed", "qty 3->5" in details)

    # An edit that changes nothing must not fabricate a record of a change.
    same = db.update_gate_pass(conn, gp["id"], edited_by="Dinesh D",
                               customer_name="NEW CUSTOMER")
    check("a no-op edit records nothing", same == [])
    check("and adds no audit row", len(conn.execute(
        "SELECT 1 FROM audit_log WHERE gate_pass_id = ? AND action = 'edit'",
        (gp["id"],)).fetchall()) == 1)

    # The database still refuses what it always refused.
    check("the serial cannot be edited even directly",
          raises(sqlite3.IntegrityError, lambda: conn.execute(
              "UPDATE gate_passes SET serial_no = 'XX-99999' WHERE id = ?", (gp["id"],))))

    # A cancelled pass is a closed record.
    db.cancel_gate_pass(conn, gp["id"], "printed in error", cancelled_by="Dinesh D")
    check("a cancelled pass cannot be edited",
          raises(ValueError, db.update_gate_pass, conn, gp["id"],
                 None, "Dinesh D", customer_name="ANYTHING"))
    check("its edit screen turns you away",
          admin.get(f"/gate-passes/{gp['id']}/edit").status_code == 302)
    check("and the register offers no Edit for it",
          admin.get("/register").data.count(b"/edit") == 0)

    # An empty pass is not a pass.
    live = db.create_gate_pass(conn, None, "S", "C", "FR 2", "01-01-2026", "",
                                sample_items(), prepared_by="Rakesh")
    check("a pass cannot be emptied of every item",
          raises(ValueError, db.update_gate_pass, conn, live["id"], [], "Dinesh D"))
    conn.close()


def test_a_cancelled_pass_says_so_on_paper(tmpdir):
    """Cancelling keeps the number and the row, so the paper must say what it is.

    Nothing deletes a gate pass. A cancelled one still prints, still carries its
    serial, and would otherwise be indistinguishable from a live pass in a
    drawer — which is the whole reason for the watermark.
    """
    flask_app, client = logged_in_app(tmpdir, "watermark")
    conn = db.connect(flask_app.config["DB_PATH"])
    live = db.create_gate_pass(conn, None, "S", "C", "I1", "01-01-2026", "",
                                sample_items(), prepared_by="Ravi Kumar")
    doomed = db.create_gate_pass(conn, None, "S", "C", "I2", "01-01-2026", "",
                                  sample_items(), prepared_by="Ravi Kumar")

    page = client.get(f"/print/{live['id']}").get_data(as_text=True)
    check("an issued pass carries no watermark", "cancelled-watermark" not in page)

    db.cancel_gate_pass(conn, doomed["id"], "printed in error", cancelled_by="Ravi Kumar")
    page = client.get(f"/print/{doomed['id']}").get_data(as_text=True)
    check("a cancelled pass is marked CANCELLED", "cancelled-watermark" in page)
    check("and the word is on it", ">CANCELLED<" in page)
    # One per copy: an A4 sheet carries two, to be cut apart. A watermark on
    # only one half leaves the other looking live.
    # Counted with a boundary: 'class="pass' also matches pass-head, pass-end
    # and pass-foot, which would make any number look right.
    copies = len(re.findall(r'class="pass(?:\s[^"]*)?"', page))
    check("both copies on the sheet are marked",
          copies == 2 and page.count("cancelled-watermark") == copies)

    # The stored value is lower case — the column has a CHECK constraint
    # allowing only 'issued' and 'cancelled'. Testing against 'CANCELLED'
    # would render nothing at all, on every pass.
    card = (ROOT / "templates" / "_pass_card.html").read_text()
    check("the template tests the value the database actually stores",
          "gate_pass.status == 'cancelled'" in card)
    check("the row really is stored lower case",
          db.get_gate_pass(conn, doomed["id"])["status"] == "cancelled")

    css = (ROOT / "static" / "css" / "print.css").read_text()
    rule = re.search(r"\.cancelled-watermark \{(.*?)\}", css, re.S).group(1)
    check("it is laid over the pass, not inserted into it", "position: absolute" in rule)
    check("it cannot be clicked through", "pointer-events: none" in rule)
    check("it sits above the content", "z-index: 100" in rule)
    # Browsers drop colours when printing unless told not to. A watermark that
    # vanishes on paper is worse than none: the screen would say cancelled and
    # the document in somebody's hand would not.
    check("its colour survives printing", "print-color-adjust: exact" in rule)
    # Absolute positioning needs a positioned ancestor, or it escapes to the
    # page and both copies stack theirs across the cut line.
    pass_rule = re.search(r"^\.pass \{(.*?)\}", css, re.S | re.M).group(1)
    check("the pass is the box it is positioned against",
          "position: relative" in pass_rule)
    conn.close()


def test_printed_sheet_has_no_rule_between_the_copies(tmpdir):
    """The gap between the two copies is white space, not a cut guide.

    A dashed rule down the middle of a printed document reads as part of the
    pass rather than as an instruction to whoever is holding the scissors.
    """
    print_css = (ROOT / "static" / "css" / "print.css").read_text()
    cut_rule = print_css[print_css.index(".cut-line {"):]
    cut_rule = cut_rule[:cut_rule.index("}")]
    check("no border is drawn between the copies",
          "border" not in cut_rule)
    check("but they are still held apart", "margin" in cut_rule)
    without_comments = re.sub(r"/\*.*?\*/", "", print_css, flags=re.S)
    check("and nothing anywhere draws a dashed or dotted rule",
          not re.search(r"border[^;:]*:[^;]*(dashed|dotted)", without_comments))


def test_remarks_print_on_more_than_one_line(tmpdir):
    """A typed newline becomes a line break on the printed pass, safely."""
    flask_app, client = logged_in_app(tmpdir, "remarks_lines")
    conn = db.connect(flask_app.config["DB_PATH"])
    draft_id = db.create_draft(conn, supplier_name="S", customer_name="C",
                                invoice_no="I", invoice_date="01-01-2026",
                                items=sample_items(),
                                remarks="line one\nline two\n<b>three</b>")
    review = client.get(f"/review/{draft_id}").get_data(as_text=True)
    check("remarks is a textarea, so more than one line can be typed",
          '<textarea id="remarks"' in review)
    check("and the existing newlines survive the round trip",
          "line one\nline two" in review)

    resp = client.post(f"/review/{draft_id}", headers={"Origin": "http://localhost"},
                       data=MultiDict([
                           ("supplier_name", "S"), ("customer_name", "C"),
                           ("invoice_no", "I"), ("invoice_date", "01-01-2026"),
                           ("vehicle_no", ""), ("remarks", "line one\nline two"),
                           ("item_name", "MICRON MODREN OAK"), ("quantity", "1"),
                           ("cartons", ""), ("action", "issue")]))
    printed = client.get(resp.headers["Location"]).get_data(as_text=True)
    check("the printed pass breaks the line", "line one<br>line two" in printed)

    # A hanging indent, so line two sits under line one rather than under the
    # word "Remarks". Done with a float: the obvious padding-left plus negative
    # text-indent silently half-worked, because `tbody td` sets padding with
    # !important, and the label ended up hanging outside its own cell.
    check("the label and the text are separate elements",
          'class="remarks-label"' in printed and 'class="remarks-body"' in printed)
    print_css = (ROOT / "static" / "css" / "print.css").read_text()
    # Matched on the rule itself, not the first mention of the name — the
    # selector is also referred to in a comment further up.
    body_rule = re.search(r"^\.remarks-body\s*\{(.*?)\}", print_css, re.S | re.M)
    body_rule = body_rule.group(1) if body_rule else ""
    check("the label floats out of the text flow", ".remarks-label { float: left;" in print_css)
    check("and the text keeps its own block, so every line starts together",
          "display: block;" in body_rule and "overflow: hidden;" in body_rule)
    check("the indent does not depend on padding that !important overrides",
          not re.search(r"^\s*text-indent\s*:", print_css, re.M))
    # The Remarks box shares its table row with the Document No / Date box,
    # which is two lines tall. A third line grows the row and pushes the bottom
    # of the pass off the page, silently, because .pass is overflow:hidden.
    check("the printed remark is capped so it cannot grow the header",
          "max-height" in body_rule)
    check("and does not print a literal backslash-n", "line one\\nline two" not in printed)

    # The remark is typed by a person and printed on a document the company
    # signs. Markup in it must appear as text, not be interpreted.
    escaped = flask_app.jinja_env.filters["line_breaks"]("<script>x</script>\nnext")
    check("markup in a remark is escaped, not rendered",
          "&lt;script&gt;" in str(escaped) and "<script>" not in str(escaped))
    check("while the newline still becomes a break", "<br>" in str(escaped))
    conn.close()


def test_carton_lookup_rules(tmpdir):
    """The matching itself: what counts as the same item, and what stays blank."""
    storage = Path(tmpdir) / "carton_rules"
    flask_app = create_app(db_path=storage / "gate_pass.db", storage_dir=storage)
    conn = db.connect(flask_app.config["DB_PATH"])

    added, updated, unchanged = db.upsert_carton_mappings(
        conn, [("AARI - APRICOT GOLD", 1), ("CRYSTAL - FANDELIER", 2)])
    check("loading the master list reports what it did", (added, updated, unchanged) == (2, 0, 0))
    check("reloading the same rows changes nothing",
          db.upsert_carton_mappings(conn, [("AARI - APRICOT GOLD", 1)]) == (0, 0, 1))
    check("a corrected count updates in place",
          db.upsert_carton_mappings(conn, [("AARI - APRICOT GOLD", 3)]) == (0, 1, 0))
    check("and does not add a second row", db.carton_mapping_count(conn) == 2)
    db.upsert_carton_mappings(conn, [("AARI - APRICOT GOLD", 1)])

    # Invoices spell the same fan several ways. All of these are one item.
    for spelling in ("AARI - APRICOT GOLD", "aari - apricot gold", "AARI-APRICOT GOLD",
                     "AARI  -  APRICOT  GOLD", "  AARI - APRICOT GOLD  ",
                     "AARI \u2013 APRICOT GOLD", "AARI \u2014 APRICOT GOLD"):
        check(f"matches {spelling!r}", db.carton_count_for(conn, spelling) == 1)

    check("a different model does not match", db.carton_count_for(conn, "AARI - MATTE WHITE") is None)
    check("an empty name does not match", db.carton_count_for(conn, "") is None)
    check("a two-carton item reads back as 2",
          db.carton_count_for(conn, "CRYSTAL - FANDELIER") == 2)
    check("and scales: three of them is six boxes",
          db.cartons_for_line(conn, "CRYSTAL - FANDELIER", "3") == 6)

    # Spares are blank even if some future master row happens to match them.
    db.upsert_carton_mappings(conn, [("SPARE BLADE SET", 1)])
    check("a spare stays blank despite being in the master list",
          db.carton_count_for(conn, "SPARE BLADE SET") is None)
    for phrase in ("SPARE PART", "SPARES", "SERVICE", "INSTALLATION SERVICES",
                   "FREIGHT", "FREIGHT CHARGES", "HARDWARE KIT", "ACCESSORY",
                   "ACCESSORIES BOX"):
        check(f"treated as non-stock: {phrase}", db.is_non_stock_item(phrase))
    # Whole words only, or a model name would be caught by accident.
    for phrase in ("SERVICEABLE FAN", "SPAREBLADE", "AARI - APRICOT GOLD"):
        check(f"not caught by a substring: {phrase}", not db.is_non_stock_item(phrase))

    # An unreadable quantity means the arithmetic cannot be done, so the box is
    # left for a human rather than filled with a number nobody worked out.
    check("a blank quantity leaves it blank", db.cartons_for_line(conn, "AARI - APRICOT GOLD", "") is None)
    check("a non-numeric quantity leaves it blank",
          db.cartons_for_line(conn, "AARI - APRICOT GOLD", "a few") is None)
    check("a zero quantity leaves it blank",
          db.cartons_for_line(conn, "AARI - APRICOT GOLD", "0") is None)

    # A count the document itself carries beats the master list.
    filled = db.fill_cartons(conn, [
        {"item_name": "AARI - APRICOT GOLD", "quantity": "2", "cartons": "9"},
        {"item_name": "AARI - APRICOT GOLD", "quantity": "2", "cartons": ""},
    ])
    check("an existing carton value is never overwritten", filled[0]["cartons"] == "9")
    check("a blank one is filled in", filled[1]["cartons"] == "2")
    conn.close()


def test_setting_the_numbering(tmpdir):
    """An admin can point the numbering at any prefix and any free number.

    A prefix that has been used before is allowed. It used to be refused, on
    the reasoning that a new run starts at 00001 and would hand out a number
    already on a printed pass — but that is only true of a FRESH prefix, since
    the next number is read from MAX(serial_seq) for the run. Reusing FZ after
    FZ-00003 continues at FZ-00004.

    What still cannot happen is issuing a number twice, and that is enforced by
    the database, not by this check: (serial_scope, serial_seq) is UNIQUE.
    """
    flask_app, admin = logged_in_app(tmpdir, "run_storage",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])

    check("a fresh book is numbered from FZ-00001",
          db.next_serial_preview(conn) == "FZ-00001")
    first = [db.create_gate_pass(conn, None, "S", "C", f"I{i}", "03-08-2026", "",
                                  sample_items(), prepared_by="Ravi Kumar")
             for i in range(3)]
    check("the first run numbers straight through",
          [gp["serial_no"] for gp in first] == ["FZ-00001", "FZ-00002", "FZ-00003"])

    # Reuse resumes; it does not restart.
    prefix, start = db.start_new_run(conn, "FZ", started_by="Ravi Kumar")
    check("reusing the current prefix carries on from the last number", start == 4)
    check("and the preview agrees", db.next_serial_preview(conn) == "FZ-00004")
    check("spelled with a trailing separator it is the same run",
          db.start_new_run(conn, "FZ-", started_by="Ravi Kumar") == ("FZ-", 4))
    check("and prints only one separator", db.next_serial_preview(conn) == "FZ-00004")
    check("lower case is the same run too",
          db.start_new_run(conn, "fz", started_by="Ravi Kumar") == ("FZ", 4))

    check("a prefix with a space is refused",
          raises(ValueError, db.start_new_run, conn, "FZ 27"))
    check("a number already issued is refused",
          raises(ValueError, db.start_new_run, conn, "FZ", 1))
    try:
        db.start_new_run(conn, "FZ", 1)
    except ValueError as exc:
        check("and the message names the first free number", "FZ-00004" in str(exc))
    check("the failed attempts changed nothing",
          db.next_serial_preview(conn) == "FZ-00004")

    # A whole example serial sets prefix and starting number together.
    check("FZ-00001 parses into a prefix and a start",
          db.parse_run_spec("FZ-00001") == ("FZ-", 1))
    check("but FZ27 is a prefix in its own right, not FZ starting at 27",
          db.parse_run_spec("FZ27") == ("FZ27", None))

    db.start_new_run(conn, "FZ27-", 1, started_by="Ravi Kumar")
    check("a new run starts where it was asked to",
          db.next_serial_preview(conn) == "FZ27-00001")
    second = db.create_gate_pass(conn, None, "S", "C", "NEW", "03-08-2026", "",
                                 sample_items(), prepared_by="Ravi Kumar")
    check("the first pass of the new run is numbered 00001",
          second["serial_no"] == "FZ27-00001")
    check("a second one follows it rather than colliding",
          db.create_gate_pass(conn, None, "S", "C", "NEW2", "03-08-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")["serial_no"]
          == "FZ27-00002")

    # Jumping forward, for a spoiled box of pre-printed stationery.
    db.start_new_run(conn, "FZ27-", 500, started_by="Ravi Kumar")
    check("the numbering can be moved forward", db.next_serial_preview(conn) == "FZ27-00500")
    check("and issues from there",
          db.create_gate_pass(conn, None, "S", "C", "J", "03-08-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")["serial_no"]
          == "FZ27-00500")
    # The jump is stored as one setting, so it must not follow a different run.
    db.start_new_run(conn, "FZ", started_by="Ravi Kumar")
    check("moving back to another prefix does not inherit the jump",
          db.next_serial_preview(conn) == "FZ-00004")

    # Nothing from any earlier run is lost or altered.
    check("every earlier pass is still in the register", len(db.list_gate_passes(conn)) == 6)
    check("an earlier pass keeps its own number",
          db.get_gate_pass_by_serial(conn, "FZ-00003") is not None)
    check("the two 00001s are different passes",
          db.get_gate_pass_by_serial(conn, "FZ-00001")["id"]
          != db.get_gate_pass_by_serial(conn, "FZ27-00001")["id"])
    check("the change is recorded in the audit log",
          conn.execute("SELECT 1 FROM audit_log WHERE action = 'new_serial_run'").fetchone()
          is not None)
    conn.close()

    # Only admins may change it, and only deliberately.
    staff = flask_app.test_client()
    sign_in(staff, "asha")
    check("a non-admin cannot change the numbering",
          staff.post("/settings/new-run",
                      data={"prefix": "XX", "confirm": "yes"}).status_code == 302)
    admin.post("/settings/new-run", data={"prefix": "XX"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("changing it without ticking the confirmation does nothing",
          db.next_serial_preview(conn) == "FZ-00004")
    conn.close()

    admin.post("/settings/new-run", data={"prefix": "XX-", "confirm": "yes"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("an admin can set the numbering from the settings page",
          db.next_serial_preview(conn) == "XX-00001")
    check("the settings page shows the next number",
          b"XX-00001" in admin.get("/settings").data)

    admin.post("/settings/new-run",
               data={"prefix": "XX-", "start_at": "250", "confirm": "yes"})
    check("a starting number typed on its own is used",
          db.next_serial_preview(db.connect(flask_app.config["DB_PATH"])) == "XX-00250")
    conn.close()

    check("a non-admin does not see the numbering control",
          b"Set Numbering" not in staff.get("/settings").data)


def test_sequence_derives_from_the_book(tmpdir):
    """The next number is read from the passes themselves, so it can never drift
    out of step with what was actually issued."""
    conn = db.connect(Path(tmpdir) / "derived.db")

    check("an empty book starts at 1", db.next_seq(conn, "FZ") == 1)
    for i in range(1, 6):
        gp = db.create_gate_pass(conn, None, "S", "C", f"I{i}", "04-08-2026", "",
                                  sample_items(), prepared_by="Ravi Kumar")
        check(f"pass {i} is numbered {i}", gp["serial_seq"] == i)

    check("the next number follows the highest issued", db.next_seq(conn, "FZ") == 6)

    # Cancelling keeps the number, so the next pass must not reuse it.
    db.cancel_gate_pass(conn, db.get_gate_pass_by_serial(conn, "FZ-00005")["id"], "test")
    check("a cancelled pass still holds its number", db.next_seq(conn, "FZ") == 6)
    after_cancel = db.create_gate_pass(conn, None, "S", "C", "I6", "04-08-2026", "",
                                        sample_items(), prepared_by="Ravi Kumar")
    check("the run continues past a cancelled number", after_cancel["serial_seq"] == 6)

    # A different run counts separately.
    check("a fresh prefix starts again at 1", db.next_seq(conn, "XX") == 1)
    db.close(conn)


def test_batch_issue_keeps_the_run_unbroken(tmpdir):
    """50 invoices at once, some of them unreadable."""
    flask_app, client = logged_in_app(tmpdir, "batch_storage")
    conn = db.connect(flask_app.config["DB_PATH"])

    good_ids, bad_ids = [], []
    for i in range(45):
        good_ids.append(db.create_draft(
            conn, supplier_name="Golden Touch Exports", customer_name="FANZART LLP",
            invoice_no=f"FR 2627{i:05d}", invoice_date="04-08-2026",
            items=[{"item_name": "MICRON MODREN OAK", "quantity": "1"}]))
    # Five that could not be read — the scanned-PDF case.
    for i in range(5):
        bad_ids.append(db.create_draft(
            conn, parse_notes="this PDF has no text in it", items=[]))

    check("nothing has taken a number yet", db.count_gate_passes(conn) == 0)
    check("a readable draft is ready",
          db.draft_problem(db.get_draft(conn, good_ids[0])) is None)
    check("an unreadable draft is held back",
          db.draft_problem(db.get_draft(conn, bad_ids[0])) is not None)

    issued, skipped = db.create_gate_passes_batch(
        conn, good_ids + bad_ids, prepared_by="Ravi Kumar", prepared_by_user_id=1)

    check("every readable invoice was issued", len(issued) == 45)
    check("every unreadable invoice was skipped", len(skipped) == 5)
    seqs = [gp["serial_seq"] for gp in issued]
    check("the batch is numbered as one unbroken block", seqs == list(range(1, 46)))
    check("no number was handed out twice", len(set(seqs)) == 45)
    check("the failures consumed no numbers", db.next_seq(conn, "FZ") == 46)
    check("the failed drafts are still there to fix",
          all(db.get_draft(conn, i) is not None for i in bad_ids))
    check("the issued drafts are gone",
          all(db.get_draft(conn, i) is None for i in good_ids))

    # A second batch carries straight on from the first.
    more = [db.create_draft(conn, supplier_name="S", customer_name="C",
                             invoice_no=f"B2-{i}", invoice_date="04-08-2026",
                             items=[{"item_name": "Fan", "quantity": "1"}])
            for i in range(10)]
    issued2, _ = db.create_gate_passes_batch(conn, more, prepared_by="Ravi Kumar")
    check("the next batch continues the run",
          [gp["serial_seq"] for gp in issued2] == list(range(46, 56)))

    everything = sorted(gp["serial_seq"] for gp in db.list_gate_passes(conn, limit=None))
    check("the whole book has no gaps", everything == list(range(1, 56)))
    db.close(conn)


def test_batch_rolls_back_as_one(tmpdir):
    """If one pass in a batch cannot be written, none of them are — the run is
    never left with a hole in the middle."""
    conn = db.connect(Path(tmpdir) / "batch_rollback.db")
    db.create_gate_pass(conn, None, "S", "C", "FIRST", "04-08-2026", "",
                         sample_items(), prepared_by="Ravi Kumar")

    draft_ids = [db.create_draft(conn, supplier_name="S",
                                  customer_name="EXPLODE" if i == 2 else "C",
                                  invoice_no=f"R-{i}", invoice_date="04-08-2026",
                                  items=[{"item_name": "Fan", "quantity": "1"}])
                 for i in range(5)]

    # Force a failure on the third pass of the batch. A duplicate serial cannot
    # be used to provoke this: the number comes from MAX(serial_seq) + 1, so it
    # is unique by construction — which is the point of deriving it from the
    # book. A trigger is the only way to make a mid-batch insert fail.
    conn.execute("""
        CREATE TRIGGER trg_test_explode BEFORE INSERT ON gate_passes
        FOR EACH ROW WHEN NEW.customer_name = 'EXPLODE'
        BEGIN SELECT RAISE(ABORT, 'boom'); END""")

    before = db.count_gate_passes(conn)
    check("the batch fails rather than issuing a partial run",
          raises(sqlite3.IntegrityError, db.create_gate_passes_batch, conn, draft_ids,
                 "Ravi Kumar"))
    conn.execute("DROP TRIGGER trg_test_explode")
    check("no pass from the failed batch was written",
          db.count_gate_passes(conn) == before)
    check("every draft survives for a retry",
          all(db.get_draft(conn, i) is not None for i in draft_ids))
    check("the connection is usable again", not conn.in_transaction)
    db.close(conn)


def test_two_batches_at_once(tmpdir):
    """Two people batch-uploading simultaneously must not interleave numbers."""
    import threading

    db_path = Path(tmpdir) / "batch_concurrent.db"
    setup = db.connect(db_path)
    batches = []
    for who in range(4):
        batches.append([db.create_draft(
            setup, supplier_name="S", customer_name=f"C{who}",
            invoice_no=f"{who}-{i}", invoice_date="04-08-2026",
            items=[{"item_name": "Fan", "quantity": "1"}]) for i in range(15)])
    setup.close()

    results, errors = {}, []
    barrier = threading.Barrier(4)

    def run(who):
        conn = db.connect(db_path)
        try:
            barrier.wait(timeout=20)
            issued, _ = db.create_gate_passes_batch(
                conn, batches[who], prepared_by=f"Staff {who}")
            results[who] = [gp["serial_seq"] for gp in issued]
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{type(exc).__name__}: {exc}")
        finally:
            db.close(conn)

    threads = [threading.Thread(target=run, args=(w,)) for w in range(4)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=60)

    check("no batch failed under contention", not errors)
    check("all four batches completed", len(results) == 4)
    allocated = sorted(n for seqs in results.values() for n in seqs)
    check("60 numbers were handed out", len(allocated) == 60)
    check("none was handed out twice", len(set(allocated)) == 60)
    check("the run has no gaps", allocated == list(range(1, 61)))
    check("each batch got a contiguous block of its own",
          all(sorted(seqs) == list(range(min(seqs), max(seqs) + 1))
              for seqs in results.values()))


@needs_fixtures
def test_batch_upload_over_http(tmpdir):
    """Uploading several PDFs in one go, the way the form posts them."""
    flask_app, client = logged_in_app(tmpdir, "batch_http")

    uploads = []
    for name in (SAMPLE_INVOICE, TWO_PAGE_INVOICE, SERVICE_LINE_INVOICE, SCANNED_INVOICE):
        uploads.append((io.BytesIO(name.read_bytes()), name.name))
    resp = client.post("/upload", data={"invoice": uploads},
                        content_type="multipart/form-data")
    check("a batch upload lands on the drafts screen",
          resp.headers.get("Location", "").endswith("/drafts"))

    conn = db.connect(flask_app.config["DB_PATH"])
    check("one draft per file", len(db.list_drafts(conn)) == 4)
    check("still nothing numbered", db.count_gate_passes(conn) == 0)
    conn.close()

    page = client.get("/drafts")
    check("the drafts screen flags the unreadable one", b"Needs attention" in page.data)
    check("the drafts screen marks the readable ones ready", b"Ready" in page.data)

    conn = db.connect(flask_app.config["DB_PATH"])
    all_ids = [d["id"] for d in db.list_drafts(conn)]
    conn.close()
    resp = client.post("/drafts/issue", data={"draft_id": [str(i) for i in all_ids]},
                        follow_redirects=True)

    conn = db.connect(flask_app.config["DB_PATH"])
    issued = db.list_gate_passes(conn, limit=None)
    check("the three readable invoices were issued", len(issued) == 3)
    check("they are numbered 1, 2, 3",
          sorted(gp["serial_seq"] for gp in issued) == [1, 2, 3])
    check("the unreadable one is still a draft", len(db.list_drafts(conn)) == 1)
    check("the operator is told nothing was wasted", b"No number was used" in resp.data)
    conn.close()

    check("issuing nothing selected is refused",
          b"Select at least one" in client.post("/drafts/issue", data={},
                                                 follow_redirects=True).data)


def _pass_issued_on(conn, day, invoice_no, prepared_by="Ravi Kumar"):
    """A pass back-dated to a given day, for exercising the date filters."""
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                              invoice_no, "04-08-2026", "", sample_items(),
                              prepared_by=prepared_by)
    conn.execute("UPDATE gate_passes SET issued_at = ? WHERE id = ?",
                 (f"{day} 10:30:00", gp["id"]))
    return db.get_gate_pass(conn, gp["id"])


def test_long_invoice_is_one_pass_over_several_pages(tmpdir):
    """A 32-item invoice is one gate pass with one number, printed over 2 pages.

    Not two gate passes, and not a suffixed number: the same shipment leaving
    under two serials is exactly what the unbroken run exists to prevent.
    """
    check("pages are counted per 26 items",
          [db.page_count(n) for n in (0, 1, 26, 27, 32, 52, 53)] == [1, 1, 1, 2, 2, 2, 3])

    paged = db.paginate_items([{"item_name": f"Item {i}"} for i in range(1, 33)])
    check("32 items become 2 pages", len(paged) == 2)
    check("the first page holds 26", len(paged[0]) == 26)
    check("the second page holds the remaining 6", len(paged[1]) == 6)
    check("numbering runs straight through rather than restarting",
          paged[0][0][0] == 1 and paged[0][-1][0] == 26
          and paged[1][0][0] == 27 and paged[1][-1][0] == 32)
    check("an empty pass still yields one page", db.paginate_items([]) == [[]])

    flask_app, client = logged_in_app(tmpdir, "multipage")
    conn = db.connect(flask_app.config["DB_PATH"])
    items = [{"item_name": f"VIENNA VARIANT {i:02d}", "quantity": str(i)}
             for i in range(1, 33)]
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                              "FR 262702140", "31-07-2026", "", items,
                              prepared_by="Ravi Kumar")
    conn.close()

    check("a 32-item invoice issues rather than being refused", gp is not None)
    check("it holds all 32 items", len(gp["items"]) == 32)
    check("it took exactly one serial number", gp["serial_no"].count("-") == 1)
    check("the serial carries no page suffix",
          not gp["serial_no"].endswith("-1") and not gp["serial_no"].endswith("-2"))

    page = client.get(f"/print/{gp['id']}").data.decode()

    check("two sheets are printed", page.count('class="sheet') == 2)
    check("each sheet still carries two copies to cut apart",
          page.count('<div class="pass">') == 4)
    check("every page shows the same serial number",
          page.count(gp["serial_no"]) >= 4)
    check("page 1 of 2 is labelled", "Page 1 of 2" in page)
    check("page 2 of 2 is labelled", "Page 2 of 2" in page)
    check("there is no page 3", "Page 3 of" not in page)

    check("the first item is on page one", "VIENNA VARIANT 01" in page)
    check("the last item is on page two", "VIENNA VARIANT 32" in page)
    check("item 27 carries its running number, not a restart",
          '<td class="col-sl">27</td>' in page)
    check("the header repeats on both pages",
          page.count("FR 262702140") >= 4 and page.count("FANZART LLP") >= 4)
    check("both pages break onto their own sheet", "page-break-after" in
          (ROOT / "static" / "css" / "print.css").read_text())

    # A short pass is unchanged: one sheet, no page label.
    conn = db.connect(flask_app.config["DB_PATH"])
    short = db.create_gate_pass(conn, None, "S", "C", "SHORT", "05-08-2026", "",
                                 sample_items(), prepared_by="Ravi Kumar")
    conn.close()
    short_page = client.get(f"/print/{short['id']}").data.decode()
    check("a single-page pass prints one sheet", short_page.count('class="sheet') == 1)
    check("a single-page pass shows no page indicator", "Page 1 of" not in short_page)


def test_review_screen_explains_multiple_pages(tmpdir):
    """The review screen informs rather than blocks."""
    flask_app, client = logged_in_app(tmpdir, "multipage_ui")
    conn = db.connect(flask_app.config["DB_PATH"])
    draft_id = db.create_draft(
        conn, supplier_name="Golden Touch Exports", customer_name="FANZART LLP",
        invoice_no="FR 262702140", invoice_date="31-07-2026",
        items=[{"item_name": f"Item {i}", "quantity": "1"} for i in range(32)])
    check("a 32-item draft is ready to issue, not held back",
          db.draft_problem(db.get_draft(conn, draft_id)) is None)
    conn.close()

    page = client.get(f"/review/{draft_id}").data
    check("the review screen carries the page notice", b'id="page-note"' in page)
    check("the notice is informational, not an error", b'class="notice info"' in page)
    check("the old blocking wording is gone", b"split this into two" not in page)
    check("the screen knows the page size", b"ITEMS_PER_PAGE = 26" in page)

    check("the drafts screen offers it for issuing",
          b"Needs attention" not in client.get("/drafts").data)

    resp = client.post(f"/review/{draft_id}", data=_issue_form(32))
    check("issuing a 32-item draft goes straight to print",
          "/print/" in resp.headers.get("Location", ""))


def _issue_form(count):
    form = MultiDict([("supplier_name", "Golden Touch Exports"),
                       ("customer_name", "FANZART LLP"),
                       ("invoice_no", "FR 262702140"),
                       ("invoice_date", "31-07-2026"), ("vehicle_no", "")])
    for i in range(1, count + 1):
        form.add("item_name", f"Item {i}")
        form.add("quantity", "1")
        form.add("cartons", "")
    form.add("action", "issue")
    return form


def test_print_layout_rules(tmpdir):
    """The printed pass has a few proportions that must not drift.

    These are asserted against the stylesheet rather than a rendered page —
    there is no browser here — so they are a guard against someone changing a
    number without re-rendering, not a substitute for looking at it.
    """
    css = (ROOT / "static" / "css" / "print.css").read_text()
    card = (ROOT / "templates" / "_pass_card.html").read_text()

    check("the item column is headed 'Item Name'",
          '<th class="col-item">Item Name</th>' in card)

    # Label and colon are separate now, so the colons can be aligned — assert
    # on the labels rather than on a flat "Label :" string.
    # Colons line up down each block because the label sits in a fixed-width
    # span. The colon must stay OUTSIDE that span: inside it, it follows the
    # label text and moves with its length, which is the thing being fixed.
    for label in (">Serial No.</span>:", ">Issue Date</span>:",
                  ">Document No</span>:", ">Date</span>:"):
        check(f"the colon sits outside the label span ({label[1:-8]})", label in card)
    key = re.search(r"\n\.key \{[^}]*\}", css)
    key_serial = re.search(r"\.key-serial \{[^}]*\}", css)
    check("labels are fixed-width so the colons align",
          key is not None and "display: inline-block" in key.group()
          and "min-width" in key.group())
    check("each block sizes to its own longest label",
          key_serial is not None and "min-width" in key_serial.group())
    # min-width, not width: a longer label should push the colon right rather
    # than overlap it.
    check("a longer label pushes the colon rather than overlapping it",
          key is not None and re.search(r"[^-]width: [\d.]+mm", key.group()) is None)
    # The block is left-aligned inside but still flush right on the page —
    # .pass-head is space-between. Right-aligning each line would put the colons
    # wherever the values happened to end.
    serial_css = re.search(r"\n\.serial \{[^}]*\}", css).group()
    check("the serial block is left-aligned inside", "text-align: left" in serial_css)

    check("the serial block carries the issue date under the number",
          ">Serial No.</span>" in card and ">Issue Date</span>" in card
          and card.index(">Serial No.</span>") < card.index(">Issue Date</span>"))
    # From the record, never from the clock. A pass is a document: reprinting
    # it next year must produce the sheet it produced the day it was raised.
    # Jinja comments stripped first: the note explaining this decision talks
    # about "today's date", and matching that would fail on the explanation.
    card_code = re.sub(r"\{#.*?#\}", "", card, flags=re.S)
    check("the issue date comes off the pass, not from datetime.now()",
          "gate_pass.issued_at|as_day_month_year" in card_code
          and "now()" not in card_code)

    # Document No above Date in one box, Remarks matching it on the right.
    check("Document No and Date share the left box",
          ">Document No</span>" in card and ">Date</span>" in card
          and card.index(">Document No</span>") < card.index("Remarks :"))
    check("Remarks is the right-hand box", 'class="remarks-cell"' in card)
    # Prints what was typed on the review screen, and stays empty when nothing
    # was — so the box is still there to write in by hand at the gate.
    check("Remarks prints what was typed, and nothing when it was not",
          '>Remarks :</span>' in card
          and '<span class="remarks-body">{% if gate_pass.remarks %}' in card)
    tall = re.search(r"table\.meta-box-tall td \{[^}]*\}", css)
    check("that box is given room", tall is not None and "min-height" in tall.group())

    # The header taking a second line is what forced ITEM_BODY_MM down. A full
    # page has to still fit: 26 rows at their floor, plus the header, plus the
    # footer's signing room, inside 196mm of pass.
    floor_mm = 0.85 * 2 + (7 * 1.2) / 72 * 25.4      # padding + one 7pt line
    check("a full page of rows still fits the item body",
          db.ITEMS_PER_PAGE * floor_mm <= db.ITEM_BODY_MM)
    check("and the row height asked for is above the floor",
          db.print_row_height_mm(db.ITEMS_PER_PAGE) >= floor_mm)
    # Every budgeted row is rendered, so the column dividers run the full height
    # of the box like the printed form; only the rows carrying an item are
    # tagged, and only that tag draws a horizontal rule.
    check("every budgeted row is drawn, so the box is full height",
          "{% for r in range(1, rows + 1) %}" in card)
    check("only rows with an item are tagged item-row",
          '<tr{% if entry %} class="item-row"{% endif %}>' in card)
    rule = re.search(r"table\.items-table tbody tr\.item-row td \{[^}]*\}", css)
    check("the horizontal rule is drawn per item row, not on the table",
          rule is not None and "border-bottom: 1px solid #000" in rule.group())
    check("blank rows are never ruled",
          re.search(r"tbody tr:not\(\.item-row\)[^}]*border-bottom:\s*1px", css) is None)
    # The dividers come from border-left on every cell, which is what carries
    # them past the last item to the bottom of the box.
    check("column dividers are on the cells, so they run the full height",
          re.search(r"table\.items-table th,\s*table\.items-table td \{[^}]*border-left: 1px solid #000",
                    css) is not None)

    def width_of(selector):
        match = re.search(rf"\.{selector} \{{[^}}]*width: ([\d.]+)%", css)
        return float(match.group(1)) if match else None

    # Locked deliberately — the heading text, not the digits, is what sets the
    # numeric column widths. They were rebalanced from 8/68/12/12 when the 15mm
    # punch margin took 10mm out of every pass and "No.of Cartons" no longer
    # fit inside its column.
    for column, expected in (("col-sl", 7), ("col-item", 66),
                              ("col-qty", 13), ("col-ctn", 14)):
        check(f"{column} is locked at {expected}%", width_of(column) == expected)
    check("the four columns add up to the full width",
          sum(width_of(c) for c in
              ("col-sl", "col-item", "col-qty", "col-ctn")) == 100)

    header = re.search(r"table\.items-table thead th \{[^}]*\}", css).group()
    padding = float(re.search(r"padding: ([\d.]+)mm", header).group(1))
    heading_size = float(re.search(r"font-size: ([\d.]+)pt", header).group(1))
    check("the header row padding gives it breathing room", 0.5 <= padding <= 1.0)
    check("headings never wrap, which would make the header taller",
          "white-space: nowrap" in header)
    # At 12% the Cartons column is 15.7mm. "No.of Cartons" needs 16.1mm at
    # 6.5pt and spills past its border, because the heading is nowrap and so
    # overflows rather than wrapping. 6.2pt is what makes the locked width fit.
    check("the heading is small enough for the longest one to fit 12%",
          heading_size <= 6.2)

    body = re.search(r"table\.items-table tbody td \{[^}]*\}", css).group()
    check("item rows have comfortable leading", "line-height: 1.2" in body)
    # 0.85mm vertical is the row's hard floor — a row cannot be shorter than
    # its padding plus its line box. It came down from 1mm when the header grew
    # a line and 26 rows no longer fit. The 2mm sides keep text off the borders.
    check("item rows have padding that keeps text off the borders",
          "padding: 0.85mm 2mm" in body)
    check("item text is centred in its row", "vertical-align: middle" in body)
    # The row height is now explicit rather than shared out by flex, so the same
    # pass prints identically whatever the paper size or print scale.
    check("the row height is an exact value from the server",
          "var(--row-h" in body)
    check("a row can never grow past its stated height",
          "max-height: var(--row-h" in body and "overflow: hidden" in body)
    check("the table no longer stretches with the page",
          "flex: none !important" in css)
    check("the side padding rule does not reset the vertical padding",
          "padding-left: 2mm" in css and "padding: 0 1mm" not in css)
    # Row spacing is set by how many rows share the fixed box, not by padding:
    # the padding only sets a floor the rows sit well above.
    check("a short pass is spaced by the row count, not the padding",
          db.PRINT_MIN_ROWS == 22)

    # The page margin is the band the browser draws its own header and footer
    # into — title, URL, page number, timestamp — which is exactly the chrome
    # that was landing in the corners of printed passes. Zero margin, and the
    # 7mm carried as padding on a full-sheet .sheet instead, so the content box
    # is still the 283x196mm every other number here assumes.
    for name in ("print.html", "print_batch.html"):
        # Matched to end of line, not to the closing brace: the rule contains a
        # Jinja `{{ ... }}` for the paper size, so a brace-counting pattern
        # stops inside it and the check silently passes on nothing.
        page_rule = re.search(r"@page .*", (ROOT / "templates" / name).read_text()).group()
        check(f"{name} leaves no margin for the browser to print into",
              "margin: 0;" in page_rule)

    sheet = re.search(r"\n\.sheet \{[^}]*\}", css).group()
    check("the sheet is the whole A4 landscape page",
          "width: 297mm" in sheet and "height: 210mm" in sheet)
    check("the page margin is carried as padding inside the sheet",
          "padding: 7mm" in sheet)
    check("an A5 sheet is a whole A5 page",
          "width: 148mm" in re.search(r"\.sheet\.single \{[^}]*\}", css).group())
    check("border-box keeps the content box at 283x196mm",
          "* { box-sizing: border-box; }" in css)

    # The item column is the odd one out: heading centred like the others, but
    # the names underneath left-aligned and indented so they read as a list.
    check("the item heading is centred",
          re.search(r"\.col-item \{[^}]*text-align: center", css) is not None)
    item_cell = re.search(r"table\.items-table td\.col-item \{[^}]*\}", css)
    check("the item names are left-aligned",
          item_cell is not None and "text-align: left" in item_cell.group())
    # !important is required: the tbody rule sets `padding: 1mm 2mm !important`
    # and an ordinary padding-left loses to it, measuring 2mm and doing nothing.
    check("and indented off the column divider",
          item_cell is not None and "padding-left: 2mm !important" in item_cell.group())
    # Load-bearing, not cosmetic: max-height does not constrain a table cell, so
    # a name that wraps makes its row taller and enough of those push the last
    # item off the page — silently, because the pass is overflow: hidden.
    check("an item name can never wrap and grow its row",
          "white-space: nowrap" in item_cell.group())
    check("a name too long for its column is visibly cut",
          "text-overflow: ellipsis" in item_cell.group())
    # Comments stripped first: this rule's own comment quotes the shorthand it
    # is warning about, and matching that would fail on the explanation rather
    # than on the code.
    item_decls = re.sub(r"/\*.*?\*/", "", item_cell.group(), flags=re.S) if item_cell else ""
    check("the indent does not reset the vertical padding",
          re.search(r"\bpadding:\s", item_decls) is None)
    for column in ("col-sl", "col-qty", "col-ctn"):
        check(f"{column} stays centred",
              re.search(rf"\.{column} \{{[^}}]*text-align: center", css) is not None)
    # Both names live in one footer so they share a baseline across the foot of
    # the pass. "Prepared by" used to be a separate line below it, sitting
    # 2.88mm lower and eating a row of the space the authoriser signs in.
    foot = re.search(r"\.pass-foot \{[^}]*\}", css).group()
    # Pinned by ONE auto margin on the wrapper that holds the totals and the
    # signatures together. Two — one on each — divide the leftover space
    # between them and leave the totals floating mid-page.
    check("the end block is pinned to the foot of the pass",
          "margin-top: auto" in re.search(r"\.pass-end \{[^}]*\}", css).group())
    check("and the footer no longer carries an auto margin of its own",
          "margin-top: auto" not in foot)
    check("both names sit on one baseline",
          "align-items: flex-end" in foot and "justify-content: space-between" in foot)
    # min-height is the signing room: everything inside is bottom-aligned, so
    # the height above the rule is empty space to write in. Measured at 15mm:
    # 13.26mm between the table and the rule, against 10.38mm before.
    signing_room = float(re.search(r"min-height: ([\d.]+)mm", foot).group(1))
    check("there is room above the rule to sign", signing_room >= 14)
    # The totals used to be a strip between the table and the signatures, and
    # its height was taken from here — 15mm dropped to 11.8mm on any page that
    # showed them. They are a row inside the table now, paid for out of the
    # table's own budget, so every pass gets the full signing room again.
    check("no page has its signing room cut for the totals",
          not re.search(r"\.pass\.with-totals \.pass-foot", css))
    check("and the totals row is paid for out of the table instead",
          db.print_row_height_mm(26, True) * 26 + db.TOTALS_ROW_MM <= db.ITEM_BODY_MM)
    check("Prepared by is inside the footer, not a line below it",
          card.index('class="prepared-by"') > card.index('class="pass-foot"')
          and card.index('class="prepared-by"') < card.index('class="foot-side'))

    # Both halves need a punch margin: the sheet is cut down the middle, so the
    # right-hand pass's left edge is the cut line, not the paper edge.
    pass_css = re.search(r"\n\.pass \{[^}]*\}", css).group()
    punch = re.search(r"padding: [\d.]+mm [\d.]+mm [\d.]+mm ([\d.]+)mm", pass_css)
    check("each pass carries a punch margin on its left",
          punch is not None and float(punch.group(1)) >= 15)
    check("and it is on .pass, so both halves get one, not on the sheet",
          "padding" in pass_css)

    prepared_css = re.search(r"\.prepared-by \{[^}]*\}", css).group()
    check("a long name cannot squeeze the signature block",
          "flex: none" in prepared_css and "max-width" in prepared_css)
    check("and no margin knocks it off the shared baseline",
          "margin-top" not in prepared_css)

    # The document title, set the way an official form is set.
    title = re.search(r"\.pass-head h1 \{[^}]*\}", css).group()
    check("the title is uppercase", "text-transform: uppercase" in title)
    check("the title is heavy",
          int(re.search(r"font-weight: (\d+)", title).group(1)) >= 700)
    check("the title is tracked", "letter-spacing" in title)
    check("the title never breaks across two lines", "white-space: nowrap" in title)
    # Nothing is drawn under the title. The weight and tracking already mark it
    # as the title, and a rule there only competes with the boxed rows below.
    check("no underline under the title", "text-decoration: underline" not in title)
    check("and no rule either", "border-bottom" not in title)
    # letter-spacing adds a gap after the FINAL letter too, so the text box is
    # wider than the glyphs and a centred title sits left of centre.
    check("the trailing letter-space is cancelled", "margin-right: -" in title)

    for face in ("Inter", "Roboto", "Montserrat"):
        check(f"{face} is offered for the title", face in title)
    check("and there is a fallback the machine is certain to have",
          "Arial" in title and "sans-serif" in title)
    # A print stylesheet cannot afford a webfont that may or may not arrive:
    # the office server has no guaranteed route to the internet, and a font
    # swapping in mid-print changes the width of a title centred between the
    # logo and the serial number. Measured across every fallback, the title is
    # 26.2mm (Inter) to 28.12mm (Arial) with ~26mm clear on each side.
    check("no webfont is fetched over the network",
          "@import" not in css and "fonts.googleapis" not in css
          and "@font-face" not in css)

    prepared = re.search(r"\.prepared-by \{[^}]*\}", css).group()
    check("Prepared by is small", "font-size: 5.5pt" in prepared)
    check("Prepared by is muted grey", "#888888" in prepared)
    check("Prepared by is uppercase with tracking",
          "text-transform: uppercase" in prepared and "letter-spacing" in prepared)

    # And on the page itself: one signature rule, for the authoriser only.
    flask_app, client = logged_in_app(tmpdir, "print_layout")
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "MR Sample Buyer",
                              "FR 262702176", "05-08-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")
    conn.close()
    page = client.get(f"/print/{gp['id']}").data

    check("Prepared by names the person who issued it", b"Prepared by: Ravi Kumar" in page)
    # The pass is `overflow: hidden`, so the footer must sit comfortably inside
    # it rather than finishing flush with the edge. At ITEM_BODY_MM = 142.8 the
    # line ended 0.15mm inside: Chrome fit it, Firefox clipped it in half.
    body_and_footer = db.ITEM_BODY_MM + 45      # header, meta boxes, foot, prepared-by
    check("the item box leaves room for the footer below it",
          body_and_footer <= 188)               # 196mm pass less 4mm padding each side
    check("Prepared by appears once per A5 half",
          page.count(b'class="prepared-by"') == 2)
    check("Authorised by is still on the pass", b"Authorised by" in page)
    check("only the authoriser gets a signature rule",
          page.count(b'class="sign-line"') == 2)
    check("the old signature-name block is gone", b'class="sign-name"' not in page)


def test_timestamps_are_local_not_utc(tmpdir):
    """Every timestamp is the office wall clock.

    SQLite's datetime('now') is UTC. Relying on it put issued_at 5.5 hours
    behind in India: a pass created at 02:37 on the 5th was filed as 21:07 on
    the 4th, and a report for "today" missed everything issued after 18:30.
    """
    from datetime import datetime as _dt

    conn = db.connect(Path(tmpdir) / "clock.db")

    def drift(stamp):
        return abs((_dt.now() - _dt.strptime(stamp, "%Y-%m-%d %H:%M:%S")).total_seconds())

    single = db.create_gate_pass(conn, None, "S", "C", "ONE", "05-08-2026", "",
                                  sample_items(), prepared_by="Ravi Kumar")
    check("a pass records the local time it was issued", drift(single["issued_at"]) < 120)

    draft_id = db.create_draft(conn, supplier_name="S", customer_name="C",
                                invoice_no="TWO", invoice_date="05-08-2026",
                                items=[{"item_name": "Fan", "quantity": "1"}])
    check("a draft records the local time it was uploaded",
          drift(db.get_draft(conn, draft_id)["created_at"]) < 120)

    batch, _ = db.create_gate_passes_batch(conn, [draft_id], prepared_by="Ravi Kumar")
    check("a batch-issued pass records local time too",
          drift(batch[0]["issued_at"]) < 120)

    user_id = make_user(conn, "clock", "Clock Tester")
    check("an account records the local time it was created",
          drift(db.get_user(conn, user_id)["created_at"]) < 120)

    db.cancel_gate_pass(conn, single["id"], "test", cancelled_by="Ravi Kumar")
    check("a cancellation records local time",
          drift(db.get_gate_pass(conn, single["id"])["cancelled_at"]) < 120)

    # The filters agree with the clock: a pass issued now is in "today".
    today = _dt.now().date().isoformat()
    check("a pass issued now falls inside today's filter",
          db.count_gate_passes(conn, date_from=today, date_to=today) == 2)
    db.close(conn)


def test_upgrade_converts_old_utc_timestamps(tmpdir):
    """An existing book was written with UTC timestamps and has to be corrected
    in place — and correcting it must happen exactly once."""
    path = Path(tmpdir) / "utc_book.db"
    old = sqlite3.connect(path)
    old.executescript("""
        CREATE TABLE gate_passes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            serial_no TEXT NOT NULL UNIQUE, serial_scope TEXT NOT NULL,
            serial_seq INTEGER NOT NULL, status TEXT NOT NULL DEFAULT 'issued',
            supplier_name TEXT NOT NULL DEFAULT '', customer_name TEXT NOT NULL DEFAULT '',
            invoice_no TEXT NOT NULL DEFAULT '', invoice_date TEXT NOT NULL DEFAULT '',
            vehicle_no TEXT NOT NULL DEFAULT '', invoice_pdf_path TEXT,
            issued_at TEXT NOT NULL, cancelled_at TEXT, cancel_reason TEXT);
        INSERT INTO gate_passes (serial_no, serial_scope, serial_seq, invoice_no, issued_at)
        VALUES ('FZ-00001', 'FZ', 1, 'OLD', '2026-08-04 21:07:17');
        PRAGMA user_version = 4;
    """)
    old.commit()
    old.close()

    conn = db.connect(path)
    corrected = db.get_gate_pass_by_serial(conn, "FZ-00001")["issued_at"]
    check("the old UTC timestamp was moved into local time",
          corrected != "2026-08-04 21:07:17")
    check("the pass itself is untouched otherwise",
          db.get_gate_pass_by_serial(conn, "FZ-00001")["invoice_no"] == "OLD")
    check("the file is marked as upgraded",
          conn.execute("PRAGMA user_version").fetchone()[0] == db.SCHEMA_VERSION)
    conn.close()

    # Re-opening must not shift it a second time.
    again = db.connect(path)
    check("re-opening does not shift the timestamp again",
          db.get_gate_pass_by_serial(again, "FZ-00001")["issued_at"] == corrected)
    again.close()

    # A brand new database has nothing to correct.
    fresh = db.connect(Path(tmpdir) / "fresh_clock.db")
    gp = db.create_gate_pass(fresh, None, "S", "C", "NEW", "05-08-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")
    from datetime import datetime as _dt
    check("a fresh database is not shifted on creation",
          abs((_dt.now() - _dt.strptime(gp["issued_at"], "%Y-%m-%d %H:%M:%S")
               ).total_seconds()) < 120)
    db.close(fresh)


def test_date_filters(tmpdir):
    conn = db.connect(Path(tmpdir) / "dates.db")
    _pass_issued_on(conn, "2026-01-15", "JAN")
    _pass_issued_on(conn, "2026-07-31", "JUL")
    _pass_issued_on(conn, "2026-08-01", "AUG-1")
    _pass_issued_on(conn, "2026-08-05", "AUG-5")

    def invoices(**kw):
        return sorted(gp["invoice_no"] for gp in db.list_gate_passes(conn, limit=None, **kw))

    check("no dates means everything", len(invoices()) == 4)
    check("a single day is inclusive at both ends",
          invoices(date_from="2026-08-05", date_to="2026-08-05") == ["AUG-5"])
    check("a range covers its whole end day",
          invoices(date_from="2026-08-01", date_to="2026-08-05") == ["AUG-1", "AUG-5"])
    check("an open start reaches back to the beginning",
          invoices(date_to="2026-07-31") == ["JAN", "JUL"])
    check("an open end runs to now",
          invoices(date_from="2026-08-01") == ["AUG-1", "AUG-5"])
    check("a month range works", invoices(date_from="2026-08-01", date_to="2026-08-31")
          == ["AUG-1", "AUG-5"])
    check("a year range works", invoices(date_from="2026-01-01", date_to="2026-12-31")
          == ["AUG-1", "AUG-5", "JAN", "JUL"])
    check("a range with nothing in it returns nothing",
          invoices(date_from="2026-03-01", date_to="2026-03-31") == [])
    check("dates combine with a search",
          invoices(date_from="2026-08-01", search="AUG-5") == ["AUG-5"])
    check("the count agrees with the list",
          db.count_gate_passes(conn, date_from="2026-08-01") == 2)
    check("dates combine with a status filter",
          db.count_gate_passes(conn, status="cancelled", date_from="2026-01-01") == 0)
    db.close(conn)


def test_quick_ranges_and_export_over_http(tmpdir):
    from datetime import date as _date, timedelta as _timedelta

    flask_app, client = logged_in_app(tmpdir, "export_storage")
    conn = db.connect(flask_app.config["DB_PATH"])
    today = _date.today()
    _pass_issued_on(conn, today.isoformat(), "TODAY")
    _pass_issued_on(conn, (today - _timedelta(days=200)).isoformat(), "LONG-AGO")
    conn.close()

    page = client.get("/register?range=today").data
    check("the Today filter keeps today's pass", b"TODAY" in page)
    check("the Today filter drops an older one", b"LONG-AGO" not in page)
    # The quick-range pills now live on Reports, but ?range= still filters the
    # register, so an old bookmark keeps working.
    check("a range still narrows the register", b"TODAY" in page)
    check("the reports page highlights the chosen preset",
          b"quick-range on" in client.get("/reports?range=today").data)
    check("all-time shows both", b"LONG-AGO" in client.get("/register").data)
    check("this year includes today's pass", b"TODAY" in client.get("/register?range=year").data)
    check("a nonsense date is ignored rather than trusted",
          client.get("/register?from=not-a-date").status_code == 200)

    summary = client.get("/reports/export?range=today")
    check("the export is a CSV download", "text/csv" in summary.headers["Content-Type"])
    check("the file is named for its range",
          "attachment" in summary.headers["Content-Disposition"]
          and today.isoformat() in summary.headers["Content-Disposition"])
    body = summary.data.decode("utf-8-sig")
    lines = [ln for ln in body.splitlines() if ln.strip()]
    check("the summary export has a header and one row per pass", len(lines) == 2)
    check("the header names the gate pass column", lines[0].startswith("Gate Pass No"))
    check("the export respects the date filter", "TODAY" in body and "LONG-AGO" not in body)
    check("the summary carries who prepared it", "Ravi Kumar" in body)
    check("the summary totals the quantities", "Total Quantity" in lines[0])

    detail = client.get("/reports/export?range=today&detail=1").data.decode("utf-8-sig")
    detail_lines = [ln for ln in detail.splitlines() if ln.strip()]
    check("the detailed export has a row per item", len(detail_lines) == 2)
    check("the detailed export names the item", "MICRON MODREN OAK" in detail)
    check("the detailed export has no totals column", "Total Quantity" not in detail_lines[0])

    everything = client.get("/reports/export").data.decode("utf-8-sig")
    check("an unfiltered export covers the whole book",
          "TODAY" in everything and "LONG-AGO" in everything)
    check("the export needs a login",
          flask_app.test_client().get("/reports/export").status_code == 302)


def test_register_page_is_just_search_and_filter(tmpdir):
    """The register is for reading the book. Exporting lives on Reports."""
    flask_app, client = logged_in_app(tmpdir, "register_ui")
    conn = db.connect(flask_app.config["DB_PATH"])
    _pass_issued_on(conn, "2026-08-05", "SHOWN")
    conn.close()

    page = client.get("/register").data
    check("there is one search box", page.count(b'name="q"') == 1)
    check("there is a Filter button", b'id="filter-toggle"' in page)
    check("the filter panel starts closed when nothing is filtered",
          b'id="filter-panel" hidden' in page)
    check("the export controls are gone from the register",
          b"Summary CSV" not in page and b"Detailed CSV" not in page)
    check("the quick-range pills are gone from the register",
          b"This week" not in page and b"All time" not in page)
    check("the table still lists the passes", b"SHOWN" in page)
    for column in (b"Sl. No.", b"Issued", b"Prepared by", b"From", b"Customer",
                    b"Invoice No", b"Status"):
        check(f"the table keeps its {column.decode()} column", column in page)
    check("the row actions are still there", b">Print<" in page and b">Cancel<" in page)

    # Filtering still works, and the panel opens itself when a filter is on.
    filtered = client.get("/register?status=cancelled").data
    check("the filter panel opens when a filter is active",
          b'id="filter-panel" hidden' not in filtered)
    check("the Filter button shows how many filters are on",
          b'class="count-badge"' in filtered)
    check("the status filter still applies", b"SHOWN" not in filtered)
    check("a date filter still applies",
          b"SHOWN" not in client.get("/register?from=2026-01-01&to=2026-01-31").data)
    check("searching still works", b"SHOWN" in client.get("/register?q=SHOWN").data)

    check("Reports is in the navigation", b'href="/reports"' in page)


def test_reports_page(tmpdir):
    flask_app, client = logged_in_app(tmpdir, "reports_ui")
    conn = db.connect(flask_app.config["DB_PATH"])
    _pass_issued_on(conn, "2026-08-05", "AUG")
    _pass_issued_on(conn, "2026-02-02", "FEB")
    conn.close()

    page = client.get("/reports").data
    check("the reports page loads", b"Export Report" in page)
    check("filters can be applied without exporting", b"Apply Filters" in page)
    for label in (b"Today", b"This week", b"This month", b"Custom"):
        check(f"the {label.decode()} preset is offered", label in page)
    check("status can be filtered", b'name="status"' in page)
    check("supplier can be filtered", b'name="supplier"' in page)
    check("customer can be filtered", b'name="customer"' in page)
    check("summary and detailed are offered", b'name="detail"' in page)
    check("csv and xlsx are offered", b'value="csv"' in page and b'value="xlsx"' in page)
    check("known suppliers are suggested", b"Golden Touch Exports" in page)
    # The match counter and its Update-count button were removed: the page is a
    # form for generating a report, not a dashboard.
    check("no match counter clutters the page", b"gate pass(es) match" not in page)
    check("no count button either", b"Update count" not in page)
    check("the action buttons sit under the format choices",
          page.index(b'value="xlsx"') < page.index(b"Apply Filters"))
    applied_page = client.get("/reports?status=").data
    check("the preview table follows the panel",
          applied_page.index(b"Apply Filters") < applied_page.index(b"Matching gate passes"))
    check("the reports page needs a login",
          flask_app.test_client().get("/reports").status_code == 302)

    # Nothing is listed until Apply Filters has been pressed — the page opens as
    # a form, not as a view of the whole book.
    fresh = client.get("/reports").data
    check("a bare visit lists no passes", b"Matching gate passes" not in fresh)
    check("and no table at all", b"serial-cell" not in fresh)
    check("it says what to do instead", b"press" in fresh and b"Apply Filters" in fresh)

    # Applying the filters is what fetches them.
    both = client.get("/reports?status=").data
    check("applying with nothing set lists the whole book",
          b"AUG" in both and b"FEB" in both)
    check("and shows the heading with a count", b"Matching gate passes" in both)
    filtered = client.get("/reports?from=2026-08-01").data
    check("applying a filter narrows the preview",
          b"AUG" in filtered and b"FEB" not in filtered)
    check("the count beside the heading follows the filter",
          b'<span class="count-badge neutral">1</span>' in filtered)
    check("an empty result says so rather than showing a bare table",
          b"Nothing matches these filters"
          in client.get("/reports?from=2020-01-01&to=2020-01-02").data)
    check("the preview is capped but the export is not",
          db.REPORT_PREVIEW_LIMIT < 1000)

    # Supplier / customer filters reach the export.
    by_customer = client.get("/reports/export?customer=FANZART").data.decode("utf-8-sig")
    check("filtering by customer narrows the export",
          "AUG" in by_customer and "FEB" in by_customer)
    none_match = client.get("/reports/export?customer=NOBODY").data.decode("utf-8-sig")
    check("a customer with no passes exports only a header",
          len([l for l in none_match.splitlines() if l.strip()]) == 1)


def test_xlsx_export(tmpdir):
    """The Excel download must be a real workbook, not a CSV named .xlsx."""
    from openpyxl import load_workbook

    flask_app, client = logged_in_app(tmpdir, "xlsx_export")
    conn = db.connect(flask_app.config["DB_PATH"])
    _pass_issued_on(conn, "2026-08-05", "FR-XLSX")
    conn.close()

    resp = client.get("/reports/export?format=xlsx")
    check("the workbook is served as an Excel file",
          resp.headers["Content-Type"].startswith(
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
    check("it downloads with an .xlsx name",
          resp.headers["Content-Disposition"].endswith('.xlsx"'))
    check("the body really is a zip container, as xlsx is",
          resp.data[:2] == b"PK")

    book = load_workbook(io.BytesIO(resp.data))
    sheet = book.active
    rows = list(sheet.iter_rows(values_only=True))
    check("the sheet is named for its contents", sheet.title == "Gate Passes")
    check("it has a header and one row per pass", len(rows) == 2)
    check("the header matches the CSV export", list(rows[0]) == db.EXPORT_SUMMARY_COLUMNS)
    check("the pass is in it", "FR-XLSX" in rows[1])
    check("the header row is frozen for scrolling", sheet.freeze_panes == "A2")
    check("the header is styled", sheet["A1"].font.bold)
    check("columns are given sensible widths",
          sheet.column_dimensions["A"].width == 16)
    check("counts arrive as numbers, not text",
          isinstance(rows[1][db.EXPORT_SUMMARY_COLUMNS.index("Items")], int))

    detailed = client.get("/reports/export?format=xlsx&detail=1")
    book2 = load_workbook(io.BytesIO(detailed.data))
    check("the detailed workbook is named for line items",
          book2.active.title == "Gate Pass Items")
    check("the detailed workbook uses the detail columns",
          list(next(book2.active.iter_rows(values_only=True))) == db.EXPORT_DETAIL_COLUMNS)

    # Both formats must describe the same report.
    csv_body = client.get("/reports/export?format=csv").data.decode("utf-8-sig")
    csv_rows = [ln for ln in csv_body.splitlines() if ln.strip()]
    check("CSV and Excel contain the same number of rows", len(csv_rows) == len(rows))

    empty = client.get("/reports/export?format=xlsx&from=2020-01-01&to=2020-01-02")
    check("an empty report still produces a valid workbook", empty.data[:2] == b"PK")
    check("the empty workbook still has its header",
          len(list(load_workbook(io.BytesIO(empty.data)).active.iter_rows())) == 1)


@needs_fixtures
def test_uploaded_invoices_are_deleted_once_used(tmpdir):
    """The upload is working material, not a record.

    Once its contents are on the gate pass there is nothing left to keep, and
    keeping them filled the disk at roughly 300 KB per invoice.
    """
    flask_app, client = logged_in_app(tmpdir, "invoice_cleanup")
    invoices = Path(flask_app.config["INVOICES_DIR"])

    def upload(fixture=SAMPLE_INVOICE):
        with open(fixture, "rb") as f:
            resp = client.post("/upload", data={"invoice": (f, fixture.name)},
                                content_type="multipart/form-data")
        return int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    def files_on_disk():
        return sorted(p.name for p in invoices.glob("*.pdf"))

    # An upload is kept while it is still a draft — it is what the operator is
    # working from.
    draft_id = upload()
    check("the upload is stored while the draft exists", len(files_on_disk()) == 1)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("the draft points at it",
          db.get_draft(conn, draft_id)["invoice_pdf_path"] is not None)
    conn.close()

    # Issuing it makes the file redundant.
    form = MultiDict([("supplier_name", "Golden Touch Exports"),
                       ("customer_name", "FANZART LLP"), ("invoice_no", "FR 262702171"),
                       ("invoice_date", "03-08-2026"), ("vehicle_no", ""),
                       ("item_name", "MICRON MODREN OAK"), ("quantity", "1"),
                       ("cartons", ""), ("action", "issue")])
    resp = client.post(f"/review/{draft_id}", data=form)
    gate_pass_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    check("issuing deletes the uploaded PDF", files_on_disk() == [])

    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.get_gate_pass(conn, gate_pass_id)
    check("the gate pass exists regardless", gp is not None)
    check("it keeps no path to the deleted file", gp["invoice_pdf_path"] is None)
    check("and keeps everything that was read off the invoice",
          gp["invoice_no"] == "FR 262702171"
          and gp["items"][0]["item_name"] == "MICRON MODREN OAK")
    conn.close()
    check("the pass still prints", client.get(f"/print/{gate_pass_id}").status_code == 200)

    # Discarding a draft takes its upload with it.
    discarded = upload()
    check("the discarded draft's upload is on disk", len(files_on_disk()) == 1)
    client.post(f"/drafts/{discarded}/delete")
    check("discarding a draft deletes its upload too", files_on_disk() == [])

    # A batch deletes only the uploads it actually issued.
    # Two DIFFERENT invoices: uploading one twice now blocks the batch, because
    # two drafts sharing a document number cannot both be issued.
    good = [upload(SAMPLE_INVOICE), upload(TWO_PAGE_INVOICE)]
    with open(SCANNED_INVOICE, "rb") as f:
        resp = client.post("/upload", data={"invoice": (f, "scan.pdf")},
                            content_type="multipart/form-data")
    unreadable = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    check("three uploads are waiting", len(files_on_disk()) == 3)

    client.post("/drafts/issue",
                 data={"draft_id": [str(i) for i in good + [unreadable]]})
    check("the batch deletes the uploads it issued", len(files_on_disk()) == 1)

    conn = db.connect(flask_app.config["DB_PATH"])
    still_a_draft = db.get_draft(conn, unreadable)
    check("the unreadable one is still a draft", still_a_draft is not None)
    check("and keeps its upload, because it still needs it",
          still_a_draft["invoice_pdf_path"] is not None)
    check("its file is the one left on disk",
          files_on_disk()[0] in still_a_draft["invoice_pdf_path"])
    conn.close()

    # A stored path is still checked before anything is unlinked.
    from app import _discard_invoice_file
    outside = Path(flask_app.config["STORAGE_DIR"]) / "gate_pass.db"
    check("a path outside the invoices folder is refused",
          _discard_invoice_file(flask_app, "../gate_pass.db") is False)
    check("and that file is untouched", outside.exists())
    check("an absent file is not an error",
          _discard_invoice_file(flask_app, "invoices/never-existed.pdf") is False)
    check("an empty path is not an error", _discard_invoice_file(flask_app, None) is False)


@needs_fixtures
def test_per_file_upload_endpoint(tmpdir):
    """The batch uploader posts one file at a time so it can show real progress."""
    flask_app, client = logged_in_app(tmpdir, "oneupload")

    with open(SAMPLE_INVOICE, "rb") as f:
        resp = client.post("/upload/one", data={"invoice": (f, "good.pdf")},
                            content_type="multipart/form-data")
    check("a readable PDF is accepted", resp.status_code == 200)
    body = resp.get_json()
    check("it reports success", body["ok"] is True)
    check("it returns the draft it made", isinstance(body["draft_id"], int))
    check("a readable invoice has no problem flagged", body["problem"] is None)
    check("the filename comes back for the progress list", body["filename"] == "good.pdf")

    with open(SCANNED_INVOICE, "rb") as f:
        scanned = client.post("/upload/one", data={"invoice": (f, "scan.pdf")},
                               content_type="multipart/form-data").get_json()
    check("a scan is still stored as a draft", scanned["ok"] is True)
    check("a scan is flagged as needing attention", scanned["problem"] is not None)
    check("a scan explains itself", "scan" in scanned["notes"].lower())

    bad = client.post("/upload/one", data={"invoice": (io.BytesIO(b"x"), "notes.txt")},
                       content_type="multipart/form-data")
    check("a non-PDF is refused", bad.status_code == 400)
    check("the refusal says why", bad.get_json()["error"] == "not a PDF")
    check("posting nothing is refused",
          client.post("/upload/one", data={}).status_code == 400)
    check("the endpoint needs a login",
          flask_app.test_client().post("/upload/one", data={}).status_code == 302)

    conn = db.connect(flask_app.config["DB_PATH"])
    check("two drafts exist and no numbers were used",
          len(db.list_drafts(conn)) == 2 and db.count_gate_passes(conn) == 0)
    conn.close()


@needs_fixtures
def test_upload_page_does_not_block_its_own_submit(tmpdir):
    """The picker holds the files in JavaScript, not in the file input.

    The input is cleared after each pick so choosing the same file twice still
    fires a change event — which leaves it empty. With `required` still on it,
    the browser's own validation refused to submit ("Please select one or more
    files") even though files were selected, and the page's submit handler never
    ran. The form must therefore turn native validation off and do the checking
    itself; `required` stays in the markup for the no-JavaScript path, where the
    input really does carry the files.
    """
    flask_app, client = logged_in_app(tmpdir, "upload_page")
    page = client.get("/upload").data.decode()

    check("the form turns off native validation for the JS path",
          "form.noValidate = true" in page)
    check("the file input keeps required for the no-JS path",
          'name="invoice"' in page and "required" in page)
    check("the input is cleared after a pick so re-picking still fires change",
          'input.value = ""' in page)
    check("the page checks for itself that something was chosen",
          "if (!chosen.length)" in page)
    check("there is a message for submitting with nothing chosen",
          'id="pick-error"' in page)

    # The no-JS path still has to work: a plain multi-file POST.
    with open(SAMPLE_INVOICE, "rb") as a, open(SERVICE_LINE_INVOICE, "rb") as b:
        resp = client.post("/upload",
                            data={"invoice": [(a, "one.pdf"), (b, "two.pdf")]},
                            content_type="multipart/form-data")
    check("posting the form without JavaScript still reads both files",
          resp.headers.get("Location", "").endswith("/drafts"))
    conn = db.connect(flask_app.config["DB_PATH"])
    check("both files became drafts", len(db.list_drafts(conn)) == 2)
    conn.close()


def test_connection_settings(tmpdir):
    """The PRAGMAs that keep four people out of each other's way."""
    path = Path(tmpdir) / "pragmas.db"
    conn = db.connect(path)

    def pragma(name):
        return conn.execute(f"PRAGMA {name}").fetchone()[0]

    check("WAL is on, so a reader never blocks the writer",
          pragma("journal_mode") == "wal")
    check("a writer waits rather than failing instantly",
          pragma("busy_timeout") == db.BUSY_TIMEOUT_MS)
    check("synchronous is NORMAL, the safe setting for WAL", pragma("synchronous") == 1)
    check("foreign keys are enforced", pragma("foreign_keys") == 1)
    check("the page cache is sized deliberately", pragma("cache_size") == -16000)
    check("temp tables stay in memory", pragma("temp_store") == 2)
    check("the driver is not opening transactions on its own",
          conn.isolation_level is None)

    check("the schema version is recorded in the file",
          pragma("user_version") == db.SCHEMA_VERSION)
    # Re-opening must not re-run the schema script; that used to happen on every
    # single request and took a write lock each time.
    before = conn.execute(
        "SELECT COUNT(*) AS n FROM sqlite_master WHERE type IN ('table','index','trigger')"
    ).fetchone()["n"]
    conn.close()
    again = db.connect(path)
    check("re-opening leaves the schema alone",
          again.execute(
              "SELECT COUNT(*) AS n FROM sqlite_master WHERE type IN ('table','index','trigger')"
          ).fetchone()["n"] == before)
    check("a second connection still gets its PRAGMAs",
          again.execute("PRAGMA busy_timeout").fetchone()[0] == db.BUSY_TIMEOUT_MS)
    db.close(again)


def test_indexes_are_used(tmpdir):
    """Every lookup the app does at speed must reach an index, not scan.

    The customer/supplier indexes were originally declared COLLATE NOCASE, which
    silently made them unusable for a plain `= ?` — the planner is the only
    honest way to check.
    """
    conn = db.connect(Path(tmpdir) / "indexes.db")

    def plan(sql, params=()):
        rows = conn.execute("EXPLAIN QUERY PLAN " + sql, params).fetchall()
        return " ".join(str(r[3]) for r in rows)

    cases = [
        ("a pass by its number", "SELECT * FROM gate_passes WHERE serial_no = ?", ("FZ-00001",)),
        ("the items of one pass, for printing",
         "SELECT * FROM gate_pass_items WHERE gate_pass_id = ? ORDER BY sl_no", (1,)),
        ("the register filtered by status",
         "SELECT * FROM gate_passes WHERE status = ? ORDER BY id DESC LIMIT 200", ("issued",)),
        ("passes for an invoice number",
         "SELECT * FROM gate_passes WHERE invoice_no = ?", ("FR 262702176",)),
        ("passes for a customer",
         "SELECT * FROM gate_passes WHERE customer_name = ?", ("FANZART LLP",)),
        ("passes on an invoice date",
         "SELECT * FROM gate_passes WHERE invoice_date = ?", ("04-08-2026",)),
        ("a draft's items", "SELECT * FROM draft_items WHERE draft_id = ?", (1,)),
        ("a pass's audit trail", "SELECT * FROM audit_log WHERE gate_pass_id = ?", (1,)),
    ]
    for label, sql, params in cases:
        check(f"{label} uses an index", "USING INDEX" in plan(sql, params)
              or "USING COVERING INDEX" in plan(sql, params))
    db.close(conn)


def test_register_is_capped(tmpdir):
    """The register must not load the whole book to show one screen."""
    conn = db.connect(Path(tmpdir) / "capped.db")
    for i in range(db.REGISTER_PAGE_SIZE + 25):
        db.create_gate_pass(conn, None, "S", "C", f"INV-{i}", "04-08-2026", "",
                             sample_items(), prepared_by="Ravi Kumar")

    page = db.list_gate_passes(conn)
    check("the register returns one capped page", len(page) == db.REGISTER_PAGE_SIZE)
    check("the newest pass is on the first page",
          page[0]["serial_seq"] == db.REGISTER_PAGE_SIZE + 25)
    check("the count reports the true total",
          db.count_gate_passes(conn) == db.REGISTER_PAGE_SIZE + 25)
    check("a filtered count matches its filter",
          db.count_gate_passes(conn, search="INV-7") == len(
              db.list_gate_passes(conn, search="INV-7", limit=None)))
    check("the cap can be lifted for an export",
          len(db.list_gate_passes(conn, limit=None)) == db.REGISTER_PAGE_SIZE + 25)
    db.close(conn)


def test_failed_write_rolls_back(tmpdir):
    """A write that blows up half way must leave nothing behind, and must not
    leave the connection holding the write lock."""
    conn = db.connect(Path(tmpdir) / "rollback.db")
    db.create_gate_pass(conn, None, "S", "C", "KEEP", "04-08-2026", "",
                         sample_items(), prepared_by="Ravi Kumar")
    before = db.count_gate_passes(conn)

    class Boom(Exception):
        pass

    try:
        with db.writing(conn):
            conn.execute("INSERT INTO settings (key, value) VALUES ('doomed', '1')")
            raise Boom()
    except Boom:
        pass

    check("the half-finished write was rolled back",
          conn.execute("SELECT COUNT(*) AS n FROM settings WHERE key = 'doomed'"
                       ).fetchone()["n"] == 0)
    check("the connection is no longer in a transaction", not conn.in_transaction)
    check("the earlier pass is untouched", db.count_gate_passes(conn) == before)

    # And the connection still works afterwards.
    later = db.create_gate_pass(conn, None, "S", "C", "AFTER", "04-08-2026", "",
                                 sample_items(), prepared_by="Ravi Kumar")
    check("writing still works after a rollback", later["serial_seq"] == 2)
    db.close(conn)


def test_users_and_login(tmpdir):
    storage = Path(tmpdir) / "auth_storage"
    flask_app = create_app(db_path=storage / "gate_pass.db", storage_dir=storage)
    flask_app.testing = True
    conn = db.connect(flask_app.config["DB_PATH"])

    check("a fresh database has no users", db.count_users(conn) == 0)
    make_user(conn, "ravi", "Ravi Kumar")
    make_user(conn, "asha", "Asha Nair")
    check("users are created", db.count_users(conn) == 2)

    check("passwords are not stored in the clear",
          "gatepass-test-pw" not in db.get_user_by_username(conn, "ravi")["password_hash"])
    check("the correct password authenticates",
          db.authenticate(conn, "ravi", "gatepass-test-pw") is not None)
    check("a wrong password does not authenticate",
          db.authenticate(conn, "ravi", "wrong") is None)
    check("an unknown user does not authenticate",
          db.authenticate(conn, "nobody", "gatepass-test-pw") is None)
    check("the username is not case sensitive",
          db.authenticate(conn, "RAVI", "gatepass-test-pw") is not None)
    check("a duplicate username is refused",
          raises(ValueError, db.create_user, conn, "ravi", "Someone Else", "another-pw"))
    check("a user with no display name is refused",
          raises(ValueError, db.create_user, conn, "x", "  ", "some-password"))
    check("a short password is refused",
          raises(ValueError, db.create_user, conn, "shorty", "Short Pw", "abc"))

    db.set_user_status(conn, db.get_user_by_username(conn, "asha")["id"], db.DISABLED)
    check("a disabled account cannot sign in",
          db.authenticate(conn, "asha", "gatepass-test-pw") is None)
    db.set_user_status(conn, db.get_user_by_username(conn, "asha")["id"], db.APPROVED)
    conn.close()

    client = flask_app.test_client()
    for path in ("/", "/upload", "/drafts", "/register", "/settings", "/print/1"):
        resp = client.get(path)
        check(f"{path} redirects a signed-out visitor to the login page",
              resp.status_code == 302 and "/login" in resp.headers.get("Location", ""))

    resp = client.post("/login", data={"username": "ravi", "password": "wrong"})
    check("signing in with a wrong password fails", resp.status_code == 401)
    check("the failure does not say whether the username exists",
          b"Wrong username or password" in resp.data)

    resp = client.post("/login", data={"username": "ravi", "password": "gatepass-test-pw"})
    check("signing in with the right password works",
          resp.status_code == 302 and "/upload" in resp.headers.get("Location", ""))
    check("the signed-in name is shown in the app",
          b"Ravi Kumar" in client.get("/upload").data)

    client.post("/logout")
    check("signing out ends the session",
          client.get("/upload").status_code == 302)

    resp = client.get("/register")
    location = unquote(resp.headers.get("Location", ""))
    check("the login page remembers where you were headed",
          "next=/register" in location)

    # ...and only ever bounces back to a path on this site.
    sign_in(client, "ravi")
    client.post("/logout")
    resp = client.post("/login?next=https://example.com/steal",
                        data={"username": "ravi", "password": "gatepass-test-pw"})
    check("an off-site next target is ignored",
          "example.com" not in resp.headers.get("Location", ""))
    resp = client.get("/upload")
    check("signing in still works after a rejected redirect target",
          resp.status_code == 200)


def test_accounts_are_created_by_admins_only(tmpdir):
    """There is no self-signup. Accounts come from an admin, or from the CLI for
    the very first one."""
    storage = Path(tmpdir) / "nosignup"
    flask_app = create_app(db_path=storage / "gate_pass.db", storage_dir=storage)
    flask_app.testing = True
    stranger = flask_app.test_client()

    check("there is no public sign-up page", stranger.get("/signup").status_code == 404)
    check("there is no public sign-up endpoint",
          stranger.post("/signup", data={"username": "x", "display_name": "X",
                                          "password": "gatepass-test-pw",
                                          "confirm_password": "gatepass-test-pw"}
                         ).status_code == 404)
    check("the login page offers no way to request an account",
          b"Request one" not in stranger.get("/login").data)
    # There used to be a check here that the page told a stranger to ask an
    # admin. That line was removed from the sign-in page by hand; what matters
    # is that no self-signup exists, which the four checks above cover.
    check("a stranger cannot reach the People page",
          stranger.get("/people").status_code == 302)
    check("nobody was created", db.count_users(db.connect(flask_app.config["DB_PATH"])) == 0)

    # The first account comes from the command line and is forced to admin.
    conn = db.connect(flask_app.config["DB_PATH"])
    first_id = db.create_user(conn, "ravi", "Ravi Kumar", "gatepass-test-pw")
    first = db.get_user(conn, first_id)
    check("the first account is an approved admin",
          first["is_admin"] == 1 and first["status"] == db.APPROVED)
    conn.close()

    admin = flask_app.test_client()
    sign_in(admin, "ravi")

    # From then on the admin creates people, password and all.
    resp = admin.post("/people/add", data={
        "username": "asha", "display_name": "Asha Nair",
        "password": "asha-initial-pw", "confirm_password": "asha-initial-pw"},
        follow_redirects=True)
    check("an admin can create an account", b"can now sign in" in resp.data)

    conn = db.connect(flask_app.config["DB_PATH"])
    asha = db.get_user_by_username(conn, "asha")
    check("the new account is usable immediately", asha["status"] == db.APPROVED)
    check("the new account is staff by default", asha["is_admin"] == 0)
    check("the password is stored only as a hash",
          "asha-initial-pw" not in asha["password_hash"])
    conn.close()

    staff = flask_app.test_client()
    check("the person can sign in with the password they were given",
          sign_in(staff, "asha", "asha-initial-pw").status_code == 302)

    check("mismatched passwords are refused",
          b"do not match" in admin.post("/people/add", data={
              "username": "bob", "display_name": "Bob", "password": "gatepass-test-pw",
              "confirm_password": "something-else"}, follow_redirects=True).data)
    check("a duplicate username is refused",
          b"already exists" in admin.post("/people/add", data={
              "username": "asha", "display_name": "Impostor",
              "password": "gatepass-test-pw", "confirm_password": "gatepass-test-pw"},
              follow_redirects=True).data)
    check("a short password is refused",
          b"at least 8" in admin.post("/people/add", data={
              "username": "tiny", "display_name": "Tiny", "password": "abc",
              "confirm_password": "abc"}, follow_redirects=True).data)

    check("a non-admin cannot create accounts",
          staff.post("/people/add", data={
              "username": "sneaky", "display_name": "Sneaky",
              "password": "gatepass-test-pw", "confirm_password": "gatepass-test-pw"}
              ).status_code == 302)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("only the intended accounts exist", db.count_users(conn) == 2)
    conn.close()

    # An admin can hand out a new password.
    conn = db.connect(flask_app.config["DB_PATH"])
    asha_id = db.get_user_by_username(conn, "asha")["id"]
    conn.close()
    admin.post(f"/people/{asha_id}/password",
                data={"password": "asha-new-pw", "confirm_password": "asha-new-pw"})
    fresh = flask_app.test_client()
    check("the new password works", sign_in(fresh, "asha", "asha-new-pw").status_code == 302)
    check("the old password no longer works",
          sign_in(flask_app.test_client(), "asha", "asha-initial-pw").status_code == 401)
    check("a non-admin cannot reset anyone's password",
          staff.post(f"/people/{asha_id}/password",
                      data={"password": "hijack-pw", "confirm_password": "hijack-pw"}
                      ).status_code == 302)


def test_permissions_are_stored_per_feature(tmpdir):
    conn = db.connect(Path(tmpdir) / "perms.db")

    admin = db.get_user(conn, db.create_user(conn, "ravi", "Ravi Kumar", "gatepass-test-pw"))
    check("the first account gets every permission",
          all(admin["permissions"].values()))
    check("it holds exactly the known permissions",
          set(admin["permissions"]) == set(db.PERMISSIONS))

    staff = db.get_user(conn, db.create_user(conn, "asha", "Asha Nair", "gatepass-test-pw",
                                              status=db.APPROVED))
    check("a new account starts with no permissions",
          not any(staff["permissions"].values()))
    check("is_admin mirrors can_manage_people", staff["is_admin"] == 0)

    db.set_user_permissions(conn, staff["id"], {"can_search_register": True,
                                                 "can_cancel_passes": True})
    staff = db.get_user(conn, staff["id"])
    check("granting two permissions grants exactly those",
          [k for k, v in staff["permissions"].items() if v]
          == ["can_search_register", "can_cancel_passes"])
    check("user_can answers for a granted permission",
          db.user_can(staff, "can_search_register"))
    check("user_can answers for one not granted",
          not db.user_can(staff, "can_access_settings"))
    check("an unknown permission name is never granted",
          not db.user_can(staff, "can_do_anything"))
    check("a missing user is never granted anything",
          not db.user_can(None, "can_search_register"))

    db.set_user_permissions(conn, staff["id"], {"can_manage_people": True})
    check("granting can_manage_people sets is_admin",
          db.get_user(conn, staff["id"])["is_admin"] == 1)
    check("and drops the ones not listed",
          not db.user_can(db.get_user(conn, staff["id"]), "can_search_register"))

    # Unreadable stored permissions must deny, never grant.
    conn.execute("UPDATE users SET permissions = 'not json' WHERE id = ?", (staff["id"],))
    check("corrupt permissions read as none",
          not any(db.get_user(conn, staff["id"])["permissions"].values()))
    conn.execute('UPDATE users SET permissions = ? WHERE id = ?',
                 ('["can_manage_people"]', staff["id"]))
    check("permissions stored as the wrong shape read as none",
          not any(db.get_user(conn, staff["id"])["permissions"].values()))

    # The last account able to manage people cannot be stripped of it.
    conn.execute("UPDATE users SET permissions = ?, is_admin = 1 WHERE id = ?",
                 (json.dumps(db.ALL_PERMISSIONS), staff["id"]))
    db.set_user_permissions(conn, staff["id"], db.NO_PERMISSIONS)
    check("one of two people-managers can be stripped",
          not db.user_can(db.get_user(conn, staff["id"]), "can_manage_people"))
    check("the last one cannot",
          raises(ValueError, db.set_user_permissions, conn, admin["id"], db.NO_PERMISSIONS))
    db.close(conn)


def test_permissions_gate_each_feature(tmpdir):
    """Each permission opens exactly one door, and no more."""
    flask_app, admin = logged_in_app(tmpdir, "perm_routes",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])
    asha_id = db.get_user_by_username(conn, "asha")["id"]
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "MANGALDEEP",
                              "FR-1", "05-08-2026", "", sample_items(),
                              prepared_by="Ravi Kumar")
    conn.close()

    staff = flask_app.test_client()
    sign_in(staff, "asha")

    def grant(**permissions):
        conn = db.connect(flask_app.config["DB_PATH"])
        db.set_user_permissions(conn, asha_id, permissions)
        conn.close()

    GATED = [("can_export_reports", "/reports"),
              ("can_access_settings", "/settings"),
              ("can_manage_people", "/people")]

    for permission, path in GATED:
        grant()
        resp = staff.get(path, follow_redirects=True)
        check(f"{path} is closed without {permission}", resp.request.path == "/register")
        check(f"{path} explains why", b"403 Unauthorized" in resp.data)

        grant(**{permission: True})
        check(f"{permission} opens {path}", staff.get(path).status_code == 200)

        # ...and opens nothing else.
        for other_permission, other_path in GATED:
            if other_path == path:
                continue
            check(f"{permission} does not open {other_path}",
                  staff.get(other_path, follow_redirects=True).request.path == "/register")

    # Searching and filtering the register are ONE permission: nobody wanted to
    # search a register they could not narrow, or narrow one they could not
    # search, so the two tick boxes became one.
    grant()
    page = staff.get("/register").data
    check("without it there is no search box", b'name="q"' not in page)
    check("and no filter button", b'id="filter-toggle"' not in page)
    check("and a search term in the URL is ignored",
          staff.get("/register?q=NOTHING").data.count(b'serial-cell reg-serial') == 1)
    check("and so is a status filter",
          staff.get("/register?status=cancelled").data.count(b'serial-cell reg-serial') == 1)

    grant(can_search_register=True)
    page = staff.get("/register").data
    check("the one permission shows the search box", b'name="q"' in page)
    check("and the filter button with it", b'id="filter-toggle"' in page)
    check("searching works",
          staff.get("/register?q=MANGALDEEP").data.count(b'serial-cell reg-serial') == 1)
    check("and filtering works",
          staff.get("/register?status=cancelled").data.count(b'serial-cell reg-serial') == 0)
    check("can_filter_register no longer exists",
          "can_filter_register" not in db.PERMISSIONS)

    # Cancelling is its own permission.
    grant()
    check("cancel is hidden without the permission",
          b">Cancel<" not in staff.get("/register").data)
    check("and refused if posted directly",
          staff.post(f"/gate-passes/{gp['id']}/cancel",
                      data={"reason": "no"}).status_code == 302)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("the pass is untouched",
          db.get_gate_pass(conn, gp["id"])["status"] == "issued")
    conn.close()

    grant(can_cancel_passes=True)
    check("cancel appears with the permission", b">Cancel<" in staff.get("/register").data)
    staff.post(f"/gate-passes/{gp['id']}/cancel", data={"reason": "damaged"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("and now works", db.get_gate_pass(conn, gp["id"])["status"] == "cancelled")
    conn.close()

    # Everything else stays open to everyone.
    grant()
    for path in ("/upload", "/register"):
        check(f"{path} needs no permission", staff.get(path).status_code == 200)
    # ...except Drafts, which is the switch between "check it first" and
    # "issue it straight away". Without it there is no draft list to show.
    check("/drafts is behind can_review_drafts",
          staff.get("/drafts").status_code == 302)


def test_managing_permissions_from_the_people_page(tmpdir):
    flask_app, admin = logged_in_app(tmpdir, "perm_ui",
                                      users=(("ravi", "Ravi Kumar"),))
    page = admin.get("/people").data
    check("the create form offers the permission checkboxes",
          all(f'name="{key}"'.encode() in page for key in db.PERMISSIONS))
    # Creating an account is a dialog opened from the Accounts heading, not a
    # panel wedged under the table.
    check("there is an Add User button", b'id="add-user-open"' in page)
    check("and the form lives in a dialog", b'id="add-dialog"' in page)
    check("each tick box says what it does",
          all(hint.encode() in page for hint in
              (db.PERMISSION_HINTS["can_review_drafts"][:30],
               db.PERMISSION_HINTS["can_batch_print"][:30])))
    check("the old admin tick is gone", b'name="is_admin"' not in page)

    admin.post("/people/add", data={
        "username": "asha", "display_name": "Asha Nair",
        "password": "asha-initial-pw", "confirm_password": "asha-initial-pw",
        "can_search_register": "1", "can_cancel_passes": "1"})

    conn = db.connect(flask_app.config["DB_PATH"])
    asha = db.get_user_by_username(conn, "asha")
    check("an account is created with the ticked permissions",
          [k for k, v in asha["permissions"].items() if v]
          == ["can_search_register", "can_cancel_passes"])
    check("and nothing else", asha["is_admin"] == 0)
    conn.close()

    page = admin.get("/people").data
    check("the table shows how many permissions are held",
          f"2 of {len(db.PERMISSIONS)}".encode() in page)
    check("a full-access account is labelled", b"Full access" in page)
    check("there is an Edit permissions button", b"edit-perms" in page)
    check("the button carries the current permissions", b"data-perms" in page)
    check("there is a permissions dialog", b'id="perm-dialog"' in page)

    admin.post(f"/people/{asha['id']}/permissions",
                data={"can_export_reports": "1", "can_access_settings": "1"})
    conn = db.connect(flask_app.config["DB_PATH"])
    updated = db.get_user(conn, asha["id"])
    check("the dialog replaces the whole set rather than adding to it",
          [k for k, v in updated["permissions"].items() if v]
          == ["can_export_reports", "can_access_settings"])
    conn.close()

    staff = flask_app.test_client()
    sign_in(staff, "asha", "asha-initial-pw")
    check("the new permissions apply immediately",
          staff.get("/reports").status_code == 200)
    check("a permission not granted is still closed",
          staff.get("/people", follow_redirects=True).request.path == "/register")

    check("a non-manager cannot edit anyone's permissions",
          staff.post(f"/people/{asha['id']}/permissions",
                      data={"can_manage_people": "1"}).status_code == 302)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("and nothing changed",
          not db.user_can(db.get_user(conn, asha["id"]), "can_manage_people"))
    ravi_id = db.get_user_by_username(conn, "ravi")["id"]
    conn.close()

    resp = admin.post(f"/people/{ravi_id}/permissions", data={}, follow_redirects=True)
    check("an admin cannot remove their own people-management permission",
          b"cannot remove your own" in resp.data)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("so they keep it",
          db.user_can(db.get_user(conn, ravi_id), "can_manage_people"))
    conn.close()


def test_permissions_backfill_on_upgrade(tmpdir):
    """An account from before permissions existed keeps what it could already do."""
    path = Path(tmpdir) / "old_users.db"
    old = sqlite3.connect(path)
    old.executescript("""
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            display_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'approved',
            is_admin INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            decided_at TEXT, decided_by TEXT);
        INSERT INTO users (username, display_name, password_hash, status, is_admin)
        VALUES ('boss', 'The Boss', 'x', 'approved', 1),
               ('hand', 'A Hand',   'x', 'approved', 0);
        PRAGMA user_version = 5;
    """)
    old.commit()
    old.close()

    conn = db.connect(path)
    boss = db.get_user_by_username(conn, "boss")
    hand = db.get_user_by_username(conn, "hand")
    check("a former admin keeps every permission", all(boss["permissions"].values()))
    check("former staff start with none", not any(hand["permissions"].values()))
    check("both accounts survive the upgrade", db.count_users(conn) == 2)
    db.close(conn)


def test_batch_printing(tmpdir):
    """Several passes printed as one document, straight to the print dialog."""
    flask_app, admin = logged_in_app(tmpdir, "batchprint",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])
    made = [db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                                 f"FR-{i}", "05-08-2026", "", sample_items(),
                                 prepared_by="Ravi Kumar") for i in range(4)]
    # One long pass, so a batch has to carry a multi-page member too.
    long_pass = db.create_gate_pass(
        conn, None, "Golden Touch Exports", "MANGALDEEP", "FR-LONG", "05-08-2026", "",
        [{"item_name": f"Item {i}", "quantity": "1"} for i in range(32)],
        prepared_by="Ravi Kumar")
    cancelled = db.create_gate_pass(conn, None, "S", "C", "FR-X", "05-08-2026", "",
                                     sample_items(), prepared_by="Ravi Kumar")
    db.cancel_gate_pass(conn, cancelled["id"], "test")
    asha_id = db.get_user_by_username(conn, "asha")["id"]
    conn.close()

    ids = ",".join(str(gp["id"]) for gp in made[:3])
    page = admin.get(f"/print/batch?ids={ids}").data.decode()

    check("all three passes are in one document",
          all(gp["serial_no"] in page for gp in made[:3]))
    check("one sheet per pass", page.count('class="sheet') == 3)
    check("each sheet still carries two copies to cut apart",
          page.count('<div class="pass">') == 6)
    check("the print dialog opens by itself", "window.print()" in page)
    check("it waits for load so the logo is there",
          'addEventListener("load"' in page)
    check("the page breaks between sheets", "page-break-after" in
          (ROOT / "static" / "css" / "print.css").read_text())
    check("nothing is offered as a download",
          "Content-Disposition" not in str(admin.get(f"/print/batch?ids={ids}").headers))
    check("it is an HTML page, not a file",
          admin.get(f"/print/batch?ids={ids}").headers["Content-Type"].startswith("text/html"))

    # A multi-page pass contributes all of its pages.
    both = f"{made[0]['id']},{long_pass['id']}"
    page = admin.get(f"/print/batch?ids={both}").data.decode()
    check("a 32-item pass brings both its pages into the batch",
          page.count('class="sheet') == 3)
    check("and its page indicators come with it",
          "Page 1 of 2" in page and "Page 2 of 2" in page)
    check("while the short pass shows none",
          page.count("Page 1 of 2") == 2)

    # Input handling.
    check("duplicate ids are printed once",
          admin.get(f"/print/batch?ids={made[0]['id']},{made[0]['id']}"
                     ).data.decode().count('class="sheet') == 1)
    check("rubbish ids are ignored",
          admin.get(f"/print/batch?ids=abc,{made[0]['id']},,-1"
                     ).data.decode().count('class="sheet') == 1)
    check("no ids at all is refused",
          admin.get("/print/batch?ids=", follow_redirects=True).request.path == "/register")
    check("ids that match nothing are refused",
          admin.get("/print/batch?ids=999999", follow_redirects=True
                     ).request.path == "/register")
    resp = admin.get(f"/print/batch?ids={made[0]['id']},999999", follow_redirects=True)
    check("a missing pass is left out rather than failing the batch",
          b"left out" in resp.data)
    too_many = ",".join(str(i) for i in range(1, db.MAX_BATCH_PRINT + 2))
    resp = admin.get(f"/print/batch?ids={too_many}", follow_redirects=True)
    check("an oversized batch is refused rather than hanging the browser",
          resp.request.path == "/register" and b"most that can be printed" in resp.data)

    # The permission gates it, on the page and on the route.
    def grant(**permissions):
        conn = db.connect(flask_app.config["DB_PATH"])
        db.set_user_permissions(conn, asha_id, permissions)
        conn.close()

    staff = flask_app.test_client()
    sign_in(staff, "asha")

    grant()
    page = staff.get("/register").data
    check("without the permission there are no tick boxes",
          b'class="pass-tick"' not in page)
    check("and no select-all", b'id="select-all"' not in page)
    check("and no action bar", b'id="bulk-bar"' not in page)
    check("and the route itself is closed",
          staff.get(f"/print/batch?ids={ids}", follow_redirects=True).request.path
          == "/register")

    grant(can_batch_print=True)
    page = staff.get("/register").data
    check("the permission brings the tick boxes", b'class="pass-tick"' in page)
    check("and the select-all box", b'id="select-all"' in page)
    check("and the action bar", b'id="bulk-bar"' in page)
    check("and the print button", b"Print Selected Passes" in page)
    check("the route opens too",
          staff.get(f"/print/batch?ids={ids}").status_code == 200)
    check("a cancelled pass cannot be selected",
          page.count(b'class="pass-tick"') == 5)
    check("batch print does not imply any other permission",
          staff.get("/reports", follow_redirects=True).request.path == "/register")


def test_role_based_access(tmpdir):
    """What each role can reach, enforced on the server rather than by hiding
    links — a hidden nav item is a courtesy, not a control."""
    flask_app, admin = logged_in_app(tmpdir, "rbac",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    staff = flask_app.test_client()
    sign_in(staff, "asha")

    ADMIN_ONLY = ["/reports", "/reports/export", "/settings", "/people", "/drafts"]
    EVERYONE = ["/upload", "/register"]

    for path in ADMIN_ONLY:
        check(f"an admin can open {path}", admin.get(path).status_code == 200)
        resp = staff.get(path, follow_redirects=True)
        check(f"a regular user is redirected away from {path}",
              resp.request.path == "/register")
        check(f"and is told why for {path}", b"403 Unauthorized" in resp.data)

    for path in EVERYONE:
        check(f"a regular user can open {path}", staff.get(path).status_code == 200)
        check(f"an admin can open {path}", admin.get(path).status_code == 200)

    # Navigation hides what the person cannot use.
    staff_nav = staff.get("/register").data
    admin_nav = admin.get("/register").data
    for link in (b'href="/reports"', b'href="/settings"', b'href="/people"'):
        check(f"{link.decode()} is hidden from a regular user", link not in staff_nav)
        check(f"{link.decode()} is shown to an admin", link in admin_nav)
    check("a regular user still sees the everyday links",
          b'href="/upload"' in staff_nav and b'href="/register"' in staff_nav)
    # No draft list without the permission — their uploads go straight to a
    # printed pass, so there is nothing for that page to hold.
    check("Drafts is hidden from a user who does not review",
          b'href="/drafts"' not in staff_nav)
    check("Drafts is shown to an admin", b'href="/drafts"' in admin_nav)

    # The register's search and filter are admin-only, and enforced server side.
    check("a regular user gets no search box", b'name="q"' not in staff_nav)
    check("a regular user gets no filter button", b'id="filter-toggle"' not in staff_nav)
    check("an admin gets the search box", b'name="q"' in admin_nav)
    check("an admin gets the filter button", b'id="filter-toggle"' in admin_nav)

    conn = db.connect(flask_app.config["DB_PATH"])
    for i in range(3):
        db.create_gate_pass(conn, None, "Golden Touch Exports",
                             "MANGALDEEP" if i == 0 else "FANZART LLP",
                             f"INV-{i}", "05-08-2026", "", sample_items(),
                             prepared_by="Ravi Kumar")
    conn.close()

    filtered = staff.get("/register?q=MANGALDEEP&status=cancelled&from=2020-01-01").data
    check("a regular user's query string is ignored, not obeyed",
          filtered.count(b'serial-cell reg-serial') == 3)
    check("an admin's search still works",
          admin.get("/register?q=MANGALDEEP").data.count(b'serial-cell reg-serial') == 1)

    # Losing admin takes effect on the next click, not at the next sign-in.
    conn = db.connect(flask_app.config["DB_PATH"])
    ravi_id = db.get_user_by_username(conn, "ravi")["id"]
    asha_id = db.get_user_by_username(conn, "asha")["id"]
    db.set_user_admin(conn, asha_id, True)
    conn.close()
    check("a promoted user reaches the admin pages at once",
          staff.get("/reports").status_code == 200)
    conn = db.connect(flask_app.config["DB_PATH"])
    db.set_user_admin(conn, asha_id, False)
    conn.close()
    check("a demoted user loses them at once",
          staff.get("/reports", follow_redirects=True).request.path == "/register")


def test_admin_powers_and_guardrails(tmpdir):
    flask_app, admin = logged_in_app(tmpdir, "admin_storage",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])
    ravi_id = db.get_user_by_username(conn, "ravi")["id"]
    asha_id = db.get_user_by_username(conn, "asha")["id"]
    conn.close()

    staff = flask_app.test_client()
    sign_in(staff, "asha")
    # A signed-in non-admin is sent back to the register with a message rather
    # than shown a bare 403 — they followed a link that is not theirs.
    resp = staff.get("/people", follow_redirects=True)
    check("a non-admin cannot open the People page", resp.request.path == "/register")
    check("they are told why", b"403 Unauthorized" in resp.data)
    check("a non-admin cannot approve anyone",
          staff.post(f"/people/{asha_id}/status",
                      data={"status": "approved"}).status_code == 302)
    check("a non-admin cannot make themselves an admin",
          staff.post(f"/people/{asha_id}/role", data={"is_admin": "1"}).status_code == 302)
    check("a signed-out visitor is sent to login, not shown a 403",
          flask_app.test_client().get("/people").status_code == 302)
    check("the People link is hidden from non-admins",
          b"/people" not in staff.get("/upload").data)

    # An admin can promote someone else — the point of the feature.
    admin.post(f"/people/{asha_id}/role", data={"is_admin": "1"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("an admin can make another admin",
          db.get_user_by_username(conn, "asha")["is_admin"] == 1)
    check("there are now two admins", db.count_admins(conn) == 2)
    conn.close()
    check("the new admin can open the People page",
          staff.get("/people").status_code == 200)

    # Guardrails against locking the office out.
    check("an admin cannot remove their own admin rights",
          b"cannot remove your own" in
          admin.post(f"/people/{ravi_id}/role",
                      data={"is_admin": "0"}, follow_redirects=True).data)
    check("an admin cannot switch off their own account",
          b"cannot switch off your own" in
          admin.post(f"/people/{ravi_id}/status",
                      data={"status": "disabled"}, follow_redirects=True).data)

    # Demote the second admin, then check the last one cannot be removed.
    admin.post(f"/people/{asha_id}/role", data={"is_admin": "0"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("an admin can demote another admin",
          db.get_user_by_username(conn, "asha")["is_admin"] == 0)
    check("the last admin cannot be demoted at the database level",
          raises(ValueError, db.set_user_admin, conn, ravi_id, False))
    check("the last admin cannot be switched off at the database level",
          raises(ValueError, db.set_user_status, conn, ravi_id, db.DISABLED))
    check("an unapproved account cannot be made an admin",
          raises(ValueError, db.set_user_admin, conn,
                 make_user(conn, "new", "New Person", status=db.PENDING), True))
    conn.close()

    # Switching someone off takes effect on their very next click.
    admin.post(f"/people/{asha_id}/status", data={"status": "disabled"})
    check("a disabled person is thrown out mid-session",
          staff.get("/upload").status_code == 302)


def test_remarks_are_typed_and_printed(tmpdir):
    """Remarks are entered on the review screen and printed on the pass.

    Unlike cartons (rule 5) a remark CAN be known at the keyboard — it is a note
    about the consignment, not a count of how it was packed — so it is captured
    with the pass. The printed box still prints empty when nobody typed
    anything, so it can be written on by hand at the gate.
    """
    flask_app, client = logged_in_app(tmpdir, "remarks")
    conn = db.connect(flask_app.config["DB_PATH"])
    draft_id = db.create_draft(conn, supplier_name="Golden Touch Exports",
                                customer_name="FANZART LLP", invoice_no="FR 1",
                                invoice_date="01-01-2026", items=sample_items())
    conn.close()

    page = client.get(f"/review/{draft_id}").get_data(as_text=True)
    check("the review screen offers a Remarks field", 'name="remarks"' in page)
    check("and marks it optional", "Remarks" in page and "(optional)" in page)

    note = "Fragile — 2 boxes short-shipped, balance to follow"
    resp = client.post(f"/review/{draft_id}", headers={"Origin": "http://localhost"},
                       data=MultiDict([
                           ("supplier_name", "Golden Touch Exports"),
                           ("customer_name", "FANZART LLP"), ("invoice_no", "FR 1"),
                           ("invoice_date", "01-01-2026"), ("vehicle_no", ""),
                           ("remarks", note),
                           ("item_name", "MICRON MODREN OAK"), ("quantity", "1"),
                           ("cartons", ""), ("action", "issue"),
                       ]))
    gate_pass_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    conn = db.connect(flask_app.config["DB_PATH"])
    check("the remark is stored on the pass",
          db.get_gate_pass(conn, gate_pass_id)["remarks"] == note)
    conn.close()

    printed = " ".join(re.sub(r"<[^>]+>", " ",
                              client.get(f"/print/{gate_pass_id}").get_data(as_text=True)).split())
    check("and printed in the Remarks box", f"Remarks : {note}" in printed)

    # Nothing typed: the box still prints, empty, to be written in.
    conn = db.connect(flask_app.config["DB_PATH"])
    blank = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                                 sample_items(), prepared_by="Ravi Kumar")
    conn.close()
    printed = " ".join(re.sub(r"<[^>]+>", " ",
                              client.get(f"/print/{blank['id']}").get_data(as_text=True)).split())
    check("an empty remark still prints the box",
          "Remarks :" in printed and note not in printed)

    # Never parsed from an invoice, so always the operator's to type — the same
    # reason vehicle_no is always editable.
    draft = {"supplier_name": "S", "customer_name": "C", "invoice_no": "I",
             "invoice_date": "D", "vehicle_no": "", "items": [], "parse_notes": ""}
    fields, _items = appmod._fill_blanks_only(draft, MultiDict([("remarks", "typed")]))
    check("remarks is not something the parser can lock", "remarks" not in fields)


def test_item_table_keyboard_navigation(tmpdir):
    """Enter goes across the row; the arrows go down the column.

    Enter used to do the same as Down. That is the wrong direction for reading
    a line off an invoice — item, then quantity, then cartons — so it now moves
    along the row and falls into the next row at the end of one. The arrows
    keep the vertical movement, for filling one column down a whole delivery.

    Asserted against the template source here; test_ui_flows.py drives it in a
    real browser, which is what actually proves the focus moves.
    """
    flask_app, client = logged_in_app(tmpdir, "keys")
    conn = db.connect(flask_app.config["DB_PATH"])
    draft_id = db.create_draft(conn, supplier_name="S", customer_name="C",
                                invoice_no="I", invoice_date="D", items=sample_items())
    conn.close()
    page = client.get(f"/review/{draft_id}").get_data(as_text=True)

    check("the items table listens for keys", 'tbody.addEventListener("keydown"' in page)
    check("the arrows are handled", '"ArrowDown"' in page and '"ArrowUp"' in page)
    check("Enter is handled", '"Enter"' in page)
    check("Enter walks the inputs in document order, so it crosses rows",
          'tbody.querySelectorAll("input")' in page)
    check("the arrows still move by row, keeping the column",
          "rowIndex + vertical" in page)

    # Enter inside a form submits it by default. On a review screen that issues
    # a gate pass on a stray keystroke, so it has to be stopped every time —
    # including in the very last cell, where there is nowhere to move to.
    check("Enter can never submit the form",
          "e.preventDefault();" in page
          and page.index("e.preventDefault();") < page.index("let next = null;"))


def test_login_cannot_be_brute_forced(tmpdir):
    """A password cannot be guessed indefinitely.

    Only matters once the site is on the public internet, which it now is: a
    Funnel hostname is published in Certificate Transparency logs the moment
    its certificate is issued, so scanners find it within hours.
    """
    storage = Path(tmpdir) / "bruteforce"
    flask_app = create_app(db_path=storage / "gate_pass.db", storage_dir=storage)
    conn = db.connect(flask_app.config["DB_PATH"])
    make_user(conn, "victim", "Victim", "the-real-password-99")
    client = flask_app.test_client()

    def guess(password, ip="203.0.113.9", username="victim"):
        return client.post("/login", data={"username": username, "password": password},
                           headers={"X-Forwarded-For": ip})

    codes = [guess("wrong-%d" % i).status_code for i in range(db.LOGIN_MAX_PER_USERNAME)]
    check("guessing is allowed up to the limit", all(c == 401 for c in codes))

    blocked = guess("wrong-again")
    check("then it is refused with 429", blocked.status_code == 429)
    check("and says how long to wait", b"Try again in about" in blocked.data)

    # The point of checking the lockout BEFORE the password: a locked-out
    # attacker must not be able to tell a right guess from a wrong one, or the
    # lockout becomes an oracle that confirms the password for later.
    right = guess("the-real-password-99")
    check("the CORRECT password is refused too while locked out",
          right.status_code == 429)
    check("and does not sign them in", db.get_user_by_username(conn, "victim") is not None
          and b"Sign in" in right.data)

    # A different account is unaffected — one person under attack must not lock
    # the rest of the office out.
    make_user(conn, "colleague", "Colleague", "another-password-77")
    other = guess("nope", username="colleague")
    check("a different username is not caught by it", other.status_code == 401)

    # Getting it right clears the slate, so three typos this morning plus five
    # this afternoon is not a lockout.
    fresh = flask_app.test_client()
    for i in range(3):
        fresh.post("/login", data={"username": "colleague", "password": "no-%d" % i},
                   headers={"X-Forwarded-For": "203.0.113.10"})
    ok = fresh.post("/login", data={"username": "colleague", "password": "another-password-77"},
                    headers={"X-Forwarded-For": "203.0.113.10"})
    check("a correct password still works after a few typos", ok.status_code in (302, 200))
    check("and the failures are forgotten",
          db.login_lockout(conn, "colleague", "203.0.113.10") == 0)

    check("attempts are recorded against the forwarded address, not the proxy",
          conn.execute("SELECT COUNT(*) FROM login_attempts WHERE ip = '127.0.0.1'"
                       ).fetchone()[0] == 0)
    conn.close()


def test_back_button_after_issuing(tmpdir):
    """Going Back to a draft that has been issued finds the pass, not a 404.

    Issuing DELETES the draft, so /review/<id> dies the moment the pass exists
    — and that is exactly where the browser's Back button lands afterwards. A
    bare 404 tells the operator nothing and appears to have lost the pass they
    just made. The audit row written at issue ends with "from draft 51", so the
    link survives even though the row does not.
    """
    flask_app, client = logged_in_app(tmpdir, "backbutton")
    conn = db.connect(flask_app.config["DB_PATH"])
    H = {"Origin": "http://localhost"}

    def a_draft(invoice_no):
        return db.create_draft(conn, supplier_name="Golden Touch Exports",
                                customer_name="FANZART LLP", invoice_no=invoice_no,
                                invoice_date="01-01-2026", items=sample_items())

    # Issued one at a time.
    draft_id = a_draft("FR 1")
    resp = client.post(f"/review/{draft_id}", headers=H, data=MultiDict([
        ("supplier_name", "Golden Touch Exports"), ("customer_name", "FANZART LLP"),
        ("invoice_no", "FR 1"), ("invoice_date", "01-01-2026"), ("vehicle_no", ""),
        ("remarks", ""), ("item_name", "MICRON MODREN OAK"), ("quantity", "1"),
        ("cartons", ""), ("action", "issue")]))
    gate_pass_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    back = client.get(f"/review/{draft_id}")
    check("Back after issuing redirects rather than 404ing", back.status_code == 302)
    check("and lands on the pass that draft became",
          back.headers.get("Location", "").endswith(f"/print/{gate_pass_id}"))

    # Issued as part of a batch — a different audit wording, same link.
    batch_id = a_draft("FR 2")
    client.post("/drafts/issue", data={"draft_id": str(batch_id)}, headers=H)
    issued = conn.execute("SELECT id FROM gate_passes ORDER BY id DESC LIMIT 1").fetchone()["id"]
    back = client.get(f"/review/{batch_id}")
    check("a batch-issued draft is found too",
          back.headers.get("Location", "").endswith(f"/print/{issued}"))

    # Discarded, not issued: it must NOT claim a pass exists.
    gone = a_draft("FR 3")
    client.post(f"/drafts/{gone}/delete", headers=H)
    back = client.get(f"/review/{gone}")
    check("a discarded draft does not point at somebody else's pass",
          "/print/" not in back.headers.get("Location", ""))
    check("and says what happened to it",
          b"issued or discarded" in client.get(f"/review/{gone}",
                                                follow_redirects=True).data)

    # The lookup matches on the id at the END of the audit line, so a draft
    # numbered 5 must never be confused with one numbered 51.
    check("a draft that never existed finds nothing",
          db.gate_pass_from_draft(conn, 99999) is None)

    # print.html renders no flash messages, so a message set on that redirect
    # would sit in the session and surface later on an unrelated screen.
    printed = (ROOT / "templates" / "print.html").read_text()
    check("the print page renders no flashes, so none is set for it",
          "get_flashed_messages" not in printed)
    later = client.get("/register", follow_redirects=True).data
    check("nothing is left hanging for the next page",
          b"Already issued" not in later)
    conn.close()


def test_stylesheets_are_cache_busted(tmpdir):
    """Every stylesheet link carries the file's modification time.

    Without it the browser keeps serving the CSS it already has, and that
    produces the most confusing possible failure: template changes appear
    immediately — they are rendered fresh every request — while CSS changes do
    not. A layout looks half-applied and the code looks wrong when it is
    correct. It cost real time more than once before this existed.
    """
    flask_app, client = logged_in_app(tmpdir, "cachebust")
    for path in ("/login", "/register", "/upload"):
        page = client.get(path, follow_redirects=True).get_data(as_text=True)
        for href in re.findall(r'<link[^>]+href="([^"]*\.css[^"]*)"', page):
            check(f"{path} asks for {href.split('?')[0]} with a version",
                  "?v=" in href)

    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "S", "C", "I", "01-01-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")
    conn.close()
    printed = client.get(f"/print/{gp['id']}").get_data(as_text=True)
    hrefs = re.findall(r'<link[^>]+href="([^"]*\.css[^"]*)"', printed)
    check("the print stylesheet is versioned too",
          hrefs and all("?v=" in h for h in hrefs))

    # The stamp must follow the file, or it is just a constant that never busts
    # anything. Touch the file and the link has to change.
    with flask_app.test_request_context():
        static_url = flask_app.jinja_env.globals["static_url"]
        before = static_url("css/print.css")
        css = Path(flask_app.static_folder) / "css" / "print.css"
        stamp = css.stat().st_mtime
        try:
            os.utime(css, (stamp + 60, stamp + 60))
            check("touching the file changes the link",
                  static_url("css/print.css") != before)
        finally:
            os.utime(css, (stamp, stamp))
        check("and restoring it restores the link",
              static_url("css/print.css") == before)
        check("a missing file does not raise", "?v=" in static_url("css/nope.css"))


def test_printed_issue_date_is_the_pass_date(tmpdir):
    """The printed Issue Date comes off the record, not off the clock.

    A gate pass is a document. Reprinting FZ-00027 a year later must produce
    the same sheet it produced the day it was raised — `datetime.now()` in the
    template would quietly restamp it with the day it was reprinted, which is a
    different claim about when the goods left.
    """
    flask_app, client = logged_in_app(tmpdir, "issuedate")
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "Golden Touch Exports", "FANZART LLP",
                              "FR 1", "01-01-2026", "", sample_items(),
                              prepared_by="Ravi Kumar")
    # Backdate it the way an old pass in the register would be.
    with db.writing(conn):
        conn.execute("UPDATE gate_passes SET issued_at = ? WHERE id = ?",
                     ("2026-02-14 09:30:00", gp["id"]))
    conn.close()

    printed = client.get(f"/print/{gp['id']}").get_data(as_text=True)
    # Tags stripped: the label and its colon are separate elements so the
    # colons can be aligned, so the pair only reads as one string once rendered.
    text = " ".join(re.sub(r"<[^>]+>", " ", printed).split())
    check("the printed pass shows the date it was issued",
          "Issue Date : 14-02-2026" in text)
    check("and not today's date",
          datetime.now().strftime("%d-%m-%Y") not in printed
          or datetime.now().strftime("%d-%m-%Y") == "14-02-2026")
    check("the serial is still there", gp["serial_no"] in printed)

    # The filter turns the stored 'YYYY-MM-DD HH:MM:SS' into the way the office
    # writes a date, and leaves anything it cannot read alone.
    with flask_app.test_request_context():
        fmt = flask_app.jinja_env.filters["as_day_month_year"]
        check("timestamps become DD-MM-YYYY", fmt("2026-02-14 09:30:00") == "14-02-2026")
        check("a bare date works too", fmt("2026-12-01") == "01-12-2026")
        check("something unreadable is passed through", fmt("") == "")
        check("and None does not raise", fmt(None) == "")


@needs_fixtures
def test_the_parse_is_the_record(tmpdir):
    """Without `can_edit_parsed_details`, a value the invoice yielded is fixed.

    The rule: **the parser's output is the record.** A person fills in what the
    document did not yield; they never overwrite what it did. That is what
    makes a draft worth landing on for someone who may not edit — the blanks
    are theirs to fill, the rest is the invoice speaking for itself.
    """
    flask_app, admin = logged_in_app(tmpdir, "parsedrec",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])
    asha = db.get_user_by_username(conn, "asha")
    # She reviews, but may not restate what the document said.
    db.set_user_permissions(conn, asha["id"], {"can_review_drafts": True})
    conn.close()
    staff = flask_app.test_client()
    sign_in(staff, "asha")

    with open(SAMPLE_INVOICE, "rb") as f:
        resp = staff.post("/upload", data={"invoice": (f, "inv.pdf")},
                           content_type="multipart/form-data")
    draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    conn = db.connect(flask_app.config["DB_PATH"])
    parsed = db.get_draft(conn, draft_id)
    conn.close()

    page = staff.get(f"/review/{draft_id}").get_data(as_text=True)
    check("a parsed value is shown as text, not an input",
          'class="read-field"' in page
          and 'id="customer_name"' not in page)
    check("and the item list cannot be added to", 'id="add-row"' not in page)

    # A readonly attribute is a courtesy to the person typing, not a control:
    # anyone can post whatever they like to this route.
    staff.post(f"/review/{draft_id}", headers={"Origin": "http://localhost"},
               data=MultiDict([
                   ("supplier_name", "SOMEONE ELSE"), ("customer_name", "NOT THEM"),
                   ("invoice_no", "FAKE-999"), ("invoice_date", "01-01-2000"),
                   ("vehicle_no", "KA 01 AB 1234"),
                   ("item_name", "GOLD BARS"), ("quantity", "999"), ("cartons", "5"),
                   ("action", "save"),
               ]))
    conn = db.connect(flask_app.config["DB_PATH"])
    after = db.get_draft(conn, draft_id)
    conn.close()
    for field in ("supplier_name", "customer_name", "invoice_no", "invoice_date"):
        check(f"a forged post cannot change {field}", after[field] == parsed[field])
    check("nor the items the invoice listed",
          [i["item_name"] for i in after["items"]] == [i["item_name"] for i in parsed["items"]])
    check("nor how many there were", len(after["items"]) == len(parsed["items"]))
    # Neither of these comes from the parser, so both are theirs to enter.
    check("the vehicle number is still theirs to enter",
          after["vehicle_no"] == "KA 01 AB 1234")
    check("and so are cartons, which the parser never returns",
          after["items"][0]["cartons"] == "5")

    # A scan yields nothing, so there is nothing to protect and everything to
    # type. Locking the whole screen would strand every scanned invoice.
    with open(SCANNED_INVOICE, "rb") as f:
        resp = staff.post("/upload", data={"invoice": (f, "scan.pdf")},
                           content_type="multipart/form-data")
    scan_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    page = staff.get(f"/review/{scan_id}").get_data(as_text=True)
    check("a scan offers every field as a box", 'class="read-field"' not in page)
    check("and lets rows be added", 'id="add-row"' in page)

    resp = staff.post(f"/review/{scan_id}", headers={"Origin": "http://localhost"},
                      data=MultiDict([
                          ("supplier_name", "MANGALDEEP"), ("customer_name", "FANZART LLP"),
                          ("invoice_no", "HAND-001"), ("invoice_date", "07-08-2026"),
                          ("vehicle_no", ""),
                          ("item_name", "VIENNA 52 WALNUT"), ("quantity", "2"),
                          ("cartons", ""), ("action", "issue"),
                      ]))
    check("a typed-in scan issues normally", "/print/" in resp.headers.get("Location", ""))
    conn = db.connect(flask_app.config["DB_PATH"])
    issued = conn.execute("SELECT * FROM gate_passes ORDER BY id DESC LIMIT 1").fetchone()
    check("with the details the operator entered", issued["supplier_name"] == "MANGALDEEP")
    conn.close()

    # An admin correcting a genuine mis-parse is the escape hatch.
    admin.post(f"/review/{draft_id}", headers={"Origin": "http://localhost"},
               data=MultiDict([
                   ("supplier_name", "Corrected Supplier"), ("customer_name", "FANZART LLP"),
                   ("invoice_no", "FR 262702171"), ("invoice_date", "03-08-2026"),
                   ("vehicle_no", ""), ("item_name", "MICRON MODREN OAK"),
                   ("quantity", "1"), ("cartons", ""), ("action", "save"),
               ]))
    conn = db.connect(flask_app.config["DB_PATH"])
    check("an admin can still correct a mis-parse",
          db.get_draft(conn, draft_id)["supplier_name"] == "Corrected Supplier")
    conn.close()


@needs_fixtures
def test_instant_issue_without_the_draft_step(tmpdir):
    """Without `can_review_drafts`, an upload becomes a printed pass in one go.

    Parse, number, insert and issue happen in a single transaction, and the
    operator lands on the print preview. With the permission, the same upload
    stops at a draft first so the parse can be corrected before a number is
    spent on it.
    """
    flask_app, admin = logged_in_app(tmpdir, "instant",
                                      users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    staff = flask_app.test_client()
    sign_in(staff, "asha")

    def upload(client, path=SAMPLE_INVOICE, name="inv.pdf"):
        with open(path, "rb") as f:
            return client.post("/upload", data={"invoice": (f, name)},
                                content_type="multipart/form-data")

    conn = db.connect(flask_app.config["DB_PATH"])
    before = conn.execute("SELECT COUNT(*) FROM gate_passes").fetchone()[0]
    conn.close()

    resp = upload(staff)
    where = resp.headers.get("Location", "")
    check("a staff upload goes straight to the print preview", "/print/" in where)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("and the pass is already issued",
          conn.execute("SELECT COUNT(*) FROM gate_passes").fetchone()[0] == before + 1)
    issued = conn.execute("SELECT * FROM gate_passes ORDER BY id DESC LIMIT 1").fetchone()
    check("it carries the next number in the run", issued["serial_no"].startswith("FZ-"))
    check("and names whoever uploaded it", issued["prepared_by"] == "Asha Nair")
    check("no draft is left behind", len(db.list_drafts(conn)) == 0)
    conn.close()

    check("the print preview really is theirs to open",
          staff.get(where).status_code == 200)

    # The Drafts page and its nav link are gone for them.
    check("Drafts is not reachable", staff.get("/drafts").status_code == 302)
    check("and not offered in the nav",
          b'href="/drafts"' not in staff.get("/register").data)

    # An admin, who reviews, gets the old behaviour.
    resp = upload(admin)
    check("an admin's upload stops at the review screen",
          "/review/" in resp.headers.get("Location", ""))
    conn = db.connect(flask_app.config["DB_PATH"])
    check("as a draft holding no number", len(db.list_drafts(conn)) == 1)
    conn.close()

    # A PDF that cannot be read has nowhere to go but the review screen: a pass
    # needs items, and a numbered pass with none is a permanent hole in the
    # register that cancelling cannot undo.
    conn = db.connect(flask_app.config["DB_PATH"])
    before = conn.execute("SELECT COUNT(*) FROM gate_passes").fetchone()[0]
    conn.close()
    resp = upload(staff, SCANNED_INVOICE, "scan.pdf")
    check("an unreadable upload lands on review instead",
          "/review/" in resp.headers.get("Location", ""))
    conn = db.connect(flask_app.config["DB_PATH"])
    check("and burns no number",
          conn.execute("SELECT COUNT(*) FROM gate_passes").fetchone()[0] == before)
    conn.close()
    # That review screen must be usable without the Drafts permission, or the
    # operator is sent somewhere they cannot open to finish the job.
    draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    check("the review screen opens without the Drafts permission",
          staff.get(f"/review/{draft_id}").status_code == 200)


@needs_fixtures
def test_serial_run_is_shared_between_people(tmpdir):
    """Different people at the office draw from one continuous run — the whole
    point of replacing the handwritten book."""
    flask_app, ravi = logged_in_app(tmpdir, "shared_storage",
                                     users=(("ravi", "Ravi Kumar"), ("asha", "Asha Nair")))
    # Both go through the draft step here, so the test exercises one flow. What
    # it is checking is that the RUN is shared, not how a pass is raised —
    # mixed flows are covered by test_instant_issue_without_the_draft_step.
    conn = db.connect(flask_app.config["DB_PATH"])
    asha_user = db.get_user_by_username(conn, "asha")
    db.set_user_permissions(conn, asha_user["id"], {"can_review_drafts": True})
    conn.close()
    asha = flask_app.test_client()
    sign_in(asha, "asha")

    serials = []
    for client, who in ((ravi, "ravi"), (asha, "asha"), (ravi, "ravi"), (asha, "asha")):
        with open(SAMPLE_INVOICE, "rb") as f:
            resp = client.post("/upload", data={"invoice": (f, "inv.pdf")},
                                content_type="multipart/form-data")
        draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
        resp = client.post(f"/review/{draft_id}", data=MultiDict([
            ("supplier_name", "Golden Touch Exports"), ("customer_name", "FANZART LLP"),
            ("invoice_no", f"INV-{who}"), ("invoice_date", "03-08-2026"), ("vehicle_no", ""),
            ("item_name", "MICRON MODREN OAK"), ("quantity", "1"), ("cartons", ""),
            ("action", "issue"),
        ]))
        gate_pass_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
        conn = db.connect(flask_app.config["DB_PATH"])
        serials.append(db.get_gate_pass(conn, gate_pass_id))
        conn.close()

    numbers = [gp["serial_seq"] for gp in serials]
    check("two people alternating still produce one unbroken run",
          numbers == [1, 2, 3, 4])
    check("no serial number is handed out twice",
          len({gp["serial_no"] for gp in serials}) == 4)
    check("each pass records who prepared it",
          [gp["prepared_by"] for gp in serials] ==
          ["Ravi Kumar", "Asha Nair", "Ravi Kumar", "Asha Nair"])
    check("the register shows both preparers",
          b"Ravi Kumar" in ravi.get("/register").data
          and b"Asha Nair" in ravi.get("/register").data)


def test_concurrent_issuing_keeps_the_run_unbroken(tmpdir):
    """Several people pressing Issue at the same moment must not collide. The
    serial is allocated inside the same immediate transaction as the insert, so
    SQLite serialises them."""
    import threading

    db_path = Path(tmpdir) / "concurrent.db"
    db.connect(db_path).close()

    results = []
    errors = []
    barrier = threading.Barrier(8)

    def issue(n):
        conn = db.connect(db_path)
        try:
            barrier.wait(timeout=10)
            gp = db.create_gate_pass(
                conn, None, "Golden Touch Exports", "FANZART LLP", f"INV-{n}",
                "03-08-2026", "", sample_items(), prepared_by=f"Person {n}")
            results.append(gp["serial_seq"])
        except Exception as exc:
            errors.append(exc)
        finally:
            conn.close()

    threads = [threading.Thread(target=issue, args=(n,)) for n in range(8)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=30)

    check("no thread failed while issuing at the same moment", not errors)
    check("every simultaneous issue got a number", len(results) == 8)
    check("no two simultaneous issues share a number", len(set(results)) == len(results))
    check("the run has no gaps", sorted(results) == list(range(1, len(results) + 1)))


@needs_fixtures
def test_prepared_by_is_recorded_and_printed(tmpdir):
    flask_app, client = logged_in_app(tmpdir, "prepared_storage")

    with open(SAMPLE_INVOICE, "rb") as f:
        resp = client.post("/upload", data={"invoice": (f, "inv.pdf")},
                            content_type="multipart/form-data")
    draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    resp = client.post(f"/review/{draft_id}", data=MultiDict([
        ("supplier_name", "Golden Touch Exports"), ("customer_name", "FANZART LLP"),
        ("invoice_no", "FR 262702171"), ("invoice_date", "03-08-2026"), ("vehicle_no", ""),
        ("item_name", "MICRON MODREN OAK"), ("quantity", "1"), ("cartons", ""),
        ("action", "issue"),
    ]))
    gate_pass_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    printed = client.get(f"/print/{gate_pass_id}").data
    check("the printed pass names who prepared it", b"Ravi Kumar" in printed)
    check("the printed pass carries a Prepared by label", b"Prepared by" in printed)
    check("the printed pass still carries Authorised by", b"Authorised by" in printed)
    check("both labels appear on each of the two passes",
          printed.count(b"Prepared by") == 2 and printed.count(b"Authorised by") == 2)
    # The box is drawn to its full budgeted height, so every row on both halves
    # has an empty cartons cell — the one item's included, since cartons are
    # written by hand.
    rows_drawn = db.print_row_count(1)
    check("the cartons column prints empty when nothing was typed",
          printed.count(b'<td class="col-ctn"></td>') == rows_drawn * 2)

    # With cartons blank there is no box count to total, so the label is dropped
    # rather than printed with nothing after it.
    conn = db.connect(flask_app.config["DB_PATH"])
    db.update_settings(conn, show_total_qty="1", show_total_cartons="1")
    conn.close()
    with_totals = client.get(f"/print/{gate_pass_id}").data
    check("the totals row still shows the quantity",
          re.search(rb'<td class="col-qty"><strong>\d+</strong>', with_totals))
    check("no empty Boxes total is printed when cartons were left blank",
          b"Boxes :" not in with_totals)

    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.get_gate_pass(conn, gate_pass_id)
    check("the preparer is stored on the record", gp["prepared_by"] == "Ravi Kumar")
    check("the preparer's account is linked", gp["prepared_by_user_id"] is not None)
    check("who prepared it cannot be edited afterwards", raises(
        sqlite3.IntegrityError, conn.execute,
        "UPDATE gate_passes SET prepared_by = 'Someone Else' WHERE id = ?", (gate_pass_id,)))
    check("the audit log records who issued it",
          "Ravi Kumar" in conn.execute(
              "SELECT details FROM audit_log WHERE gate_pass_id = ?",
              (gate_pass_id,)).fetchone()["details"])
    conn.close()


def test_migration_of_an_existing_database(tmpdir):
    """The database is the real book and is never rebuilt, so an older one has
    to gain the new columns in place, keeping every row it already held."""
    old = Path(tmpdir) / "old.db"
    conn = sqlite3.connect(old)
    conn.row_factory = sqlite3.Row
    conn.executescript("""
        CREATE TABLE gate_passes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            serial_no TEXT NOT NULL UNIQUE,
            serial_scope TEXT NOT NULL,
            serial_seq INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'issued',
            supplier_name TEXT NOT NULL DEFAULT '',
            customer_name TEXT NOT NULL DEFAULT '',
            invoice_no TEXT NOT NULL DEFAULT '',
            invoice_date TEXT NOT NULL DEFAULT '',
            vehicle_no TEXT NOT NULL DEFAULT '',
            invoice_pdf_path TEXT,
            issued_at TEXT NOT NULL DEFAULT (datetime('now')),
            cancelled_at TEXT,
            cancel_reason TEXT
        );
        INSERT INTO gate_passes (serial_no, serial_scope, serial_seq, invoice_no)
        VALUES ('GP/26-27/00001', '26-27', 1, 'OLD-INVOICE');
        CREATE TABLE serial_counters (scope TEXT PRIMARY KEY, next_seq INTEGER NOT NULL DEFAULT 1);
        INSERT INTO serial_counters (scope, next_seq) VALUES ('26-27', 2);
        -- The first release's users table: is_active, no status and no roles.
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE COLLATE NOCASE,
            display_name TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT (datetime('now'))
        );
        INSERT INTO users (username, display_name, password_hash, is_active)
        VALUES ('ravi', 'Ravi Kumar', 'x', 1), ('gone', 'Left Already', 'x', 0);
    """)
    conn.commit()
    conn.close()

    upgraded = db.connect(old)
    row = upgraded.execute("SELECT * FROM gate_passes WHERE serial_seq = 1").fetchone()
    check("the existing pass survives the upgrade", row["invoice_no"] == "OLD-INVOICE")
    check("the old pass gains an empty prepared_by", row["prepared_by"] == "")
    check("existing accounts survive the upgrade", db.count_users(upgraded) == 2)

    ravi = db.get_user_by_username(upgraded, "ravi")
    gone = db.get_user_by_username(upgraded, "gone")
    check("someone who could already sign in stays approved",
          ravi["status"] == db.APPROVED)
    check("someone already switched off comes across as disabled",
          gone["status"] == db.DISABLED)
    check("the earliest account becomes the first admin", ravi["is_admin"] == 1)
    check("the upgrade does not hand admin to everyone", gone["is_admin"] == 0)
    check("an upgraded install has an admin who can approve people",
          db.count_admins(upgraded) == 1)

    gp = db.create_gate_pass(upgraded, None, "S", "C", "NEW", "03-08-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")
    check("numbering on an upgraded book starts the new run at FZ-00001",
          gp["serial_no"] == "FZ-00001")
    check("the old pass keeps the number it was issued under",
          db.get_gate_pass_by_serial(upgraded, "GP/26-27/00001") is not None)
    check("a new pass on an upgraded database records its preparer",
          gp["prepared_by"] == "Ravi Kumar")
    upgraded.close()


def test_quantity_detection():
    check("a plain count is a quantity", invoice_parser._is_quantity("2"))
    check("a two digit count is a quantity", invoice_parser._is_quantity("26"))
    check("an 8-digit HSN code is not a quantity", not invoice_parser._is_quantity("84145120"))
    check("a price is not a quantity", not invoice_parser._is_quantity("9147.46"))
    check("a line total is not a quantity", not invoice_parser._is_quantity("18294.92"))
    check("text is not a quantity", not invoice_parser._is_quantity("Ceiling Fan"))
    check("blank is not a quantity", not invoice_parser._is_quantity(""))
    check("zero is not a quantity", not invoice_parser._is_quantity("0"))


@needs_fixtures
def test_scanned_pdf_degrades():
    """The real MANGALDEEP scan: a signed paper copy with no text layer at all.

    Nothing can be read from it, so the parser must say so once and plainly
    rather than emit five separate 'could not find' notes or invent values.
    """
    result = invoice_parser.parse_invoice(SCANNED_INVOICE)
    check("a scanned PDF yields no guessed supplier", result["supplier_name"] == "")
    check("a scanned PDF yields no guessed invoice number", result["invoice_no"] == "")
    check("a scanned PDF yields no guessed customer", result["customer_name"] == "")
    check("a scanned PDF yields no guessed date", result["invoice_date"] == "")
    check("a scanned PDF yields no items", result["items"] == [])
    check("a scanned PDF gives exactly one note", len(result["notes"]) == 1)
    check("the note explains it is a scan", "scan" in result["notes"][0].lower())


@needs_fixtures
def test_unreadable_invoice_is_still_typeable(tmpdir):
    """Degrading to manual entry has to be usable: an unreadable invoice must
    still present a blank item row rather than an empty table."""
    flask_app, client = logged_in_app(tmpdir, "scan_storage")

    with open(SCANNED_INVOICE, "rb") as f:
        resp = client.post("/upload", data={"invoice": (f, "mangaldeep_scan.pdf")},
                            content_type="multipart/form-data")
    draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    review = client.get(f"/review/{draft_id}")
    check("a scan still reaches the review screen", review.status_code == 200)
    check("the operator is told the PDF is a scan", b"scan or photo" in review.data)
    # Rendered rows carry a value attribute; the "+ Add item" JS template does not.
    check("an unreadable invoice still offers a row to type into",
          review.data.count(b'name="item_name" value=') == 1)

    issue = MultiDict([
        ("supplier_name", "Golden Touch Exports"), ("customer_name", "MANGALDEEP"),
        ("invoice_no", "FR 262702169"), ("invoice_date", "03-08-2026"),
        ("vehicle_no", ""),
        ("item_name", "VIENNA (52) (DC) (ABS) MATTE WHITE"), ("quantity", "2"), ("cartons", "1"),
        ("action", "issue"),
    ])
    resp = client.post(f"/review/{draft_id}", data=issue)
    check("a hand-typed gate pass issues normally", "/print/" in resp.headers.get("Location", ""))


@needs_fixtures
def test_full_app_flow(tmpdir):
    flask_app, client = logged_in_app(tmpdir, "app_storage")

    check("/ redirects to /upload",
          client.get("/").headers.get("Location", "").endswith("/upload"))
    check("/upload GET is 200", client.get("/upload").status_code == 200)

    resp = client.post("/upload", data={}, follow_redirects=True)
    check("uploading with no file redirects back to upload",
          resp.request.path == "/upload")

    drafts_before = len(db.list_drafts(db.connect(flask_app.config["DB_PATH"])))
    resp = client.post(
        "/upload",
        data={"invoice": (io.BytesIO(b"not a pdf"), "invoice.txt")},
        content_type="multipart/form-data",
    )
    drafts_after = len(db.list_drafts(db.connect(flask_app.config["DB_PATH"])))
    check("non-PDF upload redirects back to upload",
          resp.headers.get("Location", "").endswith("/upload"))
    check("non-PDF upload does not create a draft", drafts_after == drafts_before)

    with open(SAMPLE_INVOICE, "rb") as f:
        resp = client.post(
            "/upload",
            data={"invoice": (f, "golden_touch_sample.pdf")},
            content_type="multipart/form-data",
        )
    check("uploading a real invoice redirects to review", "/review/" in resp.headers.get("Location", ""))
    draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    review_resp = client.get(f"/review/{draft_id}")
    check("review page loads", review_resp.status_code == 200)
    check("review page shows parsed supplier", b"Golden Touch Exports" in review_resp.data)
    check("review page shows parsed item", b"MICRON MODREN OAK" in review_resp.data)
    check("the cartons box arrives empty for the operator to fill in",
          b'name="cartons" value=""' in review_resp.data)

    save_data = MultiDict([
        ("supplier_name", "Golden Touch Exports"),
        ("customer_name", "FANZART LLP"),
        ("invoice_no", "FR 262702171"),
        ("invoice_date", "03-08-2026"),
        ("vehicle_no", "KA 01 AB 1234"),
        ("item_name", "MICRON MODREN OAK"),
        ("quantity", "1"),
        ("cartons", ""),
        ("action", "save"),
    ])
    client.post(f"/review/{draft_id}", data=save_data)
    check("draft still exists after Save",
          db.get_draft(db.connect(flask_app.config["DB_PATH"]), draft_id) is not None)
    check("no gate pass exists yet",
          len(db.list_gate_passes(db.connect(flask_app.config["DB_PATH"]))) == 0)

    issue_no_item = MultiDict(save_data)
    issue_no_item["item_name"] = ""
    issue_no_item["action"] = "issue"
    resp = client.post(f"/review/{draft_id}", data=issue_no_item, follow_redirects=True)
    check("issuing with no named item is rejected", resp.request.path == f"/review/{draft_id}")
    check("rejected issue leaves the draft in place",
          db.get_draft(db.connect(flask_app.config["DB_PATH"]), draft_id) is not None)

    # Cartons stay blank on purpose — they are written by hand on the paper copy.
    issue_data = MultiDict(save_data)
    issue_data["action"] = "issue"
    resp = client.post(f"/review/{draft_id}", data=issue_data)
    check("issuing with everything filled redirects to print",
          "/print/" in resp.headers.get("Location", ""))
    gate_pass_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])

    conn = db.connect(flask_app.config["DB_PATH"])
    check("draft is gone after issuing", db.get_draft(conn, draft_id) is None)
    gp = db.get_gate_pass(conn, gate_pass_id)
    check("gate pass exists after issuing", gp is not None)
    check("issued gate pass has a serial number", gp["serial_no"].startswith("FZ-"))
    conn.close()

    print_resp = client.get(f"/print/{gate_pass_id}")
    check("print page loads", print_resp.status_code == 200)
    check("print page shows the serial number", gp["serial_no"].encode() in print_resp.data)
    check("print page shows the item", b"MICRON MODREN OAK" in print_resp.data)

    # The box runs to its full budgeted height on both halves, so the column
    # dividers reach the bottom exactly as on the printed form.
    # Counted inside <tbody> only: the totals row is a tfoot and carries a
    # col-sl cell of its own, which would inflate this without meaning a row
    # was drawn in the box.
    body = "".join(re.findall(r"<tbody>.*?</tbody>",
                              print_resp.data.decode(), re.S))
    printed_rows = body.count('<td class="col-sl">')
    check("the box is drawn full height on each of the two passes",
          printed_rows == db.print_row_count(1) * 2)
    # ...but only the one real item is ruled off horizontally.
    check("exactly the populated rows are tagged for a horizontal rule",
          print_resp.data.count(b'<tr class="item-row">') == 1 * 2)
    check("the rest are blank filler rows carrying no rule",
          body.count('<td class="col-sl"></td>') == (db.print_row_count(1) - 1) * 2)
    # print_row_count is now the height BUDGET, not the number of rows drawn:
    # it is what divides ITEM_BODY_MM into --row-h, so a short pass still gets
    # roomy rows and a full one tightens enough to fit.
    check("a short pass is budgeted the comfortable row count",
          db.print_row_count(1) == db.PRINT_MIN_ROWS)
    check("a fuller pass is budgeted a row per item instead",
          db.print_row_count(24) == 24)
    check("no pass is ever budgeted more rows than the form holds",
          db.print_row_count(99) == db.ITEMS_PER_PAGE)
    check("a 1-item pass still gets roomy rows despite drawing one row",
          db.print_row_height_mm(db.print_row_count(1))
          > db.print_row_height_mm(db.print_row_count(26)))
    check("print page shows two passes on an A4 sheet", b"cut-line" in print_resp.data)
    check("totals are off by default", b"Total qty" not in print_resp.data)
    check("vehicle is off by default", b"Vehicle :" not in print_resp.data)

    check("print page for unknown id is 404", client.get("/print/999999").status_code == 404)
    # A draft, unlike a pass, legitimately stops existing — issuing deletes it.
    # So an unknown draft id is a redirect with an explanation rather than a
    # 404: the Back button after issuing lands here, and a dead end there looks
    # like the pass was lost. See test_back_button_after_issuing.
    gone = client.get("/review/999999")
    check("an unknown draft redirects rather than 404ing", gone.status_code == 302)
    check("and explains itself",
          b"issued or discarded" in client.get("/review/999999",
                                                follow_redirects=True).data)

    register_resp = client.get("/register")
    check("register lists the issued gate pass", gp["serial_no"].encode() in register_resp.data)

    cancel_no_reason = client.post(f"/gate-passes/{gate_pass_id}/cancel", data={}, follow_redirects=True)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("cancel without a reason is rejected",
          db.get_gate_pass(conn, gate_pass_id)["status"] == "issued")
    conn.close()

    client.post(f"/gate-passes/{gate_pass_id}/cancel", data={"reason": "wrong invoice"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("cancel with a reason succeeds",
          db.get_gate_pass(conn, gate_pass_id)["status"] == "cancelled")
    conn.close()

    resp = client.post(f"/gate-passes/{gate_pass_id}/cancel", data={"reason": "again"}, follow_redirects=True)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("cancelling an already-cancelled pass is a no-op",
          db.get_gate_pass(conn, gate_pass_id)["cancel_reason"] == "wrong invoice")
    conn.close()

    settings_resp = client.get("/settings")
    check("settings page loads", settings_resp.status_code == 200)
    client.post("/settings", data={"paper_mode": "a5", "company_name": "fanzart"})
    conn = db.connect(flask_app.config["DB_PATH"])
    check("settings POST persists paper_mode", db.get_settings(conn)["paper_mode"] == "a5")
    check("settings POST turns off an unchecked checkbox",
          db.get_settings(conn)["show_total_qty"] == "0")
    conn.close()

    another_draft_id = None
    with open(SAMPLE_INVOICE, "rb") as f:
        resp = client.post(
            "/upload",
            data={"invoice": (f, "golden_touch_sample.pdf")},
            content_type="multipart/form-data",
        )
    another_draft_id = int(resp.headers["Location"].rstrip("/").rsplit("/", 1)[-1])
    drafts_resp = client.get("/drafts")
    check("drafts page lists the new draft", b"FR 262702171" in drafts_resp.data)

    client.post(f"/drafts/{another_draft_id}/delete")
    conn = db.connect(flask_app.config["DB_PATH"])
    check("discarding a draft removes it", db.get_draft(conn, another_draft_id) is None)
    conn.close()

    gp2 = db.create_gate_pass(db.connect(flask_app.config["DB_PATH"]), None, "S", "C", "I",
                               "01-01-2026", "", sample_items())
    check("discarded second draft did not burn a number", gp2["serial_no"] == "FZ-00002")


def test_hosting_hardening(tmpdir):
    """The three ways in that a hosted instance has to be closed against:
    reading files it should not serve, being driven from another site, and
    handing out a debugger."""
    flask_app, admin = logged_in_app(tmpdir, "hardening",
                                      users=(("ops", "Ravi Kumar"), ("asha", "Asha Nair")))
    conn = db.connect(flask_app.config["DB_PATH"])
    gp = db.create_gate_pass(conn, None, "S", "C", "INV", "01-01-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")
    conn.close()

    # The invoice route is rooted at the invoices folder. Rooted at storage/ it
    # served gate_pass.db (every password hash) and secret_key (enough to forge
    # an admin session) to any signed-in account, including one with no
    # permissions at all.
    staff = flask_app.test_client()
    sign_in(staff, "asha")
    for target in ("gate_pass.db", "secret_key", "../gate_pass.db", "..%2fsecret_key"):
        check(f"/invoices/{target} is not served",
              staff.get(f"/invoices/{target}").status_code != 200)

    # A real invoice still downloads — the fix must not break the feature.
    invoices = Path(flask_app.config["INVOICES_DIR"])
    invoices.mkdir(parents=True, exist_ok=True)
    (invoices / "sample.pdf").write_bytes(b"%PDF-1.4 test")
    got = admin.get("/invoices/invoices/sample.pdf")
    check("an uploaded invoice is still downloadable", got.status_code == 200)
    check("and it is the file that was stored", got.get_data() == b"%PDF-1.4 test")

    # Cross-site writes. A browser always sends Origin on a cross-site POST, so
    # a mismatch is proof of forgery; without this a page on any other site
    # could cancel a pass in a signed-in operator's name.
    forged = admin.post(f"/gate-passes/{gp['id']}/cancel",
                        data={"reason": "forged"},
                        headers={"Origin": "https://evil.example"})
    check("a cross-site POST is refused", forged.status_code == 403)
    conn = db.connect(flask_app.config["DB_PATH"])
    check("and the pass it targeted is untouched",
          db.get_gate_pass(conn, gp["id"])["status"] == "issued")
    conn.close()

    allowed = admin.post(f"/gate-passes/{gp['id']}/cancel",
                         data={"reason": "genuine"},
                         headers={"Origin": "http://localhost"})
    check("a same-origin POST still works", allowed.status_code in (200, 302))
    conn = db.connect(flask_app.config["DB_PATH"])
    check("and it did what it was asked to",
          db.get_gate_pass(conn, gp["id"])["status"] == "cancelled")
    conn.close()

    check("the session cookie is not sent cross-site",
          flask_app.config["SESSION_COOKIE_SAMESITE"] == "Lax")
    check("the session cookie is not readable from JavaScript",
          flask_app.config["SESSION_COOKIE_HTTPONLY"] is True)

    headers = admin.get("/register").headers
    check("responses set X-Content-Type-Options",
          headers.get("X-Content-Type-Options") == "nosniff")
    check("responses refuse to be framed", headers.get("X-Frame-Options") == "DENY")

    # The Werkzeug debugger is a Python console on an error page. It must never
    # be switched on by merely starting the file.
    source = (ROOT / "app.py").read_text()
    # A shell script that is not executable fails at the moment somebody needs
    # it, with a permission error that reads like a security problem rather
    # than a missing mode bit. git tracks the bit; it just has to be set.
    for script in sorted((ROOT / "deploy").glob("*.sh")):
        check(f"deploy/{script.name} is executable",
              os.access(script, os.X_OK))

    # `listen 443 ssl` means 0.0.0.0:443, which FAILS to bind if anything holds
    # 443 on any address — tailscaled does, for Funnel. nginx then rejects the
    # whole config and silently keeps the previous one, so the site stays on
    # plain HTTP and nothing looks wrong. The addresses are filled in at
    # install time instead.
    ssl_conf = (ROOT / "deploy" / "nginx-gate-pass-ssl.conf").read_text()
    check("the TLS site does not try to bind every address on 443",
          not re.search(r"^\s*listen\s+(\[::\]:)?443\b", ssl_conf, re.M))
    check("its listen addresses are substituted at install time",
          "__HTTPS_LISTEN__" in ssl_conf)
    check("and Tailscale traffic is not redirected to a port nginx cannot answer",
          "$needs_https" in ssl_conf and "100\\." in ssl_conf)

    enable = (ROOT / "deploy" / "enable-https.sh").read_text()
    # Under `set -e` a failing curl aborts inside the assignment, so `die`
    # never runs and the script ends in silence looking like success.
    check("the HTTPS script does not use set -e, which hides its own failures",
          "set -uo pipefail" in enable and "set -euo" not in enable)
    # Grepping for the bare string matched the explanatory comment in the unit
    # file and reported "already done" where it had never been set.
    check("it matches an uncommented GATE_PASS_HTTPS, not the comment about it",
          '^[[:space:]]*Environment="GATE_PASS_HTTPS=1"' in enable)
    # `systemctl reload` returning 0 does not mean the config is live.
    check("it confirms nginx really bound the socket",
          "did not bind" in enable)

    # The TLS template is filled in by plain string replacement, so a
    # placeholder NAMED in a comment gets substituted there too — which dropped
    # a bare `listen` into the middle of a comment block and made nginx reject
    # the whole file with "listen directive is not allowed here".
    check("the listen placeholder appears exactly once",
          ssl_conf.count("__HTTPS_LISTEN__") == 1)

    # Render it the way enable-https.sh does and check the result is a valid
    # shape: braces balanced, and no directive outside a block.
    rendered = (ssl_conf.replace("__APP_DIR__", "/opt/gate-pass")
                        .replace("__CERT_DIR__", "/opt/gate-pass/deploy/certs")
                        .replace("__HTTPS_LISTEN__",
                                 "    listen 127.0.0.1:443 ssl;\n"
                                 "    listen 192.168.1.45:443 ssl;"))
    depth, stray = 0, []
    for number, raw in enumerate(rendered.split("\n"), 1):
        line = raw.split("#")[0].strip()
        if not line:
            continue
        if re.match(r"^(listen|proxy_pass|ssl_certificate)\b", line) and depth == 0:
            stray.append(f"line {number}: {line}")
        depth += line.count("{") - line.count("}")
    check("the rendered config has no directive outside a block", not stray)
    check("and its braces balance", depth == 0)
    # Three: plain HTTP on 80, TLS on the LAN, and the loopback front door
    # that tailscaled proxies into so the Tailscale entrance gets the same
    # static handling, body limit and export timeouts as the LAN one.
    check("it defines a plain, a TLS and a Tailscale-facing server",
          len(re.findall(r"^server \{", rendered, re.M)) == 3)
    check("the Tailscale front door is loopback only, never on a real address",
          "listen 127.0.0.1:8091;" in rendered
          and not re.search(r"listen (?!127\.0\.0\.1)[0-9.]*:?8091", rendered))
    check("with a TLS listener per address given",
          len(re.findall(r"listen [0-9.]+:443 ssl;", rendered)) == 2)

    # A CA certificate is public and must be fetchable — every office machine
    # needs a copy before the padlock means anything. Locking the directory to
    # 700 made the one file people need impossible to get without root.
    make_cert = (ROOT / "deploy" / "make-cert.sh").read_text()
    check("the certificate directory can be read",
          'chmod 755 "$CERT_DIR"' in make_cert)
    check("but the private keys cannot",
          'chmod 600 "$CERT_DIR/fanzart-ca.key" "$CERT_DIR/server.key"' in make_cert)
    check("nginx hands the CA out so nobody carries a USB stick",
          "location = /fanzart-ca.pem" in ssl_conf)
    # make-cert.sh only runs when there is NO certificate, so a permissions fix
    # made there never reaches a machine that already has one — which is every
    # machine that matters. enable-https.sh sets them on every run.
    enable_https = (ROOT / "deploy" / "enable-https.sh").read_text()
    # Re-running after HTTPS is on must work. nginx holds 443 on exactly the
    # addresses it is about to reuse, and a graceful reload keeps old workers
    # and their sockets alive for a moment — counting either as a conflict made
    # the script work once and never again.
    check("it does not mistake its own listeners for a conflict",
          'grep -v \'users:(("nginx"\'' in enable_https)

    # The redirect must NOT catch the CA. It is the file that makes https
    # trustworthy, so requiring a trusted https connection to fetch it means a
    # machine that has not got it yet is told to come back over a connection it
    # cannot verify. An exact-match location outranks `location /`, which is
    # where the redirect now lives — it used to sit at server level, where it
    # applied to everything.
    port80 = ssl_conf[ssl_conf.index("listen 80 default_server"):
                       ssl_conf.index("__HTTPS_LISTEN__")]
    check("the CA has its own exact-match location on port 80",
          "location = /fanzart-ca.pem" in port80)
    check("and the redirect is inside location /, not at server level",
          port80.index("location = /fanzart-ca.pem") < port80.index("return 301"))
    check("so the redirect cannot reach it",
          re.search(r"location / \{[^}]*if \(\$needs_https\)", port80, re.S) is not None)

    # The CA must be reachable when HTTPS is OFF, which is precisely when a
    # machine needs it. Serving it only from the TLS site is a chicken and egg.
    http_conf = (ROOT / "deploy" / "nginx-gate-pass.conf").read_text()
    check("the plain-HTTP site serves the CA too",
          "location = /fanzart-ca.pem" in http_conf)

    check("permissions are set on every run, not only at issue",
          'chmod 755 "$CERT_DIR"' in enable_https
          and 'chmod 644 "$CERT_DIR/fanzart-ca.pem"' in enable_https)

    # The part the usual advice gets wrong: on Linux, Chrome and Brave do NOT
    # read /usr/local/share/ca-certificates. They keep their own NSS database,
    # so update-ca-certificates alone leaves them still warning — which looks
    # exactly like the certificate not working.
    trust = (ROOT / "deploy" / "trust-ca.sh").read_text()
    check("the client installer covers the system store",
          "update-ca-certificates" in trust)
    check("and the browser's own NSS store",
          "certutil" in trust and ".pki/nssdb" in trust)
    # Firefox keeps yet another NSS store, one per profile. certutil can write
    # to it directly, so there is no need to walk anybody through a dialog —
    # and profiles live in three different places depending on how Firefox was
    # installed (deb, snap, flatpak).
    check("it installs into Firefox profiles too",
          ".mozilla/firefox" in trust and "cert9.db" in trust)
    check("including snap and flatpak Firefox",
          "snap/firefox" in trust and ".var/app/org.mozilla.firefox" in trust)
    check("and falls back to instructions when there is no profile",
          "View Certificates" in trust)

    # Windows and the antivirus problem. Kaspersky intercepts HTTPS and checks
    # against its OWN authority list, so installing the CA in the Windows root
    # store does not make it accept the server — the browser trusts Kaspersky,
    # and Kaspersky refuses. Anyone following the instructions has to be told
    # that, or they will keep reinstalling a certificate that was never the
    # problem.
    for name in ("trust-ca-windows.ps1", "OFFICE-MACHINES.md"):
        doc = (ROOT / "deploy" / name).read_text()
        check(f"{name} covers the Windows root store", "Trusted Root" in doc)
        check(f"{name} explains why antivirus HTTPS scanning defeats it",
              "Kaspersky" in doc)
        check(f"{name} offers plain HTTP rather than clicking through warnings",
              "http://" in doc)
    check("and verifies afterwards rather than assuming",
          "verifies — the padlock will be clean" in trust)

    install = (ROOT / "deploy" / "install.sh").read_text()
    check("and so does the one install.sh writes inline",
          "location = /fanzart-ca.pem" in install)
    # With HTTPS on, port 80 is a 301 by design. Demanding 200 there declared a
    # perfectly healthy server broken, moments after enable-https.sh had
    # finished setting it up.
    check("the health check asks over whatever scheme is actually serving",
          'BASE="https://127.0.0.1"' in install and 'BASE="http://127.0.0.1"' in install)
    check("and it checks the CA is still fetchable without TLS",
          "fanzart-ca.pem returned" in install)
    # install.sh rewrites the nginx site every run. It used to write the
    # plain-HTTP one unconditionally, so every update silently dropped the
    # server back off TLS.
    check("updating does not silently turn HTTPS off",
          "443 ssl" in install and "HTTPS_WAS_ON" in install)
    # A neighbouring site losing its symlink is silent: nothing errors, the
    # port just stops answering. It happened once, to a site published to the
    # internet through Tailscale Funnel.
    check("it records which other nginx sites were enabled",
          "BEFORE=" in install)
    check("and puts back any that vanish",
          "RESTORED" in install)
    check("it only ever disables the stock placeholder",
          install.count("rm -f /etc/nginx/sites-enabled/") == 1
          and "sites-enabled/default" in install)

    check("the debugger is off unless explicitly asked for",
          "debug=True" not in source and "GATE_PASS_DEBUG" in source)


def test_a_fetched_copy_can_be_opened_without_touching_the_live_book(tmpdir):
    """GATE_PASS_STORAGE opens a different book, and only when asked.

    It exists so a copy pulled off the server can be read with the real app.
    The danger of such a switch is that it fires when nobody meant it to, so
    the default must be untouched when it is unset — production never sets it.
    """
    elsewhere = Path(tmpdir) / "fetched"
    os.environ["GATE_PASS_STORAGE"] = str(elsewhere)
    try:
        moved = create_app()
        check("the override moves the whole storage directory",
              Path(moved.config["STORAGE_DIR"]) == elsewhere)
        check("and the database with it",
              moved.config["DB_PATH"] == str(elsewhere / "gate_pass.db"))
        check("uploaded invoices go beside it, not into the live folder",
              Path(moved.config["INVOICES_DIR"]) == elsewhere / "invoices")
        # An explicit argument still wins, or the test suite itself would be
        # reading whatever this variable happened to be set to.
        explicit = Path(tmpdir) / "explicit"
        check("an explicit storage_dir still beats the environment",
              Path(create_app(storage_dir=explicit).config["STORAGE_DIR"]) == explicit)
    finally:
        del os.environ["GATE_PASS_STORAGE"]

    # Relative to the code, not to a directory called "gate-pass" — a clone
    # checked out under any other name is still a valid clone.
    fallback = Path(create_app().config["STORAGE_DIR"])
    check("unset, it falls back to the storage beside the code",
          fallback == Path(app_module.BASE_DIR) / "storage")
    unit = (ROOT / "deploy" / "gate-pass.service").read_text()
    check("and the live service never sets it", "GATE_PASS_STORAGE" not in unit)


def test_emptying_the_register_for_a_fresh_start(tmpdir):
    """manage_reset.py wipes the book and puts the protection back.

    It deliberately breaks the guarantee the whole app is built on — that a
    gate pass can never be deleted — so the parts that matter are that it will
    not fire by accident, that it keeps what nobody wants to set up again, and
    that the delete protection is genuinely working afterwards rather than
    merely present by name.
    """
    import manage_reset

    storage = Path(tmpdir) / "reset"
    (storage / "invoices").mkdir(parents=True)
    (storage / "invoices" / "old.pdf").write_bytes(b"%PDF-1.4 old upload")
    path = storage / "gate_pass.db"
    conn = db.connect(path)
    make_user(conn, "ravi", "Ravi Kumar")
    db.upsert_carton_mappings(conn, [("MICRON MODREN OAK", 1)])
    db.update_settings(conn, company_name="fanzart", show_total_qty="1")
    for i in range(3):
        db.create_gate_pass(conn, None, "S", "C", f"I{i}", "01-01-2026", "",
                            sample_items(), prepared_by="Ravi Kumar")
    db.create_draft(conn, supplier_name="S", customer_name="C", invoice_no="D",
                    invoice_date="01-01-2026", items=sample_items())
    # Numbering jumped forward, to prove the floor is cleared too — otherwise
    # the "fresh" book would start at 500, not 1.
    db.start_new_run(conn, "FZ", 500, started_by="Ravi Kumar")
    db.create_gate_pass(conn, None, "S", "C", "J", "01-01-2026", "",
                        sample_items(), prepared_by="Ravi Kumar")
    check("the book has passes before the reset", len(db.list_gate_passes(conn)) == 4)
    conn.close()

    # Without --wipe it must change nothing at all.
    manage_reset.main(["--db", str(path)])
    conn = db.connect(path)
    check("a dry run changes nothing", len(db.list_gate_passes(conn)) == 4)
    conn.close()

    manage_reset.main(["--db", str(path), "--wipe", "--yes"])
    conn = db.connect(path)

    check("every gate pass is gone", len(db.list_gate_passes(conn)) == 0)
    check("every draft is gone", len(db.list_drafts(conn)) == 0)
    check("the uploaded invoices are gone",
          not list((storage / "invoices").glob("*.pdf")))
    check("numbering restarts at 00001, ignoring the earlier jump to 500",
          db.next_serial_preview(conn) == "FZ-00001")
    check("and the first pass issued really is FZ-00001",
          db.create_gate_pass(conn, None, "S", "C", "NEW", "01-01-2026", "",
                              sample_items(), prepared_by="Ravi Kumar")["serial_no"]
          == "FZ-00001")

    # Kept, or somebody spends the morning of go-live setting it all up again.
    check("accounts survive", db.get_user_by_username(conn, "ravi") is not None)
    check("the carton master list survives", db.carton_mapping_count(conn) == 1)
    check("settings survive", db.get_settings(conn)["company_name"] == "fanzart")

    # The reset is itself recorded, in the log it just emptied.
    reset_rows = conn.execute(
        "SELECT details FROM audit_log WHERE action = 'register_reset'").fetchall()
    check("the reset writes its own audit line", len(reset_rows) == 1)
    check("naming how much it removed", "gate passes" in reset_rows[0]["details"])

    # And the protection is BACK — tested by trying, not by looking up a name.
    # db.connect() alone does not restore it: schema.sql only re-runs when the
    # user_version differs, and after a wipe it does not.
    try:
        with db.writing(conn):
            conn.execute("DELETE FROM gate_passes")
        check("deleting a gate pass is refused again", False)
    except sqlite3.IntegrityError:
        check("deleting a gate pass is refused again", True)
    triggers = {r["name"] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type = 'trigger'")}
    check("every trigger in the schema is back",
          set(manage_reset.TRIGGERS) <= triggers)
    conn.close()


def test_the_deploy_scripts_have_no_dangling_heredocs():
    """Every here-document in the deploy scripts is closed.

    This shipped broken. An edit to install.sh matched the delimiter on the
    `cat <<UNIT` line ITSELF rather than on the terminator, deleted only that
    first line, and left thirty lines of systemd unit sitting in the script as
    shell commands. It ran on the server as:

        install.sh: line 102: [Unit]: command not found

    `bash -n` does not catch it — `[Unit]` is a syntactically valid command —
    so a syntax check passed both before and after the damage. This looks at
    the structure instead.
    """
    for script in sorted((ROOT / "deploy").glob("*.sh")) + [ROOT / "deploy" / "install.sh"]:
        text = script.read_text()
        lines = text.splitlines()
        for number, line in enumerate(lines, 1):
            for delim in re.findall(r"<<-?\s*[\'\"]?([A-Za-z_][A-Za-z0-9_]*)[\'\"]?", line):
                closed = any(later.strip() == delim for later in lines[number:])
                check(f"{script.name}:{number} closes its <<{delim}", closed)

        # Nothing that belongs inside a unit file may sit at the top level of a
        # shell script, which is exactly what the orphaned body looked like.
        stray = [f"{script.name}:{n}: {l}" for n, l in enumerate(lines, 1)
                 if re.match(r"^(\[(Unit|Service|Install)\]|ExecStart=|WantedBy=)", l)]
        check(f"{script.name} has no loose systemd directives", not stray)
        if stray:
            for one in stray[:3]:
                print(f"    {one}")

    # And the scripts still parse.
    import subprocess
    for script in sorted((ROOT / "deploy").glob("*.sh")):
        result = subprocess.run(["bash", "-n", str(script)], capture_output=True, text=True)
        check(f"{script.name} parses", result.returncode == 0)


def test_the_deployment_config_holds_together():
    """The bits of the server setup that break quietly when they drift.

    Every one of these has either bitten already or would bite silently: a
    wrong path in a unit file is only read during an incident, and a body limit
    below the app's own rejects an upload with no explanation.
    """
    unit = (ROOT / "deploy" / "gate-pass.service").read_text()
    install = (ROOT / "deploy" / "install.sh").read_text()

    # One copy of the unit. There were two, and they had drifted on the user
    # and the working directory.
    check("install.sh renders the unit from the template, not a second copy",
          "deploy/gate-pass.service" in install
          and "cat > /etc/systemd/system/gate-pass.service <<UNIT" not in install)
    check("the template is parameterised, so it cannot hardcode a stale path",
          "__APP_DIR__" in unit and "__APP_USER__" in unit)
    # Directives only: the comment above them explains the stale path that used
    # to be here, and naming it is the point of that comment.
    directives = "\n".join(l for l in unit.splitlines() if l and not l.startswith("#"))
    check("no directive names a developer's home directory",
          "/home/" not in directives)

    # A power cut must leave the office with a working book.
    check("it restarts on failure", "Restart=always" in unit)
    check("with a delay, so a crash loop does not spin",
          re.search(r"^RestartSec=[1-9]", unit, re.M))
    check("and it is enabled at install, so it survives a reboot",
          re.search(r"systemctl enable[^\n]*gate-pass", install))

    # Optional env file: the leading "-" is what keeps a missing file from
    # stopping the service.
    check("an environment file is referenced", "EnvironmentFile=" in unit)
    check("and optionally, so an absent one is not a failure to boot",
          "EnvironmentFile=-" in unit)
    check("a template for it is committed", (ROOT / "deploy" / "env.example").exists())
    ignored = (ROOT / ".gitignore").read_text()
    check("but real env files never reach this public repository",
          ".env" in ignored and "!deploy/env.example" in ignored)

    # The service may write to the book and nothing else.
    check("the service can only write to storage/", "ReadWritePaths=__APP_DIR__/storage" in unit)
    check("and the rest of the machine is read-only to it", "ProtectSystem=strict" in unit)

    # nginx must not reject an upload the app would have accepted.
    app_limit = int(re.search(r"MAX_CONTENT_LENGTH\"\]\s*=\s*(\d+)", (ROOT / "app.py").read_text()).group(1))
    for conf in ("nginx-gate-pass.conf", "nginx-gate-pass-ssl.conf"):
        text = (ROOT / "deploy" / conf).read_text()
        for size in re.findall(r"client_max_body_size\s+(\d+)M", text):
            check(f"{conf} accepts at least what the app does ({size}M >= {app_limit}MB)",
                  int(size) >= app_limit)

    # Backups, and something to stop their log eating the disk.
    check("a nightly backup is scheduled", "/etc/cron.d/gate-pass-backup" in install)
    check("and its log is rotated", "/etc/logrotate.d/gate-pass" in install)
    check("there is a preflight check to run before handing it over",
          (ROOT / "deploy" / "preflight.sh").exists())


def test_a_draft_can_be_removed_from_the_row_it_is_on(tmpdir):
    """The ✕ on the Drafts list, and what it does to the rest of the batch.

    A blocked pair used to be a dead end on this screen: the message said to
    delete one of them and the only way to do it was to open a draft and
    discard it from the next screen. The ✕ does it from the row, and the
    blocking has to lift on its own afterwards — a warning that stays up after
    the thing it warns about is gone teaches people to ignore warnings.
    """
    flask_app, client = logged_in_app(tmpdir, "rowdelete")
    conn = db.connect(flask_app.config["DB_PATH"])

    def draft(number, source):
        return db.create_draft(conn, supplier_name="S", customer_name="C",
                                invoice_no=number, invoice_date="01-01-2026",
                                invoice_pdf_path=f"invoices/20260101120000_{source}",
                                items=sample_items())

    twin_a, twin_b = draft("RT 262700325", "5.pdf"), draft("RT-262700325", "9.pdf")
    clean = draft("FR 262702181", "3.pdf")

    page = client.get("/drafts").get_data(as_text=True)
    # data-draft-id, not the class name: the class also appears in the script
    # that binds the handler, so counting it says four for three rows.
    check("every row carries a delete button", page.count("data-draft-id") == 3)
    check("and it knows which draft it is on", f'data-draft-id="{twin_a}"' in page)
    check("the pair blocks to begin with", len(db.blocking_duplicates(
        conn, [twin_a, twin_b, clean])) == 2)

    # The page's own fetch. Says so, so the route answers in JSON rather than
    # sending it somewhere.
    resp = client.post(f"/drafts/{twin_a}/delete",
                       headers={"Origin": "http://localhost",
                                "X-Requested-With": "fetch"})
    check("the delete answers in JSON", resp.status_code == 200)
    body = resp.get_json()
    check("and says it worked", body["ok"] is True)
    check("and carries the words for the toast", body["message"] == "Draft removed")
    check("the draft is gone", db.get_draft(conn, twin_a) is None)
    check("its items went with it",
          conn.execute("SELECT COUNT(*) FROM draft_items WHERE draft_id = ?",
                       (twin_a,)).fetchone()[0] == 0)

    # The whole point: the survivor is no longer blocked BY THE SERVER, not by
    # the page deciding it looks fine now.
    check("nothing is blocked any more", body["blocked"] == [])
    check("and it counts what is left", body["remaining"] == 2)
    check("and how many can go", body["ready"] == 2)
    check("the rule agrees", db.blocking_duplicates(conn, [twin_b, clean]) == {})

    page = client.get("/drafts").get_data(as_text=True)
    check("the surviving row is no longer marked blocked",
          "&#9888; Duplicate<" not in page)
    check("and the batch is issuable again", "cannot be issued" not in page)

    # The review screen's Discard button is an ordinary form post and still has
    # to be sent somewhere. Returning it JSON would leave it on a blank page.
    resp = client.post(f"/drafts/{twin_b}/delete",
                       headers={"Origin": "http://localhost"})
    check("a plain form post still redirects", resp.status_code == 302)
    check("and to the drafts list", "/drafts" in resp.headers["Location"])

    # Two people clearing the same pair at once. The row is gone either way,
    # which is what was asked for.
    resp = client.post(f"/drafts/{twin_b}/delete",
                       headers={"Origin": "http://localhost",
                                "X-Requested-With": "fetch"})
    check("deleting an already-deleted draft is not an error",
          resp.status_code == 200 and resp.get_json()["already_gone"] is True)
    check("but a form post for a missing draft still 404s",
          client.post(f"/drafts/{twin_b}/delete",
                      headers={"Origin": "http://localhost"}).status_code == 404)

    check("no gate pass number was spent by any of it",
          len(db.list_gate_passes(conn)) == 0)
    conn.close()


def test_a_batch_is_numbered_in_document_order(tmpdir):
    """Gate pass numbers follow the document numbers, not the upload order.

    A batch used to take its order from however the drafts were listed, which
    is the order somebody fed the scanner. Fifty invoices off a pile came out
    of the register shuffled, and finding the pass for an invoice meant
    searching for it rather than reading down the column.
    """
    # The key on its own first, because the sort is only as good as this.
    check("letters at the front, digits at the end",
          db.document_sort_key("FR 262702176") == ("FR", 0, 262702176))
    check("spacing and punctuation do not change it",
          db.document_sort_key("FR-262702176") == db.document_sort_key("FR 262702176"))
    check("a transfer memo sorts with its own prefix, not under T",
          db.document_sort_key("TO NO: FR 262702176")
          == db.document_sort_key("FR 262702176"))
    # The whole reason the numbers do not order themselves: as text, the longer
    # number is the smaller one.
    check("the digits compare as a number, not as text",
          sorted(["FR 262702558", "FR 26270256"], key=db.document_sort_key)
          == ["FR 26270256", "FR 262702558"])
    check("a number with no digits goes last within its prefix",
          db.document_sort_key("FR")[1] == 1 and db.document_sort_key("FR 1")[1] == 0)

    flask_app, client = logged_in_app(tmpdir, "sortbatch")
    conn = db.connect(flask_app.config["DB_PATH"])

    # Deliberately created in the wrong order, and spanning three prefixes.
    scrambled = ["SN 262700900", "FR 262702558", "BI 262700158",
                 "FR 26270256", "SN 262700100", "FR 262702176"]
    ids = [db.create_draft(conn, supplier_name="S", customer_name="C",
                            invoice_no=number, invoice_date="01-01-2026",
                            invoice_pdf_path=f"invoices/2026_{n}.pdf",
                            items=sample_items())
           for n, number in enumerate(scrambled)]

    issued, skipped = db.create_gate_passes_batch(conn, ids, prepared_by="Ravi Kumar")
    check("all of them were issued", len(issued) == 6 and not skipped)

    order = [(p["serial_no"], p["invoice_no"]) for p in
             sorted(issued, key=lambda p: p["serial_seq"])]
    check("the numbers were handed out in document order",
          [n for _, n in order] == ["BI 262700158", "FR 26270256", "FR 262702176",
                                     "FR 262702558", "SN 262700100", "SN 262700900"])
    check("and the serials are an unbroken block",
          [s for s, _ in order] == [f"FZ-0000{i}" for i in range(1, 7)])

    # Two drafts carrying one number sort to the same key, so something else has
    # to decide — otherwise the same batch issued twice could number them the
    # other way round.
    twin_a = db.create_draft(conn, supplier_name="S", customer_name="C",
                              invoice_no="RT 262700325", invoice_date="01-01-2026",
                              invoice_pdf_path="invoices/a.pdf", items=sample_items())
    twin_b = db.create_draft(conn, supplier_name="S", customer_name="C",
                              invoice_no="RT-262700325", invoice_date="01-01-2026",
                              invoice_pdf_path="invoices/b.pdf", items=sample_items())
    check("a tied pair still has a defined order",
          db.document_sort_key("RT 262700325") == db.document_sort_key("RT-262700325"))
    issued, _ = db.create_gate_passes_batch(conn, [twin_b, twin_a],
                                             prepared_by="Ravi Kumar")
    first = min(issued, key=lambda p: p["serial_seq"])
    check("and it is the older draft that goes first",
          first["invoice_no"] == "RT 262700325")

    conn.close()


def test_the_drafts_list_previews_the_numbering_order(tmpdir):
    """What the Drafts screen shows is the order the numbers will be given out.

    These are two separate pieces of code — one lists rows for a page, the
    other allocates serial numbers — and if they order drafts differently then
    the screen is quietly lying about which draft becomes which gate pass. So
    the test is not "the list is sorted", it is "the list and the batch agree",
    compared item for item.
    """
    flask_app, client = logged_in_app(tmpdir, "previeworder")
    conn = db.connect(flask_app.config["DB_PATH"])

    scrambled = ["SN 262700900", "FR 262702558", "BI 262700158", "FR 26270256",
                 "TO NO: FR 262702176", "MU 262700307", "FR 262702176"]
    for n, number in enumerate(scrambled):
        db.create_draft(conn, supplier_name="S", customer_name="C",
                         invoice_no=number, invoice_date="01-01-2026",
                         invoice_pdf_path=f"invoices/2026_{n}.pdf",
                         items=sample_items())

    listed = db.list_drafts(conn)
    check("the list is in document order",
          [d["invoice_no"] for d in listed]
          # The transfer memo and the invoice normalize to one number, so the
          # tie-break decides — and it is the older draft, which is the memo.
          == ["BI 262700158", "FR 26270256", "TO NO: FR 262702176",
              "FR 262702176", "FR 262702558", "MU 262700307",
              "SN 262700900"])

    # The page renders them in that order too, not just the function.
    page = client.get("/drafts").get_data(as_text=True)
    # data-draft-id, not the checkbox: a blocked duplicate has no checkbox, and
    # this batch deliberately contains one.
    positions = [page.index(f'data-draft-id="{d["id"]}"') for d in listed]
    check("and the page renders them in that order",
          positions == sorted(positions))

    # The claim being made: issuing them hands out numbers in the SAME order.
    issued, skipped = db.create_gate_passes_batch(
        conn, [d["id"] for d in listed], prepared_by="Ravi Kumar")
    check("all of them were issued", len(issued) == len(scrambled) and not skipped)
    check("the numbering follows the list exactly",
          [p["invoice_no"] for p in sorted(issued, key=lambda p: p["serial_seq"])]
          == [d["invoice_no"] for d in listed])

    # And it still agrees when the ids are handed over shuffled, which is what
    # a form post of ticked checkboxes actually is.
    for number in ("RT 262700325", "RT-262700325", "BI 262700001"):
        db.create_draft(conn, supplier_name="S", customer_name="C",
                         invoice_no=number, invoice_date="01-01-2026",
                         invoice_pdf_path=f"invoices/{number}.pdf",
                         items=sample_items())
    listed = db.list_drafts(conn)
    issued, _ = db.create_gate_passes_batch(
        conn, list(reversed([d["id"] for d in listed])), prepared_by="Ravi Kumar")
    check("order comes from the documents, not from how they were submitted",
          [p["invoice_no"] for p in sorted(issued, key=lambda p: p["serial_seq"])]
          == [d["invoice_no"] for d in listed])

    conn.close()


def test_every_test_is_actually_run():
    """Every test_* in this file must be called by main().

    There is no collector here — main() lists the tests by hand — so a test
    that is written and never added to that list reports nothing, passes
    nothing, and looks exactly like a test that works. Six had accumulated,
    including one guarding a print layout that had already regressed once.
    """
    source = Path(__file__).read_text()
    defined = set(re.findall(r"^def (test_\w+)\(", source, re.M))
    body = source[source.index("def main():"):]
    called = set(re.findall(r"^\s+(test_\w+)\(", body, re.M))
    missing = sorted(defined - called - {"test_every_test_is_actually_run"})
    check("every test in this file is called by main()", not missing)
    if missing:
        for name in missing:
            print(f"    never run: {name}")


def main():
    tmpdir = tempfile.mkdtemp(prefix="gate-pass-test-")
    try:
        test_every_test_is_actually_run()
        test_the_deploy_scripts_have_no_dangling_heredocs()
        test_the_deployment_config_holds_together()
        test_emptying_the_register_for_a_fresh_start(tmpdir)
        test_a_fetched_copy_can_be_opened_without_touching_the_live_book(tmpdir)
        test_serial_format()
        test_triggers(tmpdir)
        test_serial_allocation(tmpdir)
        test_draft_never_burns_a_number(tmpdir)
        test_draft_lifecycle(tmpdir)
        test_cancel(tmpdir)
        test_settings(tmpdir)
        test_invoice_parser()
        test_multi_item_and_wrapped_names(tmpdir)
        test_two_page_invoice()
        test_service_line_invoice()
        test_name_column_is_found_by_its_heading()
        test_bi_series_invoice()
        test_stock_transfer_memo(tmpdir)
        test_transfer_rows_are_read_by_shape_not_by_column_edges()
        test_real_transfer_memos()
        test_a_gate_pass_uploaded_as_an_invoice(tmpdir)
        test_cartons_come_from_the_master_list(tmpdir)
        test_parsing_a_batch(tmpdir)
        test_remarks_print_on_more_than_one_line(tmpdir)
        test_duplicate_document_numbers_are_flagged_not_blocked(tmpdir)
        test_two_drafts_with_one_number_cannot_both_be_issued(tmpdir)
        test_a_draft_can_be_removed_from_the_row_it_is_on(tmpdir)
        test_a_batch_is_numbered_in_document_order(tmpdir)
        test_the_drafts_list_previews_the_numbering_order(tmpdir)
        test_totals_are_shown_live_and_derived_on_save(tmpdir)
        test_the_register_shows_what_has_reached_paper(tmpdir)
        test_the_printed_totals_are_bolder_than_the_rows(tmpdir)
        test_correcting_an_issued_pass(tmpdir)
        test_a_cancelled_pass_says_so_on_paper(tmpdir)
        test_printed_sheet_has_no_rule_between_the_copies(tmpdir)
        test_vehicle_prints_only_when_there_is_one(tmpdir)
        test_totals_sit_between_the_items_and_the_signatures(tmpdir)
        test_the_two_totals_are_independent(tmpdir)
        test_upgrading_keeps_a_totals_preference(tmpdir)
        test_a_two_page_pass_signs_off_only_at_the_end(tmpdir)
        test_carton_lookup_rules(tmpdir)
        test_sequence_derives_from_the_book(tmpdir)
        test_batch_issue_keeps_the_run_unbroken(tmpdir)
        test_batch_rolls_back_as_one(tmpdir)
        test_two_batches_at_once(tmpdir)
        test_batch_upload_over_http(tmpdir)
        test_long_invoice_is_one_pass_over_several_pages(tmpdir)
        test_review_screen_explains_multiple_pages(tmpdir)
        test_print_layout_rules(tmpdir)
        test_timestamps_are_local_not_utc(tmpdir)
        test_upgrade_converts_old_utc_timestamps(tmpdir)
        test_date_filters(tmpdir)
        test_quick_ranges_and_export_over_http(tmpdir)
        test_register_page_is_just_search_and_filter(tmpdir)
        test_reports_page(tmpdir)
        test_xlsx_export(tmpdir)
        test_uploaded_invoices_are_deleted_once_used(tmpdir)
        test_per_file_upload_endpoint(tmpdir)
        test_upload_page_does_not_block_its_own_submit(tmpdir)
        test_connection_settings(tmpdir)
        test_indexes_are_used(tmpdir)
        test_register_is_capped(tmpdir)
        test_failed_write_rolls_back(tmpdir)
        test_setting_the_numbering(tmpdir)
        test_users_and_login(tmpdir)
        test_accounts_are_created_by_admins_only(tmpdir)
        test_permissions_are_stored_per_feature(tmpdir)
        test_permissions_gate_each_feature(tmpdir)
        test_managing_permissions_from_the_people_page(tmpdir)
        test_permissions_backfill_on_upgrade(tmpdir)
        test_batch_printing(tmpdir)
        test_role_based_access(tmpdir)
        test_admin_powers_and_guardrails(tmpdir)
        test_remarks_are_typed_and_printed(tmpdir)
        test_item_table_keyboard_navigation(tmpdir)
        test_login_cannot_be_brute_forced(tmpdir)
        test_back_button_after_issuing(tmpdir)
        test_stylesheets_are_cache_busted(tmpdir)
        test_printed_issue_date_is_the_pass_date(tmpdir)
        test_the_parse_is_the_record(tmpdir)
        test_instant_issue_without_the_draft_step(tmpdir)
        test_serial_run_is_shared_between_people(tmpdir)
        test_concurrent_issuing_keeps_the_run_unbroken(tmpdir)
        test_prepared_by_is_recorded_and_printed(tmpdir)
        test_migration_of_an_existing_database(tmpdir)
        test_quantity_detection()
        test_scanned_pdf_degrades()
        test_unreadable_invoice_is_still_typeable(tmpdir)
        test_hosting_hardening(tmpdir)
        test_full_app_flow(tmpdir)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    passed = sum(1 for _, ok in RESULTS if ok)
    total = len(RESULTS)
    print(f"\n{passed}/{total} checks passed")
    if SKIPPED:
        print(f"{len(SKIPPED)} test(s) skipped — the sample invoices are not in the "
              f"repository (see tests/sample_invoices/README.md):")
        for name in SKIPPED:
            print(f"  - {name}")
    if passed != total:
        sys.exit(1)


if __name__ == "__main__":
    main()
