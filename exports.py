"""Turning a report's rows into a downloadable file.

Two formats, one source of rows (`db.export_rows`), so a CSV and an Excel export
of the same report always contain exactly the same data.

CSV is written with the standard library. Excel is written with openpyxl —
pandas is deliberately not used: it would add tens of megabytes and would hand
the actual .xlsx writing to openpyxl anyway.
"""

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Roughly how wide each column should be, so the sheet is readable without the
# auditor having to drag every border. Keyed by header text.
COLUMN_WIDTHS = {
    "Gate Pass No": 16, "Issued On": 19, "Status": 10, "Prepared By": 18,
    "From": 24, "Customer": 24, "Invoice No": 16, "Invoice Date": 13,
    "Vehicle No": 14, "Items": 7, "Total Quantity": 14, "Total Cartons": 14,
    "Cancelled On": 19, "Cancelled By": 18, "Cancel Reason": 40,
    "Sl No": 7, "Item": 40, "Quantity": 10, "No. of Cartons": 14,
}
DEFAULT_WIDTH = 16

HEADER_FILL = PatternFill("solid", fgColor="17171A")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Columns an auditor will want to total in Excel, so they go in as numbers.
# Everything else stays text on purpose: an invoice number like "0012" must not
# be turned into 12, and quantities are stored as free text in the database.
NUMERIC_COLUMNS = {
    "Sl No", "Items", "Quantity", "No. of Cartons", "Total Quantity", "Total Cartons",
}


def csv_bytes(rows):
    """Yields the report as CSV, a chunk at a time.

    Streamed rather than assembled: an export covering a whole year should not
    have to be built in memory before the download starts. The BOM is there so
    Excel opens UTF-8 correctly when someone double-clicks the file.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    yield "﻿".encode()
    for row in rows:
        writer.writerow(row)
        yield buffer.getvalue().encode()
        buffer.seek(0)
        buffer.truncate(0)


def xlsx_bytes(rows, title="Gate Passes"):
    """The report as a real .xlsx workbook.

    Uses openpyxl's write-only mode, which streams rows to the file instead of
    holding a cell object for each one — the difference between comfortable and
    unusable on a large audit export. A workbook cannot be streamed to the
    browser the way CSV can (the zip's index is written last), so this returns
    the finished bytes.
    """
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=title[:31])
    sheet.freeze_panes = "A2"

    rows = iter(rows)
    header = next(rows, None)
    if header is None:
        header = []

    for index, name in enumerate(header, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = \
            COLUMN_WIDTHS.get(name, DEFAULT_WIDTH)

    header_cells = []
    for name in header:
        cell = WriteOnlyCell(sheet, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)

    numeric = [name in NUMERIC_COLUMNS for name in header]
    for row in rows:
        sheet.append([_cell_value(v, numeric[i] if i < len(numeric) else False)
                      for i, v in enumerate(row)])

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _cell_value(value, as_number=False):
    if isinstance(value, (int, float)):
        return value
    if value is None or value == "":
        return ""
    text = str(value)
    if as_number:
        # Quantities are free text in the database, so only convert what really
        # is a number; anything else (a range, a note) is left exactly as typed.
        try:
            number = float(text)
        except ValueError:
            return text
        return int(number) if number.is_integer() else number
    return text


def filename(basename, detailed, fmt, date_from=None, date_to=None):
    """A filename an auditor can file away without renaming."""
    span = f"{date_from or 'start'}_to_{date_to or datetime.now().date().isoformat()}"
    kind = "detailed" if detailed else "summary"
    return f"{basename}_{kind}_{span}.{fmt}"


CONTENT_TYPES = {
    # No charset here: Flask appends its own to a text/* mimetype, and spelling
    # it out produced "text/csv; charset=utf-8; charset=utf-8".
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
